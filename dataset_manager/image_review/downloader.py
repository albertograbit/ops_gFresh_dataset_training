"""
Módulo para descarga selectiva de imágenes basada en archivo Excel
Permite descargar imágenes para revisión manual desde AWS S3 según configuración en Excel
"""

import pandas as pd
import os
import logging
import random
import shutil
import glob
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import yaml
from botocore.exceptions import ClientError, NoCredentialsError
import requests
from urllib.parse import urlparse
from utils.excel_validations import ensure_dropdown_validations
from utils.report_excel_manager import ReportExcelManager, ExcelLockedError


class ImageReviewDownloader:
    """
    Descargador selectivo de imágenes para revisión manual
    """
    
    def __init__(self, config_path: str = "config/settings.yaml", process_aware: bool = False):
        """
        Inicializar el descargador de imágenes
        
        Args:
            config_path: Ruta al archivo de configuración
            process_aware: Si usar configuración consciente de procesos
        """
        self.logger = logging.getLogger(__name__)
        self.process_aware = process_aware

        # Cargar variables de entorno del .env
        from dotenv import load_dotenv
        load_dotenv()

        # Cargar configuración
        if process_aware:
            from dataset_manager.process_aware_settings import ProcessAwareSettings
            process_settings = ProcessAwareSettings(config_path)
            full_config = process_settings.get_settings()
            self.config = full_config.get('image_review', {})
            process_info = process_settings.get_process_info()
            if process_info and 'process_dir' in process_info:
                process_dir = Path(process_info['process_dir'])
                self.carpeta_destino = str(process_dir / 'images')
            else:
                self.carpeta_destino = self.config.get('carpeta_destino', './output/images')
        else:
            with open(config_path, 'r', encoding='utf-8') as f:
                raw_cfg = yaml.safe_load(f)
            self.config = raw_cfg.get('image_review', {})
            self.carpeta_destino = self.config.get('carpeta_destino', './output/images')

        # Parámetros generales
        self.num_imagenes = self.config.get('num_imagenes_revision', 5)
        self.tipo_transacciones = self.config.get('tipo_transacciones', 'ambas')
        self.clear_output_folder = self.config.get('clear_output_folder', True)
        self.tipo_imagenes_bajar = self.config.get('tipo_imagenes_bajar', 'clase_y_similares')

        # Cliente de almacenamiento compartido
        try:
            from dataset_manager.storage.storage_client import get_storage_client
            self.s3_client, storage_env = get_storage_client(self.logger, validate=True)
            self.s3_bucket = storage_env.bucket
            self.s3_region = storage_env.region
            self.storage_type = storage_env.storage_type
            self._storage_env = storage_env
        except Exception as e:
            self.logger.warning(f"Fallo inicializando cliente de almacenamiento compartido: {e}. Se intentará inicialización específica mínima.")
            self.s3_client = None
            self.s3_bucket = os.getenv('S3_BUCKET', 'grabit-data')
            self.s3_region = os.getenv('REMOTE_STORAGE_REGION', 'eu-west-2')
            self.storage_type = 's3'

        self.download_timeout = self.config.get('download_timeout', 30)
        self.max_retries = self.config.get('max_retries', 3)
        self.image_extensions = self.config.get('image_extensions', ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'])
        self.dry_run = False

        # Tracking
        self.downloaded_images_log = []
        self.last_log_save_status = None
        self.last_log_save_error = None

        # Si no se obtuvo cliente, intentar inicialización directa legacy
        if self.s3_client is None:
            self._init_s3_client()
        os.makedirs(self.carpeta_destino, exist_ok=True)

        self.logger.info("ImageReviewDownloader inicializado:")
        self.logger.info(f"  - Imágenes por referencia: {self.num_imagenes}")
        self.logger.info(f"  - Tipo de transacciones: {self.tipo_transacciones}")
        self.logger.info(f"  - Carpeta destino: {self.carpeta_destino}")
        self.logger.info(f"  - S3 Bucket: {self.s3_bucket}")
        self.logger.info(f"  - S3 Region: {self.s3_region}")

    def _init_s3_client(self):
        """Legacy fallback initialisation if shared client creation failed."""
        try:
            from dataset_manager.storage.storage_client import get_storage_client
            self.s3_client, storage_env = get_storage_client(self.logger, validate=False)
            self.s3_bucket = storage_env.bucket
            self.s3_region = storage_env.region
            self.storage_type = storage_env.storage_type
        except Exception as e:
            self.logger.error(f"Error inicializando cliente de almacenamiento: {e}")
            self.s3_client = None

    def find_latest_excel(self) -> Optional[str]:
        """
        Encontrar el archivo Excel más reciente, priorizando el proceso activo
        
        Returns:
            Ruta al archivo Excel más reciente o None si no se encuentra
        """
        try:
            def _is_main_dataset_file(fname: str) -> bool:
                name = os.path.basename(fname)
                if not name.lower().endswith('.xlsx'):
                    return False
                if name.startswith('~$'):
                    return False
                low = name.lower()
                # Excluir ficheros de registros/parciales/backup creados por review_images
                if '_imagenes_descargadas' in low or 'backup' in low:
                    return False
                # Patrones válidos actuales y legacy
                if name.startswith('dataset_'):
                    return True
                if 'dataset_analysis_' in name:
                    return True
                return False

            # Si hay configuración consciente de procesos, buscar primero en el proceso activo
            if self.process_aware:
                from dataset_manager.process_aware_settings import ProcessAwareSettings
                process_settings = ProcessAwareSettings()
                process_info = process_settings.get_process_info()
                
                if process_info:
                    # Buscar archivo Excel en el directorio del proceso activo
                    process_reports_dir = process_settings.get_output_path('reports')
                    if os.path.exists(process_reports_dir):
                        all_files = os.listdir(process_reports_dir)
                        excel_files = [f for f in all_files if _is_main_dataset_file(f)]
                        if excel_files:
                            # Si hay múltiples archivos, tomar el más reciente
                            latest_file = max(excel_files, key=lambda f: os.path.getmtime(os.path.join(process_reports_dir, f)))
                            excel_path = os.path.join(process_reports_dir, latest_file)
                            self.logger.info(f"Archivo Excel del proceso activo encontrado: {excel_path}")
                            return excel_path
            
            # Fallback: buscar en directorios de procesos
            processes_pattern = "output/processes/**/reports/*.xlsx"
            all_excel_candidates = glob.glob(processes_pattern, recursive=True)
            excel_files = [f for f in all_excel_candidates if _is_main_dataset_file(f)]
            
            if excel_files:
                latest_file = max(excel_files, key=os.path.getmtime)
                self.logger.info(f"Archivo Excel más reciente encontrado: {latest_file}")
                return latest_file
            
            # Fallback final: buscar archivos legacy
            legacy_pattern = "output/reports/**/dataset_analysis_*.xlsx"
            legacy_files = glob.glob(legacy_pattern, recursive=True)
            
            if legacy_files:
                latest_file = max(legacy_files, key=os.path.getmtime)
                self.logger.info(f"Archivo Excel legacy encontrado: {latest_file}")
                return latest_file
            
            self.logger.warning("No se encontraron archivos Excel en ningún directorio")
            return None
            
        except Exception as e:
            self.logger.error(f"Error buscando archivo Excel: {e}")
            return None

    def load_references_data(self, excel_path: str) -> pd.DataFrame:
        """
        Cargar datos de referencias desde Excel
        
        Args:
            excel_path: Ruta al archivo Excel
            
        Returns:
            DataFrame con datos de referencias
        """
        try:
            # Leer la hoja de Referencias
            df = pd.read_excel(excel_path, sheet_name='References')
            self.logger.info(f"Cargadas {len(df)} referencias desde {excel_path}")
            
            # Filtrar solo las referencias marcadas para revisión
            df_review = df[df['revisar_imagenes'].astype(str).str.lower().isin(['si', 'sí', 'yes', 'y'])]
            self.logger.info(f"Referencias marcadas para revisión: {len(df_review)}")
            
            return df_review
            
        except Exception as e:
            self.logger.error(f"Error cargando datos de Excel: {e}")
            raise

    def load_all_references_data(self, excel_path: str) -> pd.DataFrame:
        """
        Cargar TODAS las referencias desde Excel (no solo las marcadas para revisión)
        
        Args:
            excel_path: Ruta al archivo Excel
            
        Returns:
            DataFrame con todas las referencias para cruzar códigos
        """
        try:
            # Leer la hoja completa de Referencias
            df = pd.read_excel(excel_path, sheet_name='References')
            self.logger.info(f"Cargadas {len(df)} referencias completas para cruzar códigos")
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error cargando tabla completa de referencias: {e}")
            raise

    def load_elasticsearch_data(self, excel_path: str) -> pd.DataFrame:
        """Cargar datos de Elasticsearch desde CSV externo elastic_data.csv."""
        try:
            # Intentar localizar elastic_data.csv relativo al Excel (misma carpeta o carpeta padre reports)
            excel_dir = Path(excel_path).parent
            candidates = [excel_dir / 'elastic_data.csv', excel_dir.parent / 'elastic_data.csv']
            for c in candidates:
                if c.exists():
                    df = pd.read_csv(c, sep=';', encoding='utf-8') if c.suffix.lower()=='.csv' else pd.read_csv(c)
                    self.logger.info(f"Cargados {len(df)} registros de Elasticsearch desde {c}")
                    # Asegurar columnas clave
                    needed = ['transaction_id','device_id','device_name','result_label_name','selected_reference_code','selected_reference_name','Ok','Ok_modelo','image_link']
                    missing = [col for col in needed if col not in df.columns]
                    if missing:
                        self.logger.warning(f"Columnas faltantes en CSV: {missing}")
                    if 'image_link' in df.columns:
                        df['has_image'] = (df['image_link'].notna() & (df['image_link']!='') & (df['image_link']!='null') & (df['image_link']!='None'))
                    else:
                        df['has_image'] = False
                    return df
            raise FileNotFoundError('No se encontró elastic_data.csv asociado al Excel. Ejecute primero la extracción.')
        except Exception as e:
            self.logger.error(f"Error cargando datos de Elasticsearch: {e}")
            raise

    def filter_transactions_by_type(self, df_elastic: pd.DataFrame) -> pd.DataFrame:
        """
        Filtrar transacciones según el tipo configurado
        
        Args:
            df_elastic: DataFrame con datos de Elasticsearch
            
        Returns:
            DataFrame filtrado
        """
        if self.tipo_transacciones == 'cliente':
            # Solo transacciones de cliente
            filtered_df = df_elastic[df_elastic['transaction_metric'] == 'GFRSUSERPROCESSCOM03']
        elif self.tipo_transacciones == 'manuales':
            # Solo transacciones manuales
            filtered_df = df_elastic[df_elastic['transaction_metric'] != 'GFRSUSERPROCESSCOM03']
        else:
            # Ambas
            filtered_df = df_elastic
        
        self.logger.info(f"Transacciones filtradas por tipo '{self.tipo_transacciones}': {len(filtered_df)}")
        return filtered_df

    def get_reference_transactions(self, df_elastic: pd.DataFrame, reference_name: str) -> pd.DataFrame:
        """
        Obtener transacciones para una referencia específica
        
        Args:
            df_elastic: DataFrame con datos de Elasticsearch
            reference_name: Nombre de la referencia
            
        Returns:
            DataFrame con transacciones de la referencia
        """
        ref_transactions = df_elastic[df_elastic['selected_reference_name'] == reference_name]
        
        # Filtrar por tipo de transacciones
        ref_transactions = self.filter_transactions_by_type(ref_transactions)
        
        self.logger.debug(f"Transacciones para {reference_name}: {len(ref_transactions)}")
        return ref_transactions

    def get_products_to_download(self, row: pd.Series) -> List[Dict]:
        """
        Obtener lista de productos para descargar según configuración
        
        Args:
            row: Fila del DataFrame de referencias
            
        Returns:
            Lista de diccionarios con información de productos
        """
        products = []
        
        # Producto principal (clase)
        # Si label_name no está disponible, usar reference_name (para referencias no entrenadas)
        if pd.notna(row.get('label_name')):
            # Usar label_name cuando esté disponible (referencia entrenada)
            main_product_name = row['label_name']
        else:
            # Usar reference_name cuando label_name no esté disponible (referencia no entrenada)
            main_product_name = row['reference_name'].lower()
            
        products.append({
            'name': main_product_name,
            'type': 'clase_principal',
            'result_index': None,
            'reference_name': row['reference_name'],
            'label_id': row.get('label_id'),
            'reference_code': row.get('reference_code')
        })
        
        # Si el tipo_imagenes_bajar de configuración es clase_y_similares, añadir los top3
        tipo_imagenes_bajar = self.config.get('tipo_imagenes_bajar', 'clase_y_similares')
        if tipo_imagenes_bajar.lower() == 'clase_y_similares':
            for rank in range(1, 4):
                result_index = row.get(f'top{rank}_result_index')
                result_label_name = row.get(f'top{rank}_result_label_name')
                
                if pd.notna(result_index) and pd.notna(result_label_name):
                    # Para clases similares, obtener los sub-productos específicos
                    sub_products = self.get_similar_sub_products(result_label_name, row['reference_name'])
                    
                    if sub_products:
                        # Agregar cada sub-producto encontrado
                        for sub_product in sub_products:
                            products.append({
                                'name': sub_product['label_name'],
                                'type': f'similar_top{rank}',
                                'result_index': result_index,
                                'reference_name': row['reference_name'],
                                'parent_similar_name': result_label_name,
                                'label_id': sub_product['label_id'],
                                'reference_code': sub_product['reference_code'],
                                'selected_reference_code': sub_product['selected_reference_code']
                            })
                    else:
                        # Si no se encuentran sub-productos, usar el método anterior como fallback
                        products.append({
                            'name': result_label_name,
                            'type': f'similar_top{rank}',
                            'result_index': result_index,
                            'reference_name': row['reference_name'],
                            'parent_similar_name': result_label_name
                        })
        
        return products

    def get_similar_sub_products(self, similar_product_name: str, main_reference_name: str) -> List[Dict]:
        """
        Obtener sub-productos específicos para una clase similar
        
        Args:
            similar_product_name: Nombre del producto similar (ej: "Manzana roja-amarilla")
            main_reference_name: Nombre de la referencia principal (ej: "GRANADAS")
            
        Returns:
            Lista de diccionarios con información de sub-productos específicos
        """
        try:
            # 1. Encontrar transacciones donde result_label_name coincida con el producto similar
            filtered_transactions = self.df_elastic[
                self.df_elastic['result_label_name'].str.lower() == similar_product_name.lower()
            ]
            
            # 2. Filtrar por Ok=True y Ok_modelo=True
            correct_transactions = filtered_transactions[
                (filtered_transactions['Ok'] == True) & 
                (filtered_transactions['Ok_modelo'] == True)
            ].copy()
            
            if len(correct_transactions) == 0:
                self.logger.debug(f"No hay transacciones correctas para {similar_product_name}")
                return []
            
            # 3. Extraer selected_reference_code únicos
            unique_codes = correct_transactions['selected_reference_code'].dropna().unique()
            
            if len(unique_codes) == 0:
                self.logger.debug(f"No hay códigos de referencia válidos para {similar_product_name}")
                return []
            
            # 4. Cruzar con la tabla de referencias para obtener label_id y label_name
            if not hasattr(self, 'df_references') or self.df_references is None:
                self.logger.warning("Tabla de referencias no disponible para cruzar códigos")
                return []
            
            sub_products = []
            for code in unique_codes:
                # Buscar en la tabla de referencias
                matching_refs = self.df_references[
                    self.df_references['reference_code'] == code
                ]
                
                for _, ref_row in matching_refs.iterrows():
                    sub_products.append({
                        'label_id': ref_row.get('label_id'),
                        'label_name': ref_row.get('label_name'),
                        'reference_code': ref_row.get('reference_code'),
                        'selected_reference_code': code
                    })
            
            self.logger.debug(f"Encontrados {len(sub_products)} sub-productos para {similar_product_name}: {[p['label_name'] for p in sub_products]}")
            return sub_products
            
        except Exception as e:
            self.logger.error(f"Error obteniendo sub-productos para {similar_product_name}: {e}")
            return []

    def select_random_transactions(self, df_transactions: pd.DataFrame, 
                                 product_info: Dict, num_images: int) -> List[Dict]:
        """
        Seleccionar transacciones aleatorias para un producto
        
        Args:
            df_transactions: DataFrame con transacciones
            product_info: Información del producto
            num_images: Número de imágenes a seleccionar
            
        Returns:
            Lista de transacciones seleccionadas
        """
        product_name = product_info['name']
        is_main_class = product_info['type'] == 'clase_principal'
        reference_name = product_info['reference_name']
        
        if is_main_class:
            # Para la clase principal: buscar transacciones donde el cliente seleccionó la referencia principal
            # Buscar por selected_reference_name que coincida con reference_name
            filtered_transactions = df_transactions[
                df_transactions['selected_reference_name'].str.upper() == reference_name.upper()
            ]
            
            # Para la clase principal no necesitamos filtrar por Ok/Ok_modelo porque ya fue seleccionado por el cliente
            transactions_with_images = filtered_transactions[
                filtered_transactions['has_image'] == True
            ].copy()
            
            self.logger.debug(f"Clase principal {reference_name}: {len(filtered_transactions)} transacciones filtradas, {len(transactions_with_images)} con imágenes")
        else:
            # Para clases similares: nueva lógica más específica
            if 'selected_reference_code' in product_info and product_info['selected_reference_code']:
                # Buscar transacciones específicas donde el cliente seleccionó exactamente este código de referencia
                filtered_transactions = df_transactions[
                    df_transactions['selected_reference_code'] == product_info['selected_reference_code']
                ]
                
                # Para sub-productos específicos, filtrar por Ok=True y Ok_modelo=True
                correct_transactions = filtered_transactions[
                    (filtered_transactions['Ok'] == True) & 
                    (filtered_transactions['Ok_modelo'] == True)
                ].copy()
                
                # Filtrar solo transacciones con imágenes
                transactions_with_images = correct_transactions[
                    correct_transactions['has_image'] == True
                ].copy()
                
                self.logger.debug(f"Sub-producto específico {product_name} (código: {product_info['selected_reference_code']}): {len(filtered_transactions)} filtradas, {len(correct_transactions)} correctas, {len(transactions_with_images)} con imágenes")
            else:
                # Lógica anterior como fallback
                filtered_transactions = df_transactions[
                    df_transactions['result_label_name'].str.lower() == product_name.lower()
                ]
                
                # Filtrar solo transacciones correctas (Ok=True y Ok_modelo=True) para clases similares
                correct_transactions = filtered_transactions[
                    (filtered_transactions['Ok'] == True) & 
                    (filtered_transactions['Ok_modelo'] == True)
                ].copy()
                
                # Filtrar solo transacciones con imágenes
                transactions_with_images = correct_transactions[
                    correct_transactions['has_image'] == True
                ].copy()
                
                self.logger.debug(f"Clase similar {product_name}: {len(filtered_transactions)} filtradas por result_label_name, {len(correct_transactions)} correctas, {len(transactions_with_images)} con imágenes")
        
        if len(transactions_with_images) == 0:
            # Mejor diagnóstico del problema
            self.logger.warning(f"❌ No hay transacciones con imágenes para {product_name}")
            self.logger.warning(f"   - Transacciones filtradas: {len(filtered_transactions) if 'filtered_transactions' in locals() else 0}")
            if not is_main_class and 'correct_transactions' in locals():
                self.logger.warning(f"   - Transacciones correctas (Ok=True & Ok_modelo=True): {len(correct_transactions)}")
            self.logger.warning(f"   - Transacciones con imágenes: {len(transactions_with_images)}")
            
            # Diagnóstico adicional: verificar si hay imágenes en general
            if 'filtered_transactions' in locals() and len(filtered_transactions) > 0:
                has_any_images = (filtered_transactions['has_image'] == True).sum()
                self.logger.warning(f"   - Total con has_image=True en transacciones filtradas: {has_any_images}")
                
                if has_any_images == 0:
                    # Revisar las columnas de imagen directamente
                    image_cols = [col for col in filtered_transactions.columns if any(keyword in col.lower() for keyword in ['image', 'url', 'link', 'path'])]
                    for col in image_cols[:3]:  # Solo primeras 3 columnas
                        non_empty = (filtered_transactions[col].notna() & (filtered_transactions[col] != '')).sum()
                        self.logger.warning(f"   - Columna '{col}': {non_empty} valores no vacíos")
            
            return []
        
        # Seleccionar aleatoriamente
        num_to_select = min(num_images, len(transactions_with_images))
        selected = transactions_with_images.sample(n=num_to_select, random_state=42)
        
        return selected.to_dict('records')

    def download_image_from_transaction(self, transaction: Dict, dest_folder: str, 
                                      reference_name: str = None, product_name: str = None, 
                                      product_type: str = None) -> bool:
        """
        Descargar imagen de una transacción desde S3
        
        Args:
            transaction: Diccionario con datos de la transacción
            dest_folder: Carpeta de destino
            reference_name: Nombre de la referencia
            product_name: Nombre del producto
            product_type: Tipo de producto (clase_principal, similar_top1, etc.)
            
        Returns:
            True si se descargó exitosamente, False en caso contrario
        """
        # Registrar el intento desde el inicio
        image_record = {
            'reference_name': reference_name or 'unknown',
            'product_name': product_name or 'unknown', 
            'product_type': product_type or 'unknown',
            'transaction_id': transaction.get('transaction_id'),
            'transaction_start_time': transaction.get('transaction_start_time'),
            'device_id': transaction.get('device_id'),
            'device_name': transaction.get('device_name'),
            'selected_reference_name': transaction.get('selected_reference_name'),
            'result_label_name': transaction.get('result_label_name'),
            'confidence': transaction.get('confidence'),
            'Ok': transaction.get('Ok'),
            'Ok_modelo': transaction.get('Ok_modelo'),
            's3_key': '',
            's3_url': '',
            'local_path': '',
            'download_status': 'failed',
            'error_reason': ''
        }
        
        try:
            if self.s3_client is None:
                image_record['error_reason'] = 'Cliente S3 no inicializado'
                self.downloaded_images_log.append(image_record)
                self.logger.error("Cliente S3 no inicializado")
                return False
                
            image_link = transaction.get('image_link')
            if not image_link or pd.isna(image_link):
                image_record['error_reason'] = 'No hay enlace de imagen en la transacción'
                self.downloaded_images_log.append(image_record)
                self.logger.warning(f"No hay enlace de imagen para transacción {transaction.get('transaction_id')}")
                return False
            
            # Limpiar la ruta de S3 (remover barras al inicio si las hay)
            s3_key = str(image_link).strip().lstrip('/')
            image_record['s3_key'] = s3_key
            image_record['s3_url'] = f"s3://{self.s3_bucket}/{s3_key}"
            
            # Buscar archivos de imagen en S3
            image_objects = self._list_s3_objects(s3_key)
            
            if not image_objects:
                image_record['error_reason'] = f'Imagen no encontrada en S3: {s3_key}'
                self.downloaded_images_log.append(image_record)
                self.logger.warning(f"Imagen no encontrada en S3: s3://{self.s3_bucket}/{s3_key}")
                return False
            
            # Usar la primera imagen encontrada
            s3_object_key = image_objects[0]
            image_record['s3_key'] = s3_object_key
            image_record['s3_url'] = f"s3://{self.s3_bucket}/{s3_object_key}"
            
            # Nombre del archivo de destino
            transaction_id = transaction.get('transaction_id', 'unknown')
            file_extension = os.path.splitext(s3_object_key)[1]
            if not file_extension:
                file_extension = '.jpg'  # Por defecto
                
            dest_filename = f"{transaction_id}{file_extension}"
            dest_path = os.path.join(dest_folder, dest_filename)
            image_record['local_path'] = dest_path
            
            if self.dry_run:
                self.logger.info(f"[DRY-RUN] Descargaría: s3://{self.s3_bucket}/{s3_object_key} -> {dest_path}")
                image_record['download_status'] = 'dry_run'
                image_record['error_reason'] = ''
                self.downloaded_images_log.append(image_record)
                return True
            else:
                # Descargar imagen desde S3
                download_success = self._download_from_s3(s3_object_key, dest_path)
                if download_success:
                    image_record['download_status'] = 'success'
                    image_record['error_reason'] = ''
                else:
                    image_record['error_reason'] = 'Error en descarga desde S3'
                
                self.downloaded_images_log.append(image_record)
                return download_success
            
        except Exception as e:
            image_record['error_reason'] = f'Excepción: {str(e)}'
            self.downloaded_images_log.append(image_record)
            self.logger.error(f"Error descargando imagen desde S3: {e}")
            return False

    def _list_s3_objects(self, s3_prefix: str) -> List[str]:
        """
        Listar objetos en S3 que coincidan con el prefijo y sean imágenes
        
        Args:
            s3_prefix: Prefijo para buscar objetos
            
        Returns:
            Lista de claves de objetos que son imágenes
        """
        try:
            image_objects = []
            
            # Si el prefijo ya es un archivo específico, verificar si existe
            if any(s3_prefix.lower().endswith(ext) for ext in self.image_extensions):
                try:
                    self.s3_client.head_object(Bucket=self.s3_bucket, Key=s3_prefix)
                    return [s3_prefix]
                except ClientError:
                    pass
            
            # Buscar objetos con el prefijo
            response = self.s3_client.list_objects_v2(
                Bucket=self.s3_bucket,
                Prefix=s3_prefix,
                MaxKeys=100
            )
            
            if 'Contents' in response:
                for obj in response['Contents']:
                    key = obj['Key']
                    # Verificar si es una imagen basándose en la extensión
                    if any(key.lower().endswith(ext) for ext in self.image_extensions):
                        image_objects.append(key)
            
            return image_objects
            
        except Exception as e:
            self.logger.error(f"Error listando objetos S3 con prefijo {s3_prefix}: {e}")
            return []

    def _download_from_s3(self, s3_key: str, dest_path: str) -> bool:
        """
        Descargar un archivo desde S3
        
        Args:
            s3_key: Clave del objeto en S3
            dest_path: Ruta de destino local
            
        Returns:
            True si se descargó exitosamente
        """
        try:
            # Crear directorio de destino si no existe
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            
            # Descargar archivo
            self.s3_client.download_file(self.s3_bucket, s3_key, dest_path)
            self.logger.debug(f"Imagen descargada: s3://{self.s3_bucket}/{s3_key} -> {dest_path}")
            return True
            
        except ClientError as e:
            error_code = e.response['Error']['Code']
            if error_code == 'NoSuchKey':
                self.logger.warning(f"Objeto no encontrado en S3: s3://{self.s3_bucket}/{s3_key}")
            else:
                self.logger.error(f"Error descargando desde S3: {e}")
            return False
        except Exception as e:
            self.logger.error(f"Error descargando archivo {s3_key}: {e}")
            return False

    def process_reference(self, row: pd.Series, df_elastic: pd.DataFrame) -> Dict:
        """
        Procesar una referencia para descarga de imágenes
        
        Args:
            row: Fila del DataFrame de referencias
            df_elastic: DataFrame con datos de Elasticsearch
            
        Returns:
            Diccionario con estadísticas del procesamiento
        """
        reference_name = row['reference_name']
        self.logger.info(f"Procesando referencia: {reference_name}")
        
        # Crear carpeta para la referencia dentro de 'productos/'
        productos_folder = os.path.join(self.carpeta_destino, "productos")
        os.makedirs(productos_folder, exist_ok=True)
        
        ref_folder = os.path.join(productos_folder, reference_name)
        os.makedirs(ref_folder, exist_ok=True)
        
        # Usar todas las transacciones de Elasticsearch (filtradas por tipo)
        # No limitamos por referencia aquí, ya que cada producto filtrará por su nombre
        all_transactions = self.filter_transactions_by_type(df_elastic)
        
        if len(all_transactions) == 0:
            self.logger.warning(f"No hay transacciones disponibles")
            return {'reference': reference_name, 'products': 0, 'images': 0}
        
        # Obtener productos a descargar
        products = self.get_products_to_download(row)
        
        total_images = 0
        
        for product_info in products:
            product_name = product_info['name']
            product_type = product_info['type']
            
            # Crear ruta de carpeta según el tipo de producto
            if product_type == 'clase_principal':
                # Estructura: Referencia/Producto
                product_folder = os.path.join(ref_folder, product_name)
                self.logger.info(f"  Descargando imágenes para producto principal: {product_name}")
            else:
                # Estructura: Referencia/Reference_similar_X/label_id_label_name
                parent_similar_name = product_info.get('parent_similar_name', 'similar')
                label_id = product_info.get('label_id', 'unknown')
                
                # Convertir label_id a entero si es numérico
                if isinstance(label_id, (int, float)) and label_id != 'unknown':
                    label_id = int(label_id)
                
                # Crear carpeta padre para el grupo similar
                similar_folder = os.path.join(ref_folder, f"Reference_similar_{parent_similar_name}")
                os.makedirs(similar_folder, exist_ok=True)
                
                # Crear carpeta específica para este label_id/label_name
                product_folder = os.path.join(similar_folder, f"{label_id}_{product_name}")
                self.logger.info(f"  Descargando imágenes para sub-producto: {parent_similar_name} -> {product_name} (ID: {label_id})")
            
            os.makedirs(product_folder, exist_ok=True)
            
            # Seleccionar transacciones aleatorias para este producto específico
            selected_transactions = self.select_random_transactions(
                all_transactions, product_info, self.num_imagenes
            )
            
            # Descargar imágenes
            images_downloaded = 0
            for transaction in selected_transactions:
                if self.download_image_from_transaction(
                    transaction, 
                    product_folder, 
                    reference_name, 
                    product_name, 
                    product_type
                ):
                    images_downloaded += 1
            
            total_images += images_downloaded
            self.logger.info(f"    Descargadas {images_downloaded} imágenes para {product_name}")
        
        return {
            'reference': reference_name,
            'products': len(products),
            'images': total_images
        }

    def download_review_images(self, excel_path: Optional[str] = None) -> Dict:
        """
        Descargar imágenes para revisión según configuración en Excel
        Incluye tanto referencias como devices marcados para revisión
        
        Args:
            excel_path: Ruta al archivo Excel (opcional, usa el más reciente si no se especifica)
            
        Returns:
            Diccionario con estadísticas del procesamiento
        """
        try:
            # Encontrar archivo Excel si no se especifica
            if excel_path is None:
                excel_path = self.find_latest_excel()
                if excel_path is None:
                    raise ValueError("No se encontró archivo Excel para procesar")
            
            self.logger.info(f"Iniciando descarga de imágenes para revisión desde: {excel_path}")
            
            # Limpiar carpeta de salida si está configurado
            if self.clear_output_folder and os.path.exists(self.carpeta_destino):
                if not self.dry_run:
                    self.logger.info(f"Limpiando carpeta de salida: {self.carpeta_destino}")
                    shutil.rmtree(self.carpeta_destino)
                else:
                    self.logger.info(f"[DRY-RUN] Se limpiaría carpeta de salida: {self.carpeta_destino}")
            
            # Crear carpeta de salida
            os.makedirs(self.carpeta_destino, exist_ok=True)
            
            # Cargar datos
            df_references = self.load_references_data(excel_path)
            self.df_references = self.load_all_references_data(excel_path)  # Tabla completa de referencias para cruzar
            self.df_elastic = self.load_elasticsearch_data(excel_path)  # Guardar para usar en get_similar_sub_products
            
            # Cargar datos de devices
            df_devices = self.load_devices_data(excel_path)
            devices_to_process = df_devices[df_devices['revisar_imagenes'].astype(str).str.lower().isin(['si', 'sí', 'yes', 'y'])]
            
            # Verificar si hay algo que procesar
            if len(df_references) == 0 and len(devices_to_process) == 0:
                self.logger.warning("No hay referencias ni devices marcados para revisión")
                return {
                    'references_processed': 0, 
                    'devices_processed': 0,
                    'total_images': 0
                }
            
            total_images = 0
            stats = []
            device_stats = []
            
            # Procesar referencias
            if len(df_references) > 0:
                self.logger.info(f"Procesando {len(df_references)} referencias marcadas para revisión")
                for _, row in df_references.iterrows():
                    ref_stats = self.process_reference(row, self.df_elastic)
                    stats.append(ref_stats)
                    total_images += ref_stats['images']
            else:
                self.logger.info("No hay referencias marcadas para revisión")
            
            # Procesar devices
            if len(devices_to_process) > 0:
                self.logger.info(f"Procesando {len(devices_to_process)} devices marcados para revisión")
                for _, device_row in devices_to_process.iterrows():
                    device_result = self.process_device(device_row, self.df_elastic, self.dry_run)
                    device_stats.append(device_result)
                    total_images += device_result['images']
            else:
                self.logger.info("No hay devices marcados para revisión")
            
            # Resumen final
            summary = {
                'references_processed': len(stats),
                'devices_processed': len(device_stats),
                'total_images': total_images,
                'output_folder': self.carpeta_destino,
                'details': stats,
                'device_details': device_stats
            }
            
            self.logger.info(f"Descarga completada:")
            self.logger.info(f"  - Referencias procesadas: {summary['references_processed']}")
            self.logger.info(f"  - Devices procesados: {summary['devices_processed']}")
            self.logger.info(f"  - Total imágenes descargadas: {summary['total_images']}")
            self.logger.info(f"  - Carpeta de salida: {summary['output_folder']}")

            # Guardar log de imágenes descargadas en el Excel principal (o backup si está bloqueado)
            try:
                self._save_downloaded_images_log(excel_path)
            except Exception as e:
                self.logger.warning(f"No se pudo guardar log de imágenes en Excel: {e}")
            # Informe explícito del resultado del guardado
            if self.last_log_save_status == 'success':
                self.logger.info("✅ Log de imágenes añadido al Excel")
            elif self.last_log_save_status == 'locked':
                self.logger.warning("⚠ Excel estaba abierto: cierre el archivo y vuelva a ejecutar 'review_images' para registrar el log")
            elif self.last_log_save_status == 'no_images':
                self.logger.info("No había imágenes nuevas para registrar en el Excel")
            elif self.last_log_save_status == 'error':
                self.logger.error(f"No se pudo registrar el log en Excel: {self.last_log_save_error}")
            
            return summary
            
        except Exception as e:
            self.logger.error(f"Error en descarga de imágenes para revisión: {e}")
            raise

    def _check_excel_file_accessible(self, excel_path: str) -> Tuple[bool, str]:
        """
        Verifica si el archivo Excel está accesible para escritura
        
        Returns:
            Tuple[bool, str]: (is_accessible, error_message)
        """
        # Estrategia robusta en Windows: intentar renombrar temporalmente. Si falla -> bloqueado.
        try:
            if not os.path.exists(excel_path):
                return True, ""
            temp_name = excel_path + ".locktest"
            try:
                os.rename(excel_path, temp_name)
                os.rename(temp_name, excel_path)
                return True, ""
            except PermissionError:
                return False, "El archivo está abierto en Excel (bloqueado)"
            except OSError:
                # Cualquier OSError aquí interpretamos como bloqueo para este caso
                return False, "No se pudo obtener bloqueo exclusivo (posible archivo abierto)"
        except Exception as e:
            return False, f"Error verificando acceso: {e}"

    def _save_downloaded_images_log(self, excel_path):
        """Guarda el registro de imágenes descargadas separando references y devices en hojas diferentes"""
        manager = ReportExcelManager(excel_path, interactive=True, max_retries=3)
        try:
            manager.wait_until_unlocked()
        except ExcelLockedError as e:
            self.logger.error(str(e))
            self.last_log_save_status = 'locked'
            self.last_log_save_error = str(e)
            return
        
        try:
            if not self.downloaded_images_log:
                self.logger.info("No hay imágenes descargadas que registrar")
                self.last_log_save_status = 'no_images'
                return
            
            # Construir DataFrame de registros actuales
            df_downloaded = pd.DataFrame(self.downloaded_images_log)
            
            # Separar references y devices
            df_references = df_downloaded[df_downloaded['product_type'] != 'device'].copy()
            df_devices = df_downloaded[df_downloaded['product_type'] == 'device'].copy()
            
            # Optimización: reducir columnas para acelerar escritura
            essential_save_columns = [
                'reference_name', 'product_name', 'product_type',
                'download_status', 'transaction_id', 'device_id', 'device_name',
                'result_label_name', 's3_key', 'local_path'
            ]
            
            # Aplicar columnas esenciales para acelerar guardado
            if not df_references.empty:
                existing_columns = [col for col in essential_save_columns if col in df_references.columns]
                df_references = df_references[existing_columns]
            
            if not df_devices.empty:
                existing_columns = [col for col in essential_save_columns if col in df_devices.columns]
                df_devices = df_devices[existing_columns]
                # Normalizar device_id a entero cuando sea numérico
                if 'device_id' in df_devices.columns:
                    df_devices['device_id'] = pd.to_numeric(df_devices['device_id'], errors='coerce').astype('Int64')
            
            def _modify(wb):
                # Insertar / reemplazar hojas usando openpyxl workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                insert_after = None
                if 'Devices' in wb.sheetnames:
                    insert_after = wb.sheetnames.index('Devices') + 1
                idx = insert_after
                if df_references is not None and not df_references.empty:
                    if 'revisar_imagenes_bajadas' in wb.sheetnames:
                        del wb['revisar_imagenes_bajadas']
                    s = wb.create_sheet('revisar_imagenes_bajadas', index=idx if idx is not None else None)
                    for r in dataframe_to_rows(df_references, index=False, header=True):
                        s.append(r)
                    if idx is not None:
                        idx += 1
                if df_devices is not None and not df_devices.empty:
                    if 'devices_imagenes_bajadas' in wb.sheetnames:
                        del wb['devices_imagenes_bajadas']
                    s2 = wb.create_sheet('devices_imagenes_bajadas', index=idx if idx is not None else None)
                    from openpyxl.utils.dataframe import dataframe_to_rows as _dtr
                    for r in _dtr(df_devices, index=False, header=True):
                        s2.append(r)
            # Guardar con manager (reaplica validaciones después)
            if not os.path.exists(excel_path):
                # Crear un workbook base vacío para que el manager pueda operar
                import openpyxl as _ox
                base_wb = _ox.Workbook()
                # Evitar hoja por defecto si vamos a insertar nuevas hojas
                default_sheet = base_wb.active
                base_wb.remove(default_sheet)
                base_wb.save(excel_path)
            # Guardar (creación o actualización) vía manager asegurando validaciones
            manager.save_with_validations(_modify)
            self.logger.info(f"Registro de imágenes descargadas guardado en {excel_path}")
            self.logger.info(f"Total de registros guardados (incluyendo pendientes): {len(df_downloaded)}")
            self.last_log_save_status = 'success'
            self.last_log_save_error = None
            
                        
        except Exception as e:
            self.logger.error(f"Error al guardar registro de imágenes descargadas: {e}. No se crea archivo alternativo para evitar inconsistencias.")
            self.last_log_save_status = 'error'
            self.last_log_save_error = str(e)

    def _save_with_validation_preservation(self, excel_path: str, df_references: pd.DataFrame, df_devices: pd.DataFrame):
        """Guardar hojas de log preservando validaciones existentes."""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            print(1)
            wb = openpyxl.load_workbook(excel_path)
            backup = {}
            for sn in wb.sheetnames:
                sh = wb[sn]
                backup[sn] = []
                if hasattr(sh, 'data_validations'):
                    for dv in getattr(sh.data_validations, 'dataValidation', []):
                        backup[sn].append({
                            'formula1': dv.formula1,
                            'formula2': dv.formula2,
                            'type': dv.type,
                            'operator': dv.operator,
                            'ranges': [str(r) for r in dv.ranges],
                            'allow_blank': dv.allow_blank,
                            'error': dv.error,
                            'errorTitle': dv.errorTitle,
                            'prompt': dv.prompt,
                            'promptTitle': dv.promptTitle
                        })
            insert_after = None
            if 'Devices' in wb.sheetnames:
                insert_after = wb.sheetnames.index('Devices') + 1
            idx = insert_after
            if df_references is not None and not df_references.empty:
                if 'revisar_imagenes_bajadas' in wb.sheetnames:
                    del wb['revisar_imagenes_bajadas']
                s = wb.create_sheet('revisar_imagenes_bajadas', index=idx if idx is not None else None)
                for r in dataframe_to_rows(df_references, index=False, header=True):
                    s.append(r)
                if idx is not None:
                    idx += 1
            if df_devices is not None and not df_devices.empty:
                if 'devices_imagenes_bajadas' in wb.sheetnames:
                    del wb['devices_imagenes_bajadas']
                s2 = wb.create_sheet('devices_imagenes_bajadas', index=idx if idx is not None else None)
                for r in dataframe_to_rows(df_devices, index=False, header=True):
                    s2.append(r)
            self._restore_data_validations(wb, backup)
            wb.save(excel_path)
        except Exception as e:
            self.logger.error(f"Error guardando con preservación de validaciones: {e}")
            try:
                with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                    if df_references is not None and not df_references.empty:
                        df_references.to_excel(writer, sheet_name='revisar_imagenes_bajadas', index=False)
                    if df_devices is not None and not df_devices.empty:
                        df_devices.to_excel(writer, sheet_name='devices_imagenes_bajadas', index=False)
            except Exception as e2:
                self.logger.error(f"Fallo en fallback de guardado: {e2}")

    def download_device_images(self, excel_path: str = None, dry_run: bool = False) -> Dict:
        """
        Descargar imágenes para devices marcados con revisar_imagenes=si
        
        Args:
            excel_path: Ruta al archivo Excel con datos de devices
            dry_run: Si True, solo simula la descarga
            
        Returns:
            Diccionario con resultados del procesamiento
        """
        try:
            self.logger.info("Iniciando descarga de imágenes para devices desde: {}".format(excel_path or "archivo automático"))
            
            # Buscar archivo Excel si no se proporciona
            if not excel_path:
                excel_path = self.find_latest_excel()
                if not excel_path:
                    raise FileNotFoundError("No se encontró archivo Excel para procesar")
            
            # Limpiar carpeta de salida
            # Limpiar carpeta de salida si está configurado
            if self.clear_output_folder and os.path.exists(self.carpeta_destino):
                if not dry_run:
                    self.logger.info(f"Limpiando carpeta de salida: {self.carpeta_destino}")
                    import shutil
                    shutil.rmtree(self.carpeta_destino)
                else:
                    self.logger.info(f"[DRY-RUN] Se limpiaría carpeta de salida: {self.carpeta_destino}")
            
            # Crear carpeta de salida
            os.makedirs(self.carpeta_destino, exist_ok=True)
            
            # Cargar datos de devices
            devices_df = self.load_devices_data(excel_path)
            devices_to_process = devices_df[devices_df['revisar_imagenes'].astype(str).str.lower() == 'si']
            
            if devices_to_process.empty:
                self.logger.info("No hay devices marcados para revisión de imágenes")
                return {
                    'devices_processed': 0,
                    'total_images': 0,
                    'summary': {}
                }
            
            self.logger.info(f"Devices marcados para revisión: {len(devices_to_process)}")
            
            # Cargar datos de Elasticsearch
            elastic_df = self.load_elasticsearch_data(excel_path)
            
            total_images = 0
            device_results = []
            
            # Procesar cada device
            for _, device_row in devices_to_process.iterrows():
                device_result = self.process_device(device_row, elastic_df, dry_run)
                device_results.append(device_result)
                total_images += device_result['images']
            
            self.logger.info("Descarga completada:")
            self.logger.info(f"  - Devices procesados: {len(device_results)}")
            self.logger.info(f"  - Total imágenes descargadas: {total_images}")
            self.logger.info(f"  - Carpeta de salida: {self.carpeta_destino}")
            
            return {
                'devices_processed': len(device_results),
                'total_images': total_images,
                'summary': {device['device']: device['images'] for device in device_results}
            }
            
        except Exception as e:
            self.logger.error(f"Error en descarga de imágenes para devices: {e}")
            raise

    def load_devices_data(self, excel_path: str) -> pd.DataFrame:
        """Cargar datos de devices desde Excel"""
        try:
            devices_df = pd.read_excel(excel_path, sheet_name='Devices')
            self.logger.info(f"Cargados {len(devices_df)} devices desde {excel_path}")
            return devices_df
        except Exception as e:
            self.logger.error(f"Error cargando datos de devices: {e}")
            return pd.DataFrame()

    def process_device(self, device_row: pd.Series, elastic_df: pd.DataFrame, dry_run: bool) -> Dict:
        """
        Procesar un device para descarga de imágenes
        
        Args:
            device_row: Fila del DataFrame de devices
            elastic_df: DataFrame con datos de Elasticsearch
            dry_run: Si True, solo simula la descarga
            
        Returns:
            Diccionario con estadísticas del procesamiento
        """
        device_id = device_row.get('device_id')
        device_name = device_row.get('device_name') or device_id

        self.logger.info(f"Procesando device: {device_name} (ID: {device_id})")

        # Normalizar device_id para comparación (puede venir como int o str)
        try:
            device_id_int = int(device_id) if pd.notna(device_id) else None
        except Exception:
            device_id_int = None

        # Filtrar transacciones del device
        df_tx = elastic_df.copy()
        if device_id_int is not None and 'device_id' in df_tx.columns:
            # Convertir a numérico seguro para comparar
            df_tx['__dev_id_num'] = pd.to_numeric(df_tx['device_id'], errors='coerce')
            df_tx = df_tx[df_tx['__dev_id_num'] == device_id_int]
        elif 'device_name' in df_tx.columns and pd.notna(device_name):
            df_tx = df_tx[df_tx['device_name'].astype(str).str.lower() == str(device_name).lower()]

        # Aplicar filtro de tipo de transacciones (cliente / manual / ambas)
        if not df_tx.empty:
            df_tx = self.filter_transactions_by_type(df_tx)

        if df_tx.empty:
            self.logger.warning(f"No hay transacciones para device {device_name}")
            return { 'device': device_name, 'images': 0, 'transactions': 0 }

        # Quedarnos solo con transacciones que tengan imagen
        if 'has_image' in df_tx.columns:
            df_tx = df_tx[df_tx['has_image'] == True]
        else:
            # Intentar deducir a partir de image_link
            if 'image_link' in df_tx.columns:
                df_tx = df_tx[df_tx['image_link'].notna() & (df_tx['image_link'] != '') & (df_tx['image_link'] != 'None')]

        if df_tx.empty:
            self.logger.warning(f"No hay transacciones con imágenes para device {device_name}")
            return { 'device': device_name, 'images': 0, 'transactions': 0 }

        # Seleccionar hasta N transacciones aleatorias
        num_to_select = min(self.num_imagenes, len(df_tx))
        selected_tx = df_tx.sample(n=num_to_select, random_state=42)

        # Carpeta de destino
        device_folder = os.path.join(self.carpeta_destino, 'devices', str(device_id))
        if not dry_run:
            os.makedirs(device_folder, exist_ok=True)

        images_downloaded = 0
        for _, tx in selected_tx.iterrows():
            success = self.download_image_from_transaction(
                tx,
                device_folder,
                reference_name=None,
                product_name=device_name,
                product_type='device'
            )
            if success:
                images_downloaded += 1

        self.logger.info(f"    Device {device_name}: {images_downloaded}/{num_to_select} imágenes descargadas")
        return {
            'device': device_name,
            'images': images_downloaded,
            'transactions': int(len(selected_tx))
        }
