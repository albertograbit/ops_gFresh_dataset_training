"""
CLI (Command Line Interface) para el Dataset Manager
Proporciona una interfaz de línea de comandos para facilitar el uso de la herramienta
"""

import argparse
import sys
import os
import json
import csv
from datetime import datetime
from pathlib import Path

# Agregar el directorio raíz al path para imports
current_dir = Path(__file__).parent
project_root = current_dir.parent.parent
sys.path.insert(0, str(project_root))

from dataset_manager.core.dataset_processor import DatasetProcessor
from dataset_manager.config.settings import Settings
from dataset_manager.process_manager import ProcessManager
from dataset_manager.process_aware_settings import ProcessAwareSettings

def create_parser():
    """Crea y configura el parser de argumentos"""
    # Cargar configuración para obtener defaults
    try:
        settings = Settings()
        default_days_back = getattr(settings.deployment, 'default_days_back', 30)
    except:
        default_days_back = 30
    
    parser = argparse.ArgumentParser(
        description='Dataset Manager - Herramienta para gestión de datasets de retail',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  %(prog)s analyze 122 --days-back 30
    %(prog)s download_info 122 --config config/settings.yaml
    %(prog)s status --config config/settings.yaml
  %(prog)s download_info 122 --min-appearances 100 --confidence 0.8
  %(prog)s review_images -e path/to/excel_file.xlsx
  %(prog)s create_dataset --dry-run
  %(prog)s create_dataset --force -e path/to/excel_file.xlsx
        """
    )
    
    parser.add_argument(
        '--config', '-c',
        type=str,
        help='Ruta al archivo de configuración YAML personalizado'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Activar modo verbose con logs detallados'
    )
    
    parser.add_argument(
        '--output-dir',
        type=str,
        help='Directorio base de salida (sobrescribe configuración)'
    )
    
    # Subcomandos
    subparsers = parser.add_subparsers(
        dest='command',
        help='Comandos disponibles',
        title='Comandos'
    )
    # Comando: folder_structure (nuevo nombre directo)
    folder_structure_parser = subparsers.add_parser(
        'folder_structure',
        help='Generar CSV con la estructura y conteo de imágenes por subcarpeta (modo interactivo si faltan parámetros)'
    )
    folder_structure_parser.add_argument(
        '-i', '--input', required=False,
        help='Ruta a la carpeta raíz a inspeccionar (si se omite, detecta datasets del proceso activo y permite elegir subcarpeta)'
    )
    folder_structure_parser.add_argument(
        '-o', '--output', required=False,
        help='Ruta del CSV de salida (si se omite se crea en la raíz de datasets: folder_structure_<timestamp>.csv)'
    )
    folder_structure_parser.add_argument(
        '--extensions', type=str,
        help='Lista de extensiones separadas por comas (por defecto: jpg,jpeg,png,gif,bmp,webp)'
    )
    # Comando: merge_datasets (juntar varios datasets en uno nuevo)
    merge_parser = subparsers.add_parser(
        'merge_datasets',
        help='Fusionar varios datasets existentes en uno nuevo (interactivo)'
    )
    merge_parser.add_argument(
        '-n', '--name', required=False,
        help='Nombre del dataset final (si se omite se pedirá)'
    )
    # Comando: summary_dataset
    summary_parser = subparsers.add_parser(
        'summary_dataset',
        help='Generar un CSV resumen de un dataset existente cruzando con datos de Elasticsearch'
    )
    summary_parser.add_argument(
        '-d', '--dataset', required=False,
        help='Nombre del dataset a resumir (si se omite se pedirá)'
    )
    summary_parser.add_argument(
        '--elastic-credentials', required=False,
        help='Ruta al fichero JSON de credenciales de Elasticsearch (default: config/credentials/credentials_elastic_prod.json)'
    )
    summary_parser.add_argument(
        '--index', required=False, default='device_transaction-*',
        help='Patrón de índice de Elasticsearch (default: device_transaction-*)'
    )
    # Comando: register_dataset
    register_parser = subparsers.add_parser(
        'register_dataset',
        help='Empaquetar un dataset en ZIP, subirlo a S3 y registrar metadatos en CSV global'
    )
    register_parser.add_argument(
        '-d','--dataset', required=False,
        help='Nombre del dataset a registrar (si se omite se pedirá)'
    )
    register_parser.add_argument(
        '--s3-base', required=False,
        help='Ruta base S3 (default: s3://grabit-data/datasets/)'
    )
    register_parser.add_argument(
        '--registry-csv', required=False,
        help='Ruta al CSV de registro global (default: C:/Users/AlbertoGómez/Grabit/Operaciones - General/Datasets/datasets_gfresh.csv)'
    )
    register_parser.add_argument(
        '--skip-summary-copy', action='store_true',
        help='No copiar el summary al directorio local del registro'
    )
    
    # Comando: download_info (procesamiento completo)
    process_parser = subparsers.add_parser(
        'download_info',
        help='Descargar información y análisis (datos + informes, sin imágenes)'
    )
    process_parser.add_argument(
        'deployment_id',
        type=int,
        help='ID del deployment a procesar'
    )
    process_parser.add_argument(
        '--days-back',
        type=int,
        default=default_days_back,
        help=f'Número de días hacia atrás para consulta Elasticsearch (default: {default_days_back})'
    )
    process_parser.add_argument(
        '--model',
        type=int,
        help='ID del modelo específico a usar. Si no se especifica, usa el del deployment.'
    )
    process_parser.add_argument(
        '--no-reports',
        action='store_true',
        help='Omitir generación de informes Excel'
    )
    
    # Parámetros de configuración dinámicos
    process_parser.add_argument(
        '--min-appearances',
        type=int,
        help='Mínimo número de apariciones para filtrar referencias'
    )
    process_parser.add_argument(
        '--confidence',
        type=float,
        help='Umbral de confianza para sugerencias (0.0-1.0)'
    )
    process_parser.add_argument(
        '--parallel-workers',
        type=int,
        help='Número de workers paralelos para procesamiento'
    )
    process_parser.add_argument(
        '--base-model-id',
        type=int,
        help='ID del modelo base a usar (opcional, si no se especifica usa el último modelo)'
    )
    process_parser.add_argument(
        '--process-name',
        type=str,
        help='Nombre personalizado del proceso (ej: dataset_125_010825_v1). Si no se especifica, se genera automáticamente.'
    )
    
    # Comando: status
    status_parser = subparsers.add_parser(
        'status',
        help='Mostrar estado y configuración del sistema'
    )
    
    # Comando: active-process
    active_process_parser = subparsers.add_parser(
        'active-process',
        help='Mostrar información del proceso activo'
    )
    active_process_parser.add_argument(
        '--clear',
        action='store_true',
        help='Limpiar el proceso activo'
    )
    active_process_parser.add_argument(
        '--set',
        type=str,
        help='Establecer un proceso existente como activo (nombre del proceso)'
    )
    active_process_parser.add_argument(
        '--list',
        action='store_true',
        help='Listar todos los procesos disponibles'
    )
    
    
    # Comando: review_images
    review_parser = subparsers.add_parser(
        'review_images',
        help='Descargar imágenes para revisión manual desde AWS S3'
    )
    review_parser.add_argument(
        '--excel-file', '-e',
        type=str,
        help='Ruta específica al archivo Excel. Si no se especifica, usa el más reciente'
    )
    review_parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Modo de prueba sin descargar imágenes realmente'
    )
    
    # Comando: create_dataset (renombrado desde crear_dataset)
    dataset_parser = subparsers.add_parser(
        'create_dataset',
        help='Crear dataset de entrenamiento organizando imágenes por clase y label_id'
    )
    dataset_parser.add_argument(
        '--excel-file', '-e',
        type=str,
        help='Ruta específica al archivo Excel. Si no se especifica, usa el más reciente'
    )
    dataset_parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Modo de prueba sin descargar imágenes realmente'
    )
    dataset_parser.add_argument(
        '--force',
        action='store_true',
        help='Forzar creación aunque ya exista el directorio'
    )
    # Modo rápido y límites para pruebas rápidas
    dataset_parser.add_argument(
        '--fast',
        action='store_true',
        help='Modo rápido: limita referencias/devices e imágenes por cada uno para pruebas'
    )
    dataset_parser.add_argument(
        '--limit-refs',
        type=int,
        help='Máximo de referencias a procesar en modo rápido'
    )
    dataset_parser.add_argument(
        '--limit-devices',
        type=int,
        help='Máximo de devices adicionales a procesar en modo rápido'
    )
    dataset_parser.add_argument(
        '--images-per-ref',
        type=int,
        help='Imágenes a descargar por referencia en modo rápido'
    )
    dataset_parser.add_argument(
        '--images-per-device',
        type=int,
        help='Imágenes a descargar por device adicional en modo rápido'
    )
    dataset_parser.add_argument(
        '--max-workers',
        type=int,
        help='Sobrescribe el número de workers paralelos para descarga de imágenes'
    )
    dataset_parser.add_argument(
        '--no-filter-used',
        action='store_true',
        help='No excluir transacciones usadas en datasets previos (permitir reutilización)'
    )
    
    return parser

def setup_custom_settings(args, base_settings):
    """Configura settings personalizados basados en argumentos CLI"""
    custom_settings = {}
    
    # Configuraciones de análisis de referencias
    if hasattr(args, 'min_appearances') and args.min_appearances is not None:
        custom_settings.setdefault('reference_analysis', {})
        custom_settings['reference_analysis']['min_appearances'] = args.min_appearances
    
    if hasattr(args, 'confidence') and args.confidence is not None:
        if not 0.0 <= args.confidence <= 1.0:
            raise ValueError("El umbral de confianza debe estar entre 0.0 y 1.0")
        custom_settings.setdefault('reference_analysis', {})
        custom_settings['reference_analysis']['confidence_threshold'] = args.confidence
    
    # Configuraciones de procesamiento
    if hasattr(args, 'parallel_workers') and args.parallel_workers is not None:
        custom_settings.setdefault('processing', {})
        custom_settings['processing']['parallel_workers'] = args.parallel_workers
    
    # Configuración de directorios de salida
    if args.output_dir:
        custom_settings.setdefault('image_download', {})
        custom_settings.setdefault('reports', {})
        custom_settings['image_download']['output_directory'] = os.path.join(args.output_dir, 'images')
        custom_settings['reports']['output_directory'] = os.path.join(args.output_dir, 'reports')
    
    return custom_settings

def handle_analyze_command(args):
    """Maneja el comando analyze"""
    print(f"Iniciando análisis para deployment {args.deployment_id}")
    
    try:
        # Inicializar procesador
        processor = DatasetProcessor(config_path=args.config)
        
        # Ejecutar análisis
        results = processor.process_analysis_only(
            deployment_id=args.deployment_id,
            days_back=args.days_back,
            model_id=getattr(args, 'model', None)
        )
        
        # Mostrar resumen
        print("\nAnálisis completado exitosamente")
        print_analysis_summary(results)
        
        return 0
        
    except Exception as e:
        print(f"\nError durante el análisis: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1

def handle_download_info_command(args):
    """Maneja el comando download_info"""
    print(f"Iniciando procesamiento completo para deployment {args.deployment_id}")
    
    try:
        # Inicializar gestor de procesos
        process_manager = ProcessManager()
        
        # IMPORTANTE: Limpiar proceso activo anterior
        # Un nuevo download_info siempre debe crear un proceso independiente
        process_manager.clear_active_process()
        
        # Generar nombre del proceso
        process_name = process_manager.generate_process_name(
            str(args.deployment_id), 
            getattr(args, 'process_name', None)
        )
        
        print(f"Nombre del proceso: {process_name}")
        
        # Establecer como proceso activo
        process_metadata = {
            'deployment_id': args.deployment_id,
            'days_back': args.days_back,
            'download_images': False,  # download_info nunca descarga imágenes
            'generate_reports': not args.no_reports,
            'model_id': getattr(args, 'model', None),
            'base_model_id': getattr(args, 'base_model_id', None)
        }
        process_manager.set_active_process(process_name, process_metadata)
        
        # Configurar settings personalizados
        base_settings = Settings(args.config)
        custom_settings = setup_custom_settings(args, base_settings)
        
        # Inicializar procesador
        if custom_settings:
            processor = DatasetProcessor(config_path=args.config)
            results = processor.process_with_custom_settings(
                deployment_id=args.deployment_id,
                custom_settings=custom_settings,
                days_back=args.days_back,
                download_images=False,  # download_info nunca descarga imágenes
                generate_reports=not args.no_reports,
                base_model_id=getattr(args, 'base_model_id', None),
                model_id=getattr(args, 'model', None)
            )
        else:
            processor = DatasetProcessor(config_path=args.config)
            results = processor.process_deployment(
                deployment_id=args.deployment_id,
                days_back=args.days_back,
                download_images=False,  # download_info nunca descarga imágenes
                generate_reports=not args.no_reports,
                base_model_id=getattr(args, 'base_model_id', None),
                model_id=getattr(args, 'model', None)
            )
        
        # Mostrar resumen
        print("\nProcesamiento completado exitosamente")
        
        # Obtener información del proceso para mostrar directorio
        active_process = process_manager.get_active_process()
        if active_process:
            print(f"Proceso: {process_name}")
            print(f"Directorio: {active_process.get('process_dir', 'N/A')}")
        else:
            print(f"Proceso: {process_name}")
        
        print_processing_summary(results)
        
        return 0
        
    except Exception as e:
        print(f"\nError durante el procesamiento: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1

def handle_status_command(args):
    """Maneja el comando status"""
    print("Estado del Sistema Dataset Manager")
    print("=" * 50)
    
    try:
        processor = DatasetProcessor(config_path=args.config)
        status = processor.get_processing_status()
        
        print(f"Versión: {status['system_info']['version']}")
        print(f"Configuración: {status['system_info']['config_path']}")
        print(f"Config Elasticsearch: {status['system_info']['elasticsearch_config']}")
        print(f"\nDirectorios de salida:")
        print(f"  • Imágenes: {status['system_info']['output_directories']['images']}")
        print(f"  • Informes: {status['system_info']['output_directories']['reports']}")
        
        print(f"\nConfiguración actual:")
        print(f"  • Apariciones mínimas: {status['configuration']['min_appearances']}")
        print(f"  • Umbral confianza: {status['configuration']['confidence_threshold']:.1%}")
        print(f"  • Imágenes por clase: {status['configuration']['images_per_class']}")
        print(f"  • Workers paralelos: {status['configuration']['parallel_workers']}")
        
        print(f"\nÚltima consulta: {status['timestamp']}")
        
        return 0
        
    except Exception as e:
        print(f"\nError obteniendo estado: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1

def handle_active_process_command(args):
    """Maneja el comando active-process"""
    try:
        process_manager = ProcessManager()
        
        if args.clear:
            process_manager.clear_active_process()
            print("✅ Proceso activo limpiado")
            return 0
        
        if args.list:
            # Listar todos los procesos disponibles
            print("📁 Procesos Disponibles")
            print("=" * 30)
            
            output_dir = Path("output") / "processes"
            if not output_dir.exists():
                print("No hay procesos disponibles")
                return 0
            
            processes = []
            for proc_dir in output_dir.iterdir():
                if proc_dir.is_dir():
                    # Intentar leer metadata del proceso
                    metadata_file = proc_dir / "process_metadata.json"
                    if metadata_file.exists():
                        try:
                            with open(metadata_file, 'r', encoding='utf-8') as f:
                                metadata = json.load(f)
                            processes.append({
                                'name': proc_dir.name,
                                'created': metadata.get('created_at', 'N/A'),
                                'deployment_id': metadata.get('deployment_id', 'N/A')
                            })
                        except:
                            processes.append({
                                'name': proc_dir.name,
                                'created': 'N/A',
                                'deployment_id': 'N/A'
                            })
                    else:
                        processes.append({
                            'name': proc_dir.name,
                            'created': 'N/A',
                            'deployment_id': 'N/A'
                        })
            
            if not processes:
                print("No hay procesos disponibles")
            else:
                for proc in sorted(processes, key=lambda x: x['name']):
                    print(f"  • {proc['name']}")
                    print(f"    Deployment: {proc['deployment_id']}, Creado: {proc['created']}")
            
            return 0
        
        if args.set:
            # Establecer un proceso existente como activo
            process_name = args.set
            
            # Verificar si el proceso existe (buscar en carpetas de output)
            output_dir = Path("output") / "processes"
            process_dir = output_dir / process_name
            
            if not process_dir.exists():
                print(f"❌ El proceso '{process_name}' no existe")
                print("Procesos disponibles:")
                if output_dir.exists():
                    for proc_dir in output_dir.iterdir():
                        if proc_dir.is_dir():
                            print(f"  • {proc_dir.name}")
                else:
                    print("  No hay procesos disponibles")
                return 1
            
            # Establecer como activo
            metadata = {
                'process_dir': str(process_dir),
                'set_manually': True
            }
            process_manager.set_active_process(process_name, metadata)
            print(f"✅ Proceso '{process_name}' establecido como activo")
            return 0
        
        active_process = process_manager.get_active_process()
        
        if not active_process:
            print("❌ No hay ningún proceso activo")
            return 0
        
        print("📋 Información del Proceso Activo")
        print("=" * 50)
        print(f"Nombre: {active_process['name']}")
        print(f"Creado: {active_process['created_at']}")
        print(f"Directorio: {active_process.get('process_dir', 'N/A')}")
        
        if 'metadata' in active_process and active_process['metadata']:
            metadata = active_process['metadata']
            print(f"\nDetalles:")
            print(f"  • Deployment ID: {metadata.get('deployment_id', 'N/A')}")
            print(f"  • Días hacia atrás: {metadata.get('days_back', 'N/A')}")
            print(f"  • Descarga imágenes: {'Sí' if metadata.get('download_images', False) else 'No'}")
            print(f"  • Generar informes: {'Sí' if metadata.get('generate_reports', False) else 'No'}")
            if metadata.get('model_id'):
                print(f"  • Model ID: {metadata['model_id']}")
            if metadata.get('base_model_id'):
                print(f"  • Base Model ID: {metadata['base_model_id']}")
        
        return 0
        
    except Exception as e:
        print(f"Error obteniendo proceso activo: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


def handle_create_dataset_command(args):
    """Maneja el comando crear_dataset"""
    try:
        # Importar el módulo de creación de datasets
        from dataset_manager.dataset_creator.creator import DatasetCreator
        from dataset_manager.process_logger import ProcessLoggerManager
        import logging
        
        # Configurar logging
        level = logging.DEBUG if args.verbose else logging.INFO
        logging.basicConfig(
            level=level,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        
        logger = logging.getLogger(__name__)
        
        # Usar configuración consciente de procesos
        process_aware_settings = ProcessAwareSettings(args.config)
        
        # Verificar si hay proceso activo
        process_info = process_aware_settings.get_process_info()
        if not process_info:
            logger.error("❌ No hay proceso activo. Ejecute primero 'download_info' para crear un proceso.")
            return 1
        
        # Obtener deployment_id del proceso activo
        deployment_id = process_info.get('metadata', {}).get('deployment_id')
        if not deployment_id:
            logger.error("❌ El proceso activo no tiene deployment_id válido.")
            return 1
        
        # Setup process logger
        process_directory = process_info.get('process_dir') or process_info.get('directory')
        if process_directory:
            process_logger = ProcessLoggerManager.get_logger(
                process_info['name'], 
                process_directory
            )
        else:
            process_logger = None
        
        if process_logger:
            process_logger.log_command_start(
                "create_dataset",
                f"Creación de dataset - deployment {deployment_id} - dry_run: {'Sí' if args.dry_run else 'No'}"
            )
        
        # Crear dataset creator con configuración consciente de procesos
        creator = DatasetCreator(config_path=args.config, process_aware=True)
        
        if args.dry_run:
            logger.info("🔍 MODO DRY-RUN: Solo mostrar qué se haría")
            if process_logger:
                process_logger.log_detailed("Modo DRY-RUN activado - solo simulación")
        
        # Mostrar información del proceso activo
        logger.info(f"📋 Usando proceso activo: {process_info['name']}")
        logger.info(f"🎯 Deployment ID: {deployment_id}")
        logger.info(f"📁 Directorio del proceso: {process_directory}")
        
        if process_logger:
            process_logger.log_detailed(f"Deployment ID: {deployment_id}")
            process_logger.log_detailed(f"Directorio del proceso: {process_directory}")
        
        # Ejecutar creación de dataset
        try:
            results = creator.crear_dataset(
                deployment_id=deployment_id,
                excel_path=args.excel_file,
                dry_run=args.dry_run,
                force=args.force,
                fast=args.fast,
                limit_refs=getattr(args, 'limit_refs', None),
                limit_devices=getattr(args, 'limit_devices', None),
                images_per_ref=getattr(args, 'images_per_ref', None),
                images_per_device=getattr(args, 'images_per_device', None),
                max_workers=getattr(args, 'max_workers', None),
                no_filter_used=getattr(args, 'no_filter_used', False)
            )
        except ValueError as e:
            # Si es un error de validación, mostrar la información de forma prominente
            error_msg = str(e)
            if "referencias sin campos requeridos" in error_msg:
                # Mostrar tabla de validación al final
                print("\n" + "🔴" * 25 + " PROCESO DETENIDO " + "🔴" * 25)
                print("La tabla de validación se mostró arriba con los detalles específicos.")
                print("Corrija los campos faltantes en el Excel y vuelva a ejecutar crear_dataset")
                print("🔴" * 70)
            
            if process_logger:
                process_logger.log_command_result(
                    "create_dataset",
                    f"ERROR: {error_msg}",
                    success=False
                )
            return 1
        
        # Mostrar resultados
        logger.info("🎉 DATASET CREADO EXITOSAMENTE")
        logger.info("=" * 50)
        logger.info(f"📂 Carpeta dataset: {results['dataset_folder']}")
        logger.info(f"📋 Referencias procesadas: {results['references_processed']}")
        logger.info(f"🏷️  Clases creadas: {results['classes_created']}")
        logger.info(f"📸 Total de imágenes: {results['total_images']}")
        logger.info(f"📊 Archivo Excel: {results['excel_path']}")
        
        # Log del proceso
        if process_logger:
            process_logger.log_step_completion("Creación de dataset", {
                "Carpeta dataset": results['dataset_folder'],
                "Referencias procesadas": results['references_processed'],
                "Clases creadas": results['classes_created'],
                "Total imágenes": results['total_images']
            })
            
            process_logger.log_command_result(
                "create_dataset",
                f"Dataset creado: {results['references_processed']} referencias, {results['classes_created']} clases, {results['total_images']} imágenes",
                success=True
            )
        
        # Mostrar estructura de carpetas
        from pathlib import Path
        dataset_path = Path(results['dataset_folder'])
        if dataset_path.exists():
            logger.info("\n📁 Estructura de carpetas creada:")
            for clase_folder in sorted(dataset_path.iterdir()):
                if clase_folder.is_dir():
                    label_folders = list(clase_folder.iterdir())
                    logger.info(f"   {clase_folder.name}/")
                    for label_folder in sorted(label_folders):
                        if label_folder.is_dir():
                            images_count = len(list(label_folder.glob('*.jpg'))) + len(list(label_folder.glob('*.png')))
                            logger.info(f"     └── {label_folder.name}/ ({images_count} imágenes)")
        
        logger.info(f"\n✅ Proceso completado para deployment {deployment_id}")
        return 0
        
    except KeyboardInterrupt:
        logger.info("Creación de dataset cancelada por el usuario")
        if process_logger:
            process_logger.log_command_result("create_dataset", "Cancelado por el usuario", success=False)
        return 1
    except Exception as e:
        logger.error(f"❌ Error durante la creación del dataset: {str(e)}")
        if process_logger:
            process_logger.log_command_result("create_dataset", f"ERROR: {str(e)}", success=False)
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


def handle_review_images_command(args):
    """Maneja el comando review_images"""
    try:
        # Importar el módulo de revisión de imágenes
        from dataset_manager.image_review.downloader import ImageReviewDownloader
        from dataset_manager.process_logger import ProcessLoggerManager
        import logging
        
        # Configurar logging
        level = logging.DEBUG if args.verbose else logging.INFO
        logging.basicConfig(
            level=level,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        
        logger = logging.getLogger(__name__)
        
        # Usar configuración consciente de procesos
        process_aware_settings = ProcessAwareSettings(args.config)
        
        # Setup process logger si hay proceso activo
        process_logger = None
        process_info = process_aware_settings.get_process_info()
        if process_info:
            # Usar process_dir si está disponible, sino construir el directorio
            process_directory = process_info.get('process_dir') or process_info.get('directory')
            if process_directory:
                process_logger = ProcessLoggerManager.get_logger(
                    process_info['name'], 
                    process_directory
                )
            process_logger.log_command_start(
                "review_images",
                f"Descarga de imágenes para revisión - dry_run: {'Sí' if args.dry_run else 'No'}"
            )
        
        # Crear downloader con configuración adaptada
        downloader = ImageReviewDownloader(config_path=args.config, process_aware=True)
        downloader.dry_run = args.dry_run
        
        if args.dry_run:
            logger.info("=== MODO DRY-RUN: Solo mostrar qué se haría ===")
            if process_logger:
                process_logger.log_detailed("Modo DRY-RUN activado - solo simulación")
        
        # Mostrar información del proceso activo
        if process_info:
            logger.info(f"Usando proceso activo: {process_info['name']}")
            output_path = process_aware_settings.get_output_path("images/productos")
            logger.info(f"Directorio de salida: {output_path}")
            if process_logger:
                process_logger.log_detailed(f"Directorio de salida: {output_path}")
        
        # Determinar archivo Excel
        excel_file = args.excel_file or downloader.find_latest_excel()
        if not excel_file:
            logger.error("No se encontró archivo Excel")
            if process_logger:
                process_logger.log_command_result(
                    "review_images",
                    "ERROR: No se encontró archivo Excel",
                    success=False
                )
            return 1
            
        logger.info(f"Iniciando descarga de imágenes para revisión desde: {excel_file}")
        if process_logger:
            process_logger.log_detailed(f"Archivo Excel: {excel_file}")
        
        # Ejecutar descarga
        results = downloader.download_review_images(excel_file)
        
        # Generar archivo Excel con registro de imágenes descargadas
        if downloader.downloaded_images_log:
            # Usar el archivo original para agregar la nueva hoja
            downloader._save_downloaded_images_log(excel_file)
            logger.info(f"Registro de imágenes descargadas guardado en hoja 'revisar_imagenes_bajadas' del archivo original: {excel_file}")
        
        # Mostrar resultados
        logger.info("=== RESUMEN DE RESULTADOS ===")
        logger.info(f"Referencias procesadas: {results.get('references_processed', 0)}")
        logger.info(f"Devices procesados: {results.get('devices_processed', 0)}")
        logger.info(f"Referencias con revisar_imagenes=si: {len([d for d in results.get('details', []) if d.get('images', 0) > 0])}")
        logger.info(f"Devices con revisar_imagenes=si: {len([d for d in results.get('device_details', []) if d.get('images', 0) > 0])}")
        logger.info(f"Imágenes descargadas: {results.get('total_images', 0)}")
        logger.info(f"Errores: {sum(d.get('errors', 0) for d in results.get('details', []))}")
        
        # Log del proceso
        if process_logger:
            process_logger.log_step_completion("Descarga de imágenes para revisión", {
                "Referencias procesadas": results.get('references_processed', 0),
                "Devices procesados": results.get('devices_processed', 0),
                "Referencias marcadas": len([d for d in results.get('details', []) if d.get('images', 0) > 0]),
                "Devices marcados": len([d for d in results.get('device_details', []) if d.get('images', 0) > 0]),
                "Imágenes descargadas": results.get('total_images', 0),
                "Errores": sum(d.get('errors', 0) for d in results.get('details', []))
            })
            
            process_logger.log_command_result(
                "review_images",
                f"Procesadas {results.get('references_processed', 0)} referencias y {results.get('devices_processed', 0)} devices, descargadas {results.get('total_images', 0)} imágenes",
                success=True
            )
        
        # Mostrar información del proceso
        if process_info:
            logger.info(f"Proceso: {process_info['name']}")
            process_directory = process_info.get('process_dir') or process_info.get('directory', 'N/A')
            logger.info(f"Directorio: {process_directory}")
        
        logger.info("Proceso completado exitosamente")
        return 0
        
    except Exception as e:
        print(f"Error en descarga de imágenes: {e}")
        import traceback
        if args.verbose:
            traceback.print_exc()
        return 1

def print_analysis_summary(results):
    """Imprime resumen del análisis"""
    summary = results.get('summary', {})
    
    print(f"\nRESUMEN DEL ANÁLISIS:")
    print(f"   • Registros Elasticsearch: {summary.get('data_extraction', {}).get('elastic_records', 0):,}")
    print(f"   • Referencias procesadas: {summary.get('data_extraction', {}).get('references', 0):,}")
    print(f"   • Sin asignar: {summary.get('reference_analysis', {}).get('unassigned_count', 0):,}")
    print(f"   • No entrenadas: {summary.get('reference_analysis', {}).get('untrained_count', 0):,}")
    print(f"   • Sugerencias alta confianza: {summary.get('reference_analysis', {}).get('high_confidence_suggestions', 0):,}")

def print_processing_summary(results):
    """Imprime resumen del procesamiento completo"""
    summary = results.get('summary', {})
    
    print(f"\nRESUMEN DEL PROCESAMIENTO:")
    
    # Datos
    print(f"   Datos extraídos:")
    print(f"      • Elasticsearch: {summary.get('data_extraction', {}).get('elastic_records', 0):,} registros")
    print(f"      • Referencias: {summary.get('data_extraction', {}).get('references', 0):,}")
    print(f"      • Labels: {summary.get('data_extraction', {}).get('labels', 0):,}")
    
    # Análisis
    print(f"   Análisis de referencias:")
    print(f"      • Sin asignar: {summary.get('reference_analysis', {}).get('unassigned_count', 0):,}")
    print(f"      • No entrenadas: {summary.get('reference_analysis', {}).get('untrained_count', 0):,}")
    print(f"      • Sugerencias confiables: {summary.get('reference_analysis', {}).get('high_confidence_suggestions', 0):,}")
    
    # Imágenes
    image_stats = summary.get('image_download', {})
    if image_stats:
        print(f"   📸 Imágenes descargadas:")
        print(f"      • Referencias nuevas: {image_stats.get('total_new_ref_images', 0):,}")
        print(f"      • Clases similares: {image_stats.get('total_similar_images', 0):,}")
    
    # Informes
    reports = summary.get('generated_reports', [])
    if reports:
        print(f"   Informes generados: {len(reports)}")
    
    # Tiempo
    processing_time = results.get('processing_metadata', {}).get('processing_time_seconds', 0)
    print(f"   Tiempo total: {processing_time:.1f} segundos")

def main():
    """Función principal del CLI"""
    parser = create_parser()
    args = parser.parse_args()
    
    # Mostrar ayuda si no se especifica comando
    if not args.command:
        parser.print_help()
        return 1
    
    # Configurar verbosidad
    if args.verbose:
        import logging
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Ejecutar comando correspondiente
    if args.command == 'analyze':
        return handle_analyze_command(args)
    elif args.command == 'download_info':
        return handle_download_info_command(args)
    elif args.command == 'status':
        return handle_status_command(args)
    elif args.command == 'active-process':
        return handle_active_process_command(args)
    elif args.command == 'review_images':
        return handle_review_images_command(args)
    elif args.command == 'create_dataset':
        return handle_create_dataset_command(args)
    elif args.command == 'folder_structure':
        return handle_folder_structure_command(args)
    elif args.command == 'merge_datasets':
        return handle_merge_datasets_command(args)
    elif args.command == 'summary_dataset':
        return handle_summary_dataset_command(args)
    elif args.command == 'register_dataset':
        return handle_register_dataset_command(args)
    else:
        print(f"Comando desconocido: {args.command}")
        return 1

if __name__ == '__main__':
    sys.exit(main())

def handle_folder_structure_command(args):
    """Genera un CSV con conteo de imágenes por subcarpeta.

    Reglas de salida:
      - Si se pasa -o se usa esa ruta.
      - Si NO se pasa -o, el CSV se crea SIEMPRE en la raíz de 'datasets' (datasets_root)
        aunque se elija una subcarpeta para analizar.
    """
    root: Path
    datasets_root: Path

    # Determinar ruta raíz (interactiva si no se pasa -i)
    if getattr(args, 'input', None):
        root = Path(args.input)
        datasets_root = root
        # Buscar carpeta datasets hacia arriba
        for parent in [root] + list(root.parents):
            if parent.name == 'datasets':
                datasets_root = parent
                break
    else:
        datasets_dir = None
        try:
            pa = ProcessAwareSettings(getattr(args, 'config', None))
            pinfo = pa.get_process_info()
            if pinfo and pinfo.get('process_dir'):
                candidate = Path(pinfo['process_dir']) / 'datasets'
                if candidate.exists():
                    datasets_dir = candidate
        except Exception:
            pass
        if datasets_dir is None:
            candidate_root = Path('output') / 'processes'
            if candidate_root.exists():
                for proc_dir in sorted(candidate_root.iterdir()):
                    dds = proc_dir / 'datasets'
                    if dds.exists():
                        datasets_dir = dds
                        break
        if datasets_dir is None:
            datasets_dir = Path('.')
        # Listar subcarpetas
        subfolders = [d for d in sorted(datasets_dir.iterdir()) if d.is_dir()]
        root = datasets_dir
        datasets_root = datasets_dir
        print(f"Carpeta datasets raíz detectada: {datasets_dir}")
        if subfolders and sys.stdin.isatty():
            print("Subcarpetas disponibles:")
            for idx, d in enumerate(subfolders, start=1):
                print(f"  {idx}. {d.name}")
            choice = input("Elige número de subcarpeta (ENTER para usar raíz): ").strip()
            if choice.isdigit():
                pos = int(choice)
                if 1 <= pos <= len(subfolders):
                    root = subfolders[pos - 1]
                    print(f"Seleccionada subcarpeta: {root}")
        elif not subfolders:
            print("No hay subcarpetas dentro de datasets; se usa la raíz")

    if not root.exists() or not root.is_dir():
        print(f"Ruta no válida: {root}")
        return 1

    # Determinar CSV de salida
    if getattr(args, 'output', None):
        out_csv = Path(args.output)
    else:
        out_csv = datasets_root / f"folder_structure_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    # Extensiones
    if getattr(args, 'extensions', None):
        exts = [f".{e.strip().lstrip('.').lower()}" for e in args.extensions.split(',') if e.strip()]
    else:
        exts = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']

    # Primera pasada: recolectar cuentas de imágenes por directorio
    dir_info = {}  # path -> {'images': int, 'subdirs': list[Path]}
    for curr_dir, subdirs, files in os.walk(root):
        img_count = 0
        for f in files:
            if Path(f).suffix.lower() in exts:
                img_count += 1
        dir_info[Path(curr_dir)] = {
            'images': img_count,
            'subdirs': [Path(curr_dir) / s for s in subdirs]
        }

    # Determinar qué directorios son hojas con imágenes (último nivel)
    leaf_rows = []
    for p, info in dir_info.items():
        if info['images'] == 0:
            continue  # excluir niveles sin imágenes
        # Si alguno de sus subdirs tiene imágenes, no es hoja
        has_child_with_images = any(
            (c in dir_info and dir_info[c]['images'] > 0) for c in info['subdirs']
        )
        if has_child_with_images:
            continue  # queremos solo último nivel
        rel = os.path.relpath(p, root)
        rel_norm = '.' if rel == '.' else rel.replace('\\', '/')
        if rel_norm == '.':
            # Raíz con imágenes y sin hijos con imágenes: usar nombre de la carpeta raíz
            label_id = p.name
            product = ''
        else:
            parts = rel_norm.split('/')
            label_id = parts[-1]
            product = parts[-2] if len(parts) >= 2 else ''
        # Nombre del dataset: primer componente relativo a datasets_root
        try:
            rel_ds = p.relative_to(datasets_root)
            dataset_name = rel_ds.parts[0] if rel_ds.parts else p.name
        except Exception:
            dataset_name = p.name
        leaf_rows.append({
            'dataset_name': dataset_name,
            'product': product,
            'label_id': label_id,
            'folder_rel': rel_norm,
            'images_count': info['images'],
            'folder_abs': str(p)
        })

    # Ordenar por producto y luego label_id para legibilidad
    leaf_rows.sort(key=lambda r: (r['product'], r['label_id']))

    fieldnames = ['dataset_name', 'product', 'label_id', 'folder_rel', 'images_count', 'folder_abs']
    with out_csv.open('w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        for r in leaf_rows:
            writer.writerow(r)

    total_dirs = len(leaf_rows)
    total_images = sum(r['images_count'] for r in leaf_rows)
    print(f"Directorios hoja con imágenes: {total_dirs}")
    print(f"Imágenes totales: {total_images}")
    print(f"CSV: {out_csv}")
    return 0

def handle_merge_datasets_command(args):
    """Fusionar múltiples datasets existentes en uno nuevo.

    Pasos:
      1. Detectar carpeta datasets del proceso activo (o buscar en output/processes/*/datasets primera existente).
      2. Listar datasets (directorios que empiezan por 'dataset_'). Selección iterativa por número; ENTER finaliza.
      3. Pedir nombre destino si no se pasó --name. Verificar que no exista.
      4. Copiar estructura e imágenes. Si hay colisión de nombre de archivo se sobreescribe y se contabiliza.
      5. Registrar en Excel (datasets_creados) filas nuevas con origen=juntar_datasets y dataset_original.
      6. Mostrar resumen (datasets fusionados, imágenes copiadas, sobrescritas, total final único).
    """
    import shutil
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd

    # 1. Localizar datasets_root
    datasets_root = None
    try:
        pa = ProcessAwareSettings(getattr(args, 'config', None))
        pinfo = pa.get_process_info()
        if pinfo and pinfo.get('process_dir'):
            candidate = Path(pinfo['process_dir']) / 'datasets'
            if candidate.exists():
                datasets_root = candidate
    except Exception:
        pass
    if datasets_root is None:
        candidate = Path('output') / 'processes'
        if candidate.exists():
            for proc_dir in sorted(candidate.iterdir()):
                dds = proc_dir / 'datasets'
                if dds.exists():
                    datasets_root = dds
                    break
    if datasets_root is None:
        print('No se encontró carpeta datasets')
        return 1

    # 2. Listar datasets disponibles (cualquier subcarpeta no oculta)
    available = [d for d in sorted(datasets_root.iterdir()) if d.is_dir() and not d.name.startswith('.')]
    if not available:
        print('No hay datasets para fusionar')
        return 1

    selected = []
    remaining = available.copy()
    while True:
        print('\nDatasets disponibles:')
        for idx, d in enumerate(remaining, start=1):
            print(f"  {idx}. {d.name}")
        choice = input('Selecciona un número (ENTER para terminar): ').strip()
        if choice == '':
            break
        if not choice.isdigit():
            print('Valor no válido')
            continue
        pos = int(choice)
        if not (1 <= pos <= len(remaining)):
            print('Número fuera de rango')
            continue
        ds = remaining.pop(pos-1)
        selected.append(ds)
        print(f'Añadido: {ds.name}')
        if not remaining:
            print('No quedan más datasets. Finalizando selección.')
            break

    if not selected:
        print('No se seleccionaron datasets')
        return 1

    print('\nDatasets a fusionar:')
    for d in selected:
        print(' -', d.name)

    # 3. Nombre destino
    target_name = args.name or input('Nombre del nuevo dataset: ').strip()
    if not target_name:
        print('Nombre destino vacío')
        return 1
    target_folder = datasets_root / target_name
    if target_folder.exists():
        print('El dataset destino ya existe; abortando para evitar sobrescritura accidental.')
        return 1
    target_folder.mkdir(parents=True, exist_ok=False)

    overwritten = 0
    copied = 0
    log_rows = []  # filas enriquecidas para Excel
    raw_copies = []  # tracking básico (dataset_original, clase, label, filename, dest_path)

    # 4. Copiar
    for src_ds in selected:
        for class_dir in src_ds.iterdir():
            if not class_dir.is_dir():
                continue
            for label_dir in class_dir.iterdir():
                if not label_dir.is_dir():
                    continue
                dest_label_dir = target_folder / class_dir.name / label_dir.name
                dest_label_dir.mkdir(parents=True, exist_ok=True)
                for img_file in label_dir.iterdir():
                    if not img_file.is_file():
                        continue
                    dest_file = dest_label_dir / img_file.name
                    if dest_file.exists():
                        overwritten += 1
                    try:
                        shutil.copy2(img_file, dest_file)
                        copied += 1
                        raw_copies.append({
                            'dataset_original': src_ds.name,
                            'class': class_dir.name,
                            'label_id': label_dir.name,
                            'filename': img_file.name,
                            'dest_path': str(dest_file)
                        })
                    except Exception as e:
                        print(f'Error copiando {img_file}: {e}')

    # 5. Registrar en Excel (datasets_creados)
    # Localizar Excel principal (usar el del proceso activo si existe)
    excel_path = None
    try:
        pa = ProcessAwareSettings(getattr(args, 'config', None))
        pinfo = pa.get_process_info()
        if pinfo and pinfo.get('process_dir'):
            reports_dir = Path(pinfo['process_dir']) / 'reports'
            if reports_dir.exists():
                # elegir el que coincide con el nombre del proceso o primero
                candidates = list(reports_dir.glob('*.xlsx'))
                if candidates:
                    # Preferir exacto
                    proc_name = pinfo.get('name')
                    exact = [c for c in candidates if c.stem == proc_name]
                    excel_path = str(exact[0] if exact else candidates[0])
    except Exception:
        pass

    # Enriquecer metadatos usando datasets_creados existente (si existe) y generar log_rows
    if raw_copies:
        try:
            sheet_name = 'datasets_creados'
            import pandas as pd
            rows_df = pd.DataFrame(raw_copies)
            # Cargar existente para merge
            try:
                existing_df_full = pd.read_excel(excel_path, sheet_name=sheet_name)
            except Exception:
                existing_df_full = pd.DataFrame()
            # Normalizar columnas clave para matching
            if not existing_df_full.empty:
                # Asegurar filename
                if 'filename' not in existing_df_full.columns and 'local_path' in existing_df_full.columns:
                    existing_df_full['filename'] = existing_df_full['local_path'].apply(lambda p: Path(str(p)).name if pd.notna(p) else '')
                if 'dataset_folder' in existing_df_full.columns and 'dataset_original' not in existing_df_full.columns:
                    existing_df_full['dataset_original'] = existing_df_full['dataset_folder']
                # Crear columna unificada de clase para merge sin modificar originales
                existing_df_full['merge_class'] = existing_df_full.apply(
                    lambda r: r['clase_name'] if 'clase_name' in r and pd.notna(r['clase_name']) else (r['class'] if 'class' in r else ''), axis=1
                )
                rows_df['merge_class'] = rows_df['class']
                # Dataset folder original para clave
                rows_df['dataset_folder'] = rows_df['dataset_original']
                # Definir posibles claves
                possible_cols = ['merge_class','label_id','filename','dataset_folder']
                have_cols_existing = [c for c in possible_cols if c in existing_df_full.columns]
                have_cols_rows = [c for c in possible_cols if c in rows_df.columns]
                key_cols = [c for c in possible_cols if c in have_cols_existing and c in have_cols_rows]
                # Normalizar tipos
                for col in key_cols:
                    existing_df_full[col] = existing_df_full[col].astype(str)
                    rows_df[col] = rows_df[col].astype(str)
                # Intentar merge completo
                merged = rows_df.copy()
                if key_cols:
                    try:
                        merged = rows_df.merge(existing_df_full.drop_duplicates(subset=key_cols), on=key_cols, how='left', suffixes=('','_orig'))
                    except Exception as e:
                        print(f"Aviso: fallo en merge principal ({e}); intentando merge por filename...")
                        # Fallback solo por filename
                        if 'filename' in key_cols:
                            try:
                                merged = rows_df.merge(existing_df_full.drop_duplicates(subset=['filename']), on='filename', how='left', suffixes=('','_orig'))
                            except Exception as e2:
                                print(f"Aviso: fallo en merge por filename ({e2}); usando datos básicos.")
                                merged = rows_df.copy()
                # Limpiar auxiliar
                if 'merge_class' in merged.columns:
                    merged = merged.rename(columns={'merge_class':'class'})
                # Eliminar columnas duplicadas exactas
                merged = merged.loc[:, ~merged.columns.duplicated()]
            else:
                merged = rows_df.copy()
            # Construir filas finales
            final_records = []
            # Columnas originales que intentaremos preservar
            preserved_cols = set(existing_df_full.columns.tolist() if not existing_df_full.empty else [])
            for _, r in merged.iterrows():
                base = {}
                for c in preserved_cols:
                    base[c] = r.get(c, '')
                # Overwrites específicos
                base['dataset_folder'] = target_folder.name
                base['clase_name'] = r.get('class','')  # para consistencia con logs previos
                base['class'] = r.get('class','')  # mantener si ya existía
                base['label_id'] = r.get('label_id','')
                base['local_path'] = r.get('dest_path','')
                base['filename'] = r.get('filename','')
                base['download_status'] = 'merged'
                base['origen'] = 'juntar_datasets'
                base['dataset_original'] = r.get('dataset_original','')
                base['relative_folder'] = f"{r.get('class','')}/{r.get('label_id','')}" if r.get('class') and r.get('label_id') else '.'
                final_records.append(base)
            log_rows = final_records
            # Si no hay excel_path no se puede registrar, pero se omite CSV temporal por petición.
            if not excel_path:
                print('No se encontró Excel para actualizar.')
        except Exception as e:
            print(f'Error preparando datos de merge: {e}')
            log_rows = []

    if excel_path and log_rows:
        try:
            sheet_name = 'datasets_creados'
            import pandas as pd
            rows_df = pd.DataFrame(log_rows)
            # Cargar workbook con fallback si está corrupto / no es zip
            from zipfile import BadZipFile
            try:
                workbook = openpyxl.load_workbook(excel_path)
            except (BadZipFile, OSError, KeyError) as e:
                print(f"Aviso: no se pudo abrir '{excel_path}' ({e}); se creará un nuevo Excel para logs de merge.")
                import openpyxl as _ox
                workbook = _ox.Workbook()
                # El workbook nuevo trae una hoja por defecto; la eliminaremos luego si no coincide
                # Asegurar que no haya conflicto de nombre
                if sheet_name in workbook.sheetnames:
                    del workbook[sheet_name]
            if sheet_name in workbook.sheetnames:
                existing_df = pd.read_excel(excel_path, sheet_name=sheet_name)
                # Asegurar columnas origen/dataset_original
                if 'origen' not in existing_df.columns:
                    existing_df['origen'] = 'crear_dataset'
                if 'dataset_original' not in existing_df.columns:
                    existing_df['dataset_original'] = ''
                combined_df = pd.concat([existing_df, rows_df], ignore_index=True)
                del workbook[sheet_name]
            else:
                combined_df = rows_df
            insert_after = None
            if 'Devices' in workbook.sheetnames:
                insert_after = workbook.sheetnames.index('Devices') + 1
            sheet = workbook.create_sheet(sheet_name, index=insert_after if insert_after is not None else None)
            for r in dataframe_to_rows(combined_df, index=False, header=True):
                sheet.append(r)
            try:
                workbook.save(excel_path)
            except Exception as e:
                # Último fallback: guardar a un nuevo archivo separado
                fallback_path = Path(excel_path).with_name(Path(excel_path).stem + '_merge_log.xlsx')
                print(f"Aviso: fallo guardando Excel principal ({e}); guardando en {fallback_path}")
                try:
                    workbook.save(fallback_path)
                except Exception as e2:
                    print(f"Error guardando incluso fallback: {e2}")
            print(f'Registro Excel actualizado: {excel_path}')
        except PermissionError:
            print('Excel abierto; no se pudo actualizar el log.')
        except Exception as e:
            print(f'Error actualizando Excel: {e}')

    # 6. Guardar resumen y reporte de sobrescritas
    overwrite_report = target_folder / 'merge_overwritten.txt'
    with overwrite_report.open('w', encoding='utf-8') as f:
        f.write(f'Imágenes sobrescritas: {overwritten}\n')
    # Contar solo ficheros de imagen comunes
    img_exts = {'.jpg','.jpeg','.png','.gif','.bmp','.webp'}
    unique_images = sum(1 for _ in target_folder.rglob('*') if _.is_file() and _.suffix.lower() in img_exts)
    print('\n--- RESUMEN MERGE ---')
    print(f'Datasets fusionados: {len(selected)}')
    print(f'Imágenes copiadas (incluye sobrescritas): {copied}')
    print(f'Imágenes sobrescritas: {overwritten}')
    print(f'Imágenes únicas finales: {unique_images}')
    print(f'Dataset final: {target_folder}')
    return 0

def handle_summary_dataset_command(args):
    """Generar CSV resumen de un dataset:

    Requisitos del usuario:
      - Preguntar qué dataset (si no se pasa --dataset).
      - Generar CSV en la raíz del dataset llamado _summary_{dataset}.csv
      - Columnas: transaction_id, device_id, transaction_start_time, Date, clase, label_id, filename, file_name_path
      - Rellenar transaction_id, device_id, transaction_start_time, Date desde Elastic.
      - clase, label_id, filename, file_name_path se obtienen del sistema de carpetas / nombre de fichero.
      - NO usar la hoja datasets_creados; inspeccionar directamente los ficheros existentes (puede haber borrados).
      - Para cruzar: extraer transaction_id del filename. Se asume que el nombre empieza por '<transaction_id>_' o contiene un número largo inicial antes de otro guion bajo.
    """
    import re
    import pandas as pd
    from zipfile import BadZipFile

    # 1. Localizar carpeta datasets_root igual que en merge
    datasets_root = None
    try:
        pa = ProcessAwareSettings(getattr(args, 'config', None))
        pinfo = pa.get_process_info()
        if pinfo and pinfo.get('process_dir'):
            candidate = Path(pinfo['process_dir']) / 'datasets'
            if candidate.exists():
                datasets_root = candidate
    except Exception:
        pass
    if datasets_root is None:
        candidate = Path('output') / 'processes'
        if candidate.exists():
            for proc_dir in sorted(candidate.iterdir()):
                dds = proc_dir / 'datasets'
                if dds.exists():
                    datasets_root = dds
                    break
    if datasets_root is None:
        print('No se encontró carpeta datasets')
        return 1

    # 2. Dataset target
    dataset_name = args.dataset
    available = [d for d in sorted(datasets_root.iterdir()) if d.is_dir() and not d.name.startswith('.')]
    if not available:
        print('No hay datasets disponibles')
        return 1
    if not dataset_name:
        print('Datasets disponibles:')
        for idx, d in enumerate(available, start=1):
            print(f"  {idx}. {d.name}")
        sel = input('Seleccione dataset (número): ').strip()
        if not sel.isdigit() or not (1 <= int(sel) <= len(available)):
            print('Selección no válida')
            return 1
        dataset_path = available[int(sel) - 1]
    else:
        dataset_path = next((d for d in available if d.name == dataset_name), None)
        if dataset_path is None:
            print(f"Dataset '{dataset_name}' no encontrado")
            return 1
    dataset_name = dataset_path.name

    # 3. Recolectar archivos de imagen y construir filas básicas
    img_exts = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
    records = []
    # Estructura esperada del nombre de fichero: <fecha>_<transaction id>_<original filename>
    # Ejemplo: 20240810T153045_123456789_nombreOriginal.jpg
    for class_dir in dataset_path.iterdir():
        if not class_dir.is_dir():
            continue
        for label_dir in class_dir.iterdir():
            if not label_dir.is_dir():
                continue
            for img_file in label_dir.iterdir():
                if not img_file.is_file() or img_file.suffix.lower() not in img_exts:
                    continue
                fname = img_file.name
                parts = fname.split('_', 2)
                # parts[0]=fecha, parts[1]=transaction_id, parts[2]=resto (original filename con posibles underscores)
                transaction_id = parts[1] if len(parts) >= 2 else ''
                # Ruta relativa al dataset sin incluir el nombre del fichero
                rel_folder = f"{class_dir.name}/{label_dir.name}"
                records.append({
                    'transaction_id': transaction_id,
                    'clase': class_dir.name,
                    'label_id': label_dir.name,
                    'filename': fname,
                    'file_name_path': rel_folder
                })

    if not records:
        print('No se encontraron imágenes en el dataset')
        return 1

    df_local = pd.DataFrame(records)
    # Filtrar transaction_ids no vacíos para consulta
    tx_ids = sorted(set([t for t in df_local['transaction_id'].tolist() if t]))
    print(f"Imágenes: {len(df_local)}, con transaction_id: {len(tx_ids)} únicos")

    # 4. Obtener datos ya descargados desde el Excel de reports (NO consultar Elasticsearch de nuevo)
    # Localizar Excel (similar a merge_datasets)
    excel_path = None
    try:
        pa = ProcessAwareSettings(getattr(args, 'config', None))
        pinfo = pa.get_process_info()
        if pinfo and pinfo.get('process_dir'):
            reports_dir = Path(pinfo['process_dir']) / 'reports'
            if reports_dir.exists():
                candidates = list(reports_dir.glob('*.xlsx'))
                if candidates:
                    proc_name = pinfo.get('name')
                    exact = [c for c in candidates if c.stem == proc_name]
                    excel_path = exact[0] if exact else candidates[0]
    except Exception:
        pass

    df_elastic = pd.DataFrame(columns=['transaction_id','device_id','transaction_start_time'])
    if excel_path and tx_ids:
        try:
            xls = pd.ExcelFile(excel_path)
            collected = []
            needed_cols = {'transaction_id','device_id','transaction_start_time'}
            for sheet in xls.sheet_names:
                try:
                    df_sheet = xls.parse(sheet, dtype=str)
                except Exception:
                    continue
                if needed_cols.issubset(set(df_sheet.columns)):
                    sub = df_sheet[list(needed_cols)].copy()
                    collected.append(sub)
            if collected:
                temp = pd.concat(collected, ignore_index=True)
                # Normalizar tipos a string y filtrar sólo los transaction_id de interés
                temp['transaction_id'] = temp['transaction_id'].astype(str)
                temp = temp[temp['transaction_id'].isin(tx_ids)]
                # Quitar duplicados priorizando filas con fecha válida
                if 'transaction_start_time' in temp.columns:
                    temp['transaction_start_time_parsed'] = pd.to_datetime(temp['transaction_start_time'], errors='coerce')
                    temp = temp.sort_values(by='transaction_start_time_parsed', ascending=False)
                temp = temp.drop_duplicates(subset=['transaction_id'], keep='first')
                df_elastic = temp[['transaction_id','device_id','transaction_start_time']].copy()
        except Exception as e:
            print(f'Aviso: no se pudieron extraer datos de Excel existente ({e}). Se continuá sólo con info local.')

    # Normalizar columnas claves
    for col in ['transaction_id','device_id','transaction_start_time']:
        if col not in df_elastic.columns:
            df_elastic[col] = ''
    # Convertir transaction_start_time a formato y Date
    if 'transaction_start_time' in df_elastic.columns and not df_elastic.empty:
        df_elastic['transaction_start_time'] = pd.to_datetime(df_elastic['transaction_start_time'], errors='coerce')
        df_elastic['Date'] = df_elastic['transaction_start_time'].dt.date
        df_elastic['transaction_start_time'] = df_elastic['transaction_start_time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    if 'Date' not in df_elastic.columns:
        df_elastic['Date'] = ''

    # Evitar duplicados por transaction_id quedándonos con la primera ocurrencia (o la más reciente si hay fecha)
    if not df_elastic.empty:
        if 'transaction_start_time' in df_elastic.columns:
            df_elastic = df_elastic.sort_values(by='transaction_start_time', ascending=False)
        df_elastic = df_elastic.drop_duplicates(subset=['transaction_id'], keep='first')

    # Merge left para mantener todas las imágenes aunque no haya datos Elastic
    df_final = df_local.merge(df_elastic[['transaction_id','device_id','transaction_start_time','Date']], on='transaction_id', how='left', suffixes=('','_es'))

    # Orden de columnas requerido
    ordered_cols = ['transaction_id','device_id','transaction_start_time','Date','clase','label_id','filename','file_name_path']
    for c in ordered_cols:
        if c not in df_final.columns:
            df_final[c] = ''
    df_final = df_final[ordered_cols]

    # 5. Guardar CSV
    out_csv = dataset_path / f"_summary_{dataset_name}.csv"
    try:
        df_final.to_csv(out_csv, index=False, sep=';', encoding='utf-8-sig')
        print(f'Resumen generado: {out_csv} ({len(df_final)} filas)')
    except Exception as e:
        print(f'Error guardando CSV: {e}')
        return 1
    return 0

def handle_register_dataset_command(args):
    """Registrar un dataset:
      1. Seleccionar dataset (interactivo si no se pasa --dataset).
      2. Comprimirlo a ZIP.
      3. Subir a S3 (base s3://grabit-data/datasets/ configurable). Si existe preguntar sobreescritura.
      4. Añadir/actualizar fila en CSV de registro global con columnas:
         dataset_name, created_at (dd/mm/YYYY), deployment_id, trained_model, description, s3_path
         - deployment_id se extrae del proceso activo.
         - trained_model y description se piden por pantalla.
         - Si ya existe fila para dataset_name preguntar si sobrescribir.
      5. Crear carpeta local junto al CSV con nombre del dataset y copiar summary si existe.
    """
    import shutil
    import csv as _csv
    import boto3
    from botocore.exceptions import ClientError
    from datetime import datetime as _dt
    import os
    import threading
    from zipfile import ZipFile, ZIP_DEFLATED
    from tqdm import tqdm
    import pandas as pd

    # 1. Localizar datasets_root
    datasets_root = None
    try:
        pa = ProcessAwareSettings(getattr(args, 'config', None))
        pinfo = pa.get_process_info()
        if pinfo and pinfo.get('process_dir'):
            candidate = Path(pinfo['process_dir']) / 'datasets'
            if candidate.exists():
                datasets_root = candidate
    except Exception:
        pass
    if datasets_root is None:
        candidate = Path('output') / 'processes'
        if candidate.exists():
            for proc_dir in sorted(candidate.iterdir()):
                dds = proc_dir / 'datasets'
                if dds.exists():
                    datasets_root = dds
                    break
    if datasets_root is None:
        print('No se encontró carpeta datasets')
        return 1

    # Datasets disponibles
    available = [d for d in sorted(datasets_root.iterdir()) if d.is_dir() and not d.name.startswith('.')]
    if not available:
        print('No hay datasets para registrar')
        return 1

    dataset_name = args.dataset
    if not dataset_name:
        print('Datasets disponibles:')
        for idx, d in enumerate(available, start=1):
            print(f"  {idx}. {d.name}")
        sel = input('Seleccione dataset (número): ').strip()
        if not sel.isdigit() or not (1 <= int(sel) <= len(available)):
            print('Selección no válida')
            return 1
        dataset_path = available[int(sel) - 1]
    else:
        dataset_path = next((d for d in available if d.name == dataset_name), None)
        if dataset_path is None:
            print(f"Dataset '{dataset_name}' no encontrado")
            return 1
    dataset_name = dataset_path.name

    # 2. Comprimir a ZIP en temporal con progreso
    zip_name = f"{dataset_name}.zip"
    tmp_zip_path = Path('tmp_zip')
    tmp_zip_path.mkdir(exist_ok=True)
    local_zip = tmp_zip_path / zip_name
    recreate_zip = True
    if local_zip.exists():
        ans = input(f'ZIP local {local_zip} ya existe. Regenerar? (s/N): ').strip().lower()
        recreate_zip = (ans == 's')
    if recreate_zip:
        if local_zip.exists():
            local_zip.unlink()
        # Recolectar archivos
        files = [p for p in dataset_path.rglob('*') if p.is_file()]
        print(f'Creando ZIP {local_zip} con {len(files)} ficheros ...')
        with ZipFile(local_zip, 'w', compression=ZIP_DEFLATED) as zf:
            for f in tqdm(files, desc='Zipping', unit='file'):
                # arcname incluye la carpeta raíz del dataset
                arcname = f.relative_to(dataset_path.parent)
                zf.write(f, arcname=str(arcname))
        if not local_zip.exists():
            print('Error creando ZIP')
            return 1
    else:
        print('Se reutiliza ZIP existente.')

    # 3. Subir a S3
    s3_base = args.s3_base or 's3://grabit-data/datasets/'
    if not s3_base.endswith('/'):
        s3_base += '/'
    # Parse s3 path
    if not s3_base.startswith('s3://'):
        print('Ruta S3 base inválida (debe empezar por s3://)')
        return 1
    _, s3_rest = s3_base.split('s3://', 1)
    bucket, *prefix_parts = s3_rest.split('/')
    prefix = '/'.join([p for p in prefix_parts if p])
    object_key = f"{prefix + '/' if prefix else ''}{zip_name}"
    s3_uri = f"s3://{bucket}/{object_key}"

    s3 = boto3.client('s3')
    # Verificar existencia
    exists = False
    try:
        s3.head_object(Bucket=bucket, Key=object_key)
        exists = True
    except ClientError:
        exists = False
    do_upload = True
    if exists:
        ans = input(f'El fichero {s3_uri} ya existe en S3. Sobrescribir? (s/N): ').strip().lower()
        if ans != 's':
            do_upload = False
            print('Se omite subida a S3.')
    if do_upload:
        class ProgressPercentage:
            def __init__(self, filename):
                self._filename = filename
                self._size = float(os.path.getsize(filename))
                self._seen_so_far = 0
                self._lock = threading.Lock()
            def __call__(self, bytes_amount):
                with self._lock:
                    self._seen_so_far += bytes_amount
                    percentage = (self._seen_so_far / self._size) * 100 if self._size else 100
                    print(f"\rSubiendo {self._filename}: {self._seen_so_far:,.0f}/{self._size:,.0f} bytes ({percentage:5.1f}%)", end='')
        print(f'Subiendo a {s3_uri} ...')
        s3.upload_file(str(local_zip), bucket, object_key, Callback=ProgressPercentage(str(local_zip)))
        print('\nSubida completada')
    else:
        print('Usando artefacto existente en S3.')

    # 4. Registrar en CSV global
    registry_csv = args.registry_csv or r'C:\Users\AlbertoGómez\Grabit\Operaciones - General\Datasets\datasets_gfresh.csv'
    registry_path = Path(registry_csv)
    registry_path.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    header = ['dataset_name','created_at','deployment_id','trained_model','description','s3_path']
    if registry_path.exists():
        parsed = False
        # Intentar con ';'
        try:
            df_csv = pd.read_csv(registry_path, sep=';', dtype=str, encoding='utf-8-sig')
            if set(header).issubset(df_csv.columns):
                rows = df_csv.fillna('').to_dict(orient='records')
                parsed = True
        except Exception:
            pass
        if not parsed:
            try:
                df_csv = pd.read_csv(registry_path, sep='\t', dtype=str, encoding='utf-8-sig')
                if set(header).issubset(df_csv.columns):
                    rows = df_csv.fillna('').to_dict(orient='records')
                    parsed = True
            except Exception:
                pass
        if not parsed:
            # Intentar reconstruir si única columna con cabecera concatenada
            try:
                with registry_path.open('r', encoding='utf-8-sig') as f:
                    first = f.readline().strip()
                if ';' in first:
                    cols = first.split(';')
                    if set(header).issubset(cols):
                        df_csv = pd.read_csv(registry_path, sep=';', dtype=str, encoding='utf-8-sig')
                        rows = df_csv.fillna('').to_dict(orient='records')
                        parsed = True
            except Exception:
                pass
        if not parsed:
            print('Aviso: no se pudo parsear correctamente el CSV existente; se reescribirá usando cabecera estándar.')

    # Extraer deployment_id del proceso activo
    deployment_id = ''
    try:
        pa = ProcessAwareSettings(getattr(args, 'config', None))
        pinfo = pa.get_process_info()
        deployment_id = str(pinfo.get('metadata', {}).get('deployment_id','')) if pinfo else ''
    except Exception:
        pass

    trained_model = input('ID de trained_model (ENTER si no aplica): ').strip()
    description = input('Descripción: ').strip()
    created_at = _dt.now().strftime('%d/%m/%Y')

    # Verificar si ya existe
    existing_idx = next((i for i,r in enumerate(rows) if r.get('dataset_name') == dataset_name), None)
    if existing_idx is not None:
        ans = input('Registro ya existe en CSV. Sobrescribir? (s/N): ').strip().lower()
        if ans != 's':
            print('Registro no actualizado por el usuario.')
        else:
            rows[existing_idx] = {
                'dataset_name': dataset_name,
                'created_at': created_at,
                'deployment_id': deployment_id,
                'trained_model': trained_model,
                'description': description,
                's3_path': s3_uri
            }
            print('Registro sobrescrito.')
    else:
        rows.append({
            'dataset_name': dataset_name,
            'created_at': created_at,
            'deployment_id': deployment_id,
            'trained_model': trained_model,
            'description': description,
            's3_path': s3_uri
        })
        print('Registro añadido.')

    # Guardar CSV (usar tab como en ejemplo). Si quieres ; cambia delimiter.
    # Intentar guardar con reintentos si está abierto en Excel
    max_retries = 3
    for attempt in range(1, max_retries+1):
        try:
            with registry_path.open('w', encoding='utf-8-sig', newline='') as f:
                writer = _csv.DictWriter(f, fieldnames=header, delimiter=';')
                writer.writeheader()
                for r in rows:
                    clean = {k: r.get(k,'') for k in header}
                    writer.writerow(clean)
            print(f'CSV actualizado: {registry_path}')
            break
        except PermissionError:
            if attempt == max_retries:
                print('No se pudo escribir el CSV tras varios intentos (permiso denegado). Se omite actualización.')
            else:
                ans = input(f"El fichero '{registry_path}' parece abierto. Ciérralo y pulsa ENTER para reintentar (o 'c' para cancelar): ").strip().lower()
                if ans == 'c':
                    print('Operación de actualización de CSV cancelada por el usuario.')
                    break
                continue
        except Exception as e:
            print(f'Error guardando CSV: {e}')
            break

    # 5. Crear carpeta local y copiar summary
    if not args.skip_summary_copy:
        registry_dir = registry_path.parent / dataset_name
        registry_dir.mkdir(parents=True, exist_ok=True)
        summary_file = dataset_path / f"_summary_{dataset_name}.csv"
        if summary_file.exists():
            dest_summary = registry_dir / '_summary.csv'
            shutil.copy2(summary_file, dest_summary)
            print(f'Summary copiado a {dest_summary}')
        else:
            print('No se encontró summary para copiar (ejecute summary_dataset si lo necesita).')

    # 6. Subir CSV de datasets completo a S3 (datasets.csv) opcional
    try:
        ans_upload_registry = input('¿Actualizar datasets.csv en S3? (s/N): ').strip().lower()
        if ans_upload_registry == 's':
            # Reusar s3_base (ya normalizado antes)
            s3_base2 = s3_base or 's3://grabit-data/datasets/'
            if not s3_base2.endswith('/'):
                s3_base2 += '/'
            _, s3_rest2 = s3_base2.split('s3://', 1)
            bucket2, *prefix_parts2 = s3_rest2.split('/')
            prefix2 = '/'.join([p for p in prefix_parts2 if p])
            registry_key = f"{prefix2 + '/' if prefix2 else ''}datasets.csv"
            print(f'Subiendo registro a s3://{bucket2}/{registry_key} ...')
            size_csv = registry_path.stat().st_size
            class ProgressCSV:
                def __init__(self, filename):
                    self._filename = filename
                    self._size = float(size_csv)
                    self._seen = 0
                    self._lock = threading.Lock()
                def __call__(self, bytes_amount):
                    with self._lock:
                        self._seen += bytes_amount
                        pct = (self._seen / self._size * 100) if self._size else 100
                        print(f"\rSubiendo datasets.csv: {self._seen:,.0f}/{self._size:,.0f} bytes ({pct:5.1f}%)", end='')
            s3.upload_file(str(registry_path), bucket2, registry_key, Callback=ProgressCSV(str(registry_path)))
            print('\nRegistro datasets.csv actualizado en S3.')
        else:
            print('No se actualiza datasets.csv en S3.')
    except Exception as e:
        print(f'Aviso: error subiendo datasets.csv a S3: {e}')

    print('Registro de dataset completado.')
    return 0
