import logging
from pathlib import Path
from typing import List

try:
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    OPENPYXL_AVAILABLE = False


def ensure_dropdown_validations(excel_path: str,
                                opciones_si_no: List[str] = None,
                                opciones_devices: List[str] = None,
                                max_rows: int = 5000) -> None:
    """Reaplica (idempotente) las validaciones de dropdown requeridas.

    Crea/actualiza la hoja 'Validacion' y añade data validations en las hojas
    'References' y 'Devices' para las columnas clave tras operaciones que
    hayan podido eliminar validaciones (p.ej. reescritura de hojas).

    Args:
        excel_path: Ruta al archivo Excel existente.
        opciones_si_no: Lista para columnas si/no (default ['si','no']).
        opciones_devices: Lista para devices_incluir (default ['todos','devices_seleccionados']).
        max_rows: Rango máximo a cubrir con validaciones.
    """
    logger = logging.getLogger(__name__)
    if not OPENPYXL_AVAILABLE:
        logger.debug("openpyxl no disponible; skip ensure_dropdown_validations")
        return
    path = Path(excel_path)
    if not path.exists():
        logger.debug(f"Archivo Excel no existe aún: {excel_path}")
        return

    opciones_si_no = opciones_si_no or ['si', 'no']
    opciones_devices = opciones_devices or ['todos', 'devices_seleccionados']

    try:
        wb = openpyxl.load_workbook(excel_path)

        # Crear / limpiar hoja de validación
        if 'Validacion' in wb.sheetnames:
            val_sheet = wb['Validacion']
            # limpiar filas
            for _ in range(val_sheet.max_row):
                val_sheet.delete_rows(1)
        else:
            val_sheet = wb.create_sheet('Validacion')

        # Escribir listas
        val_sheet['A1'] = 'Opciones_Devices_Incluir'
        for i, opt in enumerate(opciones_devices, start=2):
            val_sheet[f'A{i}'] = opt
        val_sheet['B1'] = 'Opciones_Si_No'
        for i, opt in enumerate(opciones_si_no, start=2):
            val_sheet[f'B{i}'] = opt
        val_sheet.sheet_state = 'hidden'

        def _apply_list_validation(ws, column_name: str, list_ref: str):
            headers = [c.value for c in ws[1]] if ws.max_row else []
            if column_name not in headers:
                return
            col_idx = headers.index(column_name) + 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            dv = DataValidation(type="list", formula1=list_ref, allow_blank=False)
            dv.error = 'Valor no válido'; dv.errorTitle = 'Validación'
            dv.add(f"{col_letter}2:{col_letter}{max_rows}")
            ws.add_data_validation(dv)

        if 'References' in wb.sheetnames:
            ws_ref = wb['References']
            _apply_list_validation(ws_ref, 'devices_incluir', f"Validacion!$A$2:$A${len(opciones_devices)+1}")
            _apply_list_validation(ws_ref, 'revisar_imagenes', f"Validacion!$B$2:$B${len(opciones_si_no)+1}")
            _apply_list_validation(ws_ref, 'incluir_dataset', f"Validacion!$B$2:$B${len(opciones_si_no)+1}")
        if 'Devices' in wb.sheetnames:
            ws_dev = wb['Devices']
            _apply_list_validation(ws_dev, 'revisar_imagenes', f"Validacion!$B$2:$B${len(opciones_si_no)+1}")
            _apply_list_validation(ws_dev, 'incluir_dataset', f"Validacion!$B$2:$B${len(opciones_si_no)+1}")

        wb.save(excel_path)
        logger.info("Validaciones de dropdown reforzadas (ensure_dropdown_validations)")
    except Exception as e:  # pragma: no cover
        logger.warning(f"No se pudieron reaplicar validaciones: {e}")
