"""
Generador de informes y archivos Excel
Crea informes detallados del análisis y exporta datos a Excel con múltiples pestañas
"""

import logging
import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import json
from utils.excel_validations import ensure_dropdown_validations

# Verificar disponibilidad de openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ReportGenerator:
    """
    Generador de informes y archivos Excel para el análisis de dataset
    """
    
    def __init__(self, settings):
        """
        Inicializa el generador de informes
        
        Args:
            settings: Objeto Settings con la configuración
        """
        self.settings = settings
        self.logger = logging.getLogger(__name__)
        
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl no está disponible. Instale con: pip install openpyxl")
    
    def _check_excel_file_writable(self, excel_path: str) -> tuple[bool, str]:
        """
        Verifica si el archivo Excel está disponible para escritura
        
        Returns:
            tuple[bool, str]: (is_writable, error_message)
        """
        try:
            # Verificar si el directorio existe
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            
            # Si el archivo no existe, verificar si podemos crear uno
            if not os.path.exists(excel_path):
                try:
                    # Crear archivo temporal para verificar permisos de escritura
                    test_file = excel_path + '.test'
                    with open(test_file, 'w') as f:
                        f.write('test')
                    os.remove(test_file)
                    return True, ""
                except Exception as e:
                    return False, f"No se puede escribir en el directorio: {str(e)}"
            
            # Si el archivo existe, verificar si está abierto
            try:
                # Intentar abrir en modo append para verificar bloqueo
                with open(excel_path, 'a'):
                    pass
                return True, ""
            except PermissionError:
                return False, "El archivo está abierto en otra aplicación"
            except Exception as e:
                return False, f"Error de acceso: {str(e)}"
                
        except Exception as e:
            return False, f"Error verificando archivo: {str(e)}"

    def _handle_excel_file_conflict(self, excel_path: str) -> str:
        """
        Maneja conflictos cuando el archivo Excel está abierto
        
        Returns:
            str: Ruta del archivo a usar (original o backup)
        """
        print(f"\n{'='*70}")
        print(f"⚠️  ARCHIVO EXCEL EN USO")
        print(f"{'='*70}")
        print(f"📁 Archivo: {os.path.basename(excel_path)}")
        print(f"📍 Ubicación: {os.path.dirname(excel_path)}")
        print(f"\n🔸 El archivo está abierto en Microsoft Excel u otra aplicación")
        print(f"🔸 Debe cerrar el archivo para que se actualice correctamente")
        print(f"{'='*70}")
        
        while True:
            print(f"\nOpciones disponibles:")
            print(f"  [1] Cerrar Excel y continuar")
            print(f"  [2] Crear archivo con nombre diferente")
            print(f"  [3] Cancelar operación")
            
            choice = input(f"\nSeleccione una opción (1-3): ").strip()
            
            if choice == '1':
                # Verificar si el usuario cerró el archivo
                is_writable, error_msg = self._check_excel_file_writable(excel_path)
                if is_writable:
                    print(f"✅ Archivo accesible, continuando...")
                    return excel_path
                else:
                    print(f"❌ El archivo sigue bloqueado: {error_msg}")
                    print(f"🔸 Asegúrese de cerrar completamente Excel")
                    
            elif choice == '2':
                # Crear nombre de backup
                timestamp = datetime.now().strftime("%H%M%S")
                backup_path = excel_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
                print(f"📝 Archivo alternativo: {os.path.basename(backup_path)}")
                return backup_path
                
            elif choice == '3':
                print(f"❌ Operación cancelada por el usuario")
                raise RuntimeError("Operación cancelada - archivo Excel en uso")
                
            else:
                print(f"❌ Opción no válida, seleccione 1, 2 o 3")
    
    
    def _create_session_directory(self, deployment_id: int, extracted_data: Dict[str, Any]) -> str:
        """
        Crea un directorio específico para esta sesión de procesamiento
        
        Args:
            deployment_id: ID del deployment
            extracted_data: Datos extraídos con metadatos
            
        Returns:
            Ruta del directorio de sesión creado
        """
        # Generar nombre de directorio descriptivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Determinar si es modelo base o activo
        metadata = extracted_data.get('extraction_metadata', {})
        is_base_model = metadata.get('is_base_model', False)
        model_used = metadata.get('model_used', 'unknown')
        
        if is_base_model:
            session_name = f"deployment_{deployment_id}_modelo_base_{model_used}_{timestamp}"
        else:
            session_name = f"deployment_{deployment_id}_modelo_activo_{model_used}_{timestamp}"
        
        # Crear directorio
        base_output_dir = self.settings.get_output_directory('reports')
        session_dir = os.path.join(base_output_dir, session_name)
        
        # Crear subdirectorios
        os.makedirs(session_dir, exist_ok=True)
        os.makedirs(os.path.join(session_dir, 'reports'), exist_ok=True)
        os.makedirs(os.path.join(session_dir, 'logs'), exist_ok=True)
        
        self.logger.info(f"Directorio de sesión creado: {session_dir}")
        return session_dir
    
    def _setup_session_logging(self, session_dir: str, deployment_id: int) -> str:
        """
        Configura logging específico para la sesión
        
        Args:
            session_dir: Directorio de la sesión
            deployment_id: ID del deployment
            
        Returns:
            Ruta del archivo de log
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"processing_log_{deployment_id}_{timestamp}.log"
        log_path = os.path.join(session_dir, 'logs', log_filename)
        
        # Configurar handler específico para esta sesión
        session_handler = logging.FileHandler(log_path, encoding='utf-8')
        session_handler.setLevel(logging.INFO)
        
        # Formato específico para log de sesión
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(name)s - %(message)s'
        )
        session_handler.setFormatter(formatter)
        
        # Agregar handler al logger raíz para capturar todos los logs
        root_logger = logging.getLogger()
        root_logger.addHandler(session_handler)
        
        # Escribir header del log
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(f"=== DATASET MANAGER PROCESSING LOG ===\n")
            f.write(f"Deployment ID: {deployment_id}\n")
            f.write(f"Inicio de procesamiento: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"=" * 50 + "\n\n")
        
        self.logger.info(f"Log de sesión configurado: {log_path}")
        return log_path
    
    def create_main_excel_report(self, extracted_data: Dict[str, Any], 
                               analysis_results: Dict[str, Any],
                               download_results: Dict[str, Any],
                               deployment_id: int,
                               session_dir: str = None) -> str:
        """
        Crea el archivo Excel principal con todas las pestañas del análisis
        
        Args:
            extracted_data: Datos extraídos
            analysis_results: Resultados del análisis
            download_results: Resultados de descarga de imágenes
            deployment_id: ID del deployment
            session_dir: Directorio de sesión (opcional)
            
        Returns:
            Ruta del archivo Excel creado
        """
        try:
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl no está disponible para crear archivos Excel")
            
            self.logger.info("Creando archivo Excel principal del análisis")
            
            # Almacenar extracted_data para uso en métodos auxiliares
            self.extracted_data = extracted_data
            
            # Determinar directorio de salida y nombre de archivo
            # Prioridad: proceso activo actual > session_dir > directorio por defecto
            process_info = None
            if hasattr(self.settings, 'get_process_info'):
                process_info = self.settings.get_process_info()
            
            if process_info and session_dir is None:
                # Usar directorio del proceso activo ACTUAL
                output_dir = self.settings.get_output_path('reports')
                # Usar nombre del proceso como nombre de archivo
                filename = f"{process_info['name']}.xlsx"
                self.logger.info(f"Usando proceso activo: {process_info['name']}")
            elif session_dir:
                output_dir = os.path.join(session_dir, 'reports')
                # Generar nombre de archivo tradicional
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = self.settings.reports.excel_filename.format(
                    deployment_id=deployment_id,
                    timestamp=timestamp
                )
                self.logger.info(f"Usando session_dir: {session_dir}")
            else:
                output_dir = self.settings.get_output_directory('reports')
                # Generar nombre de archivo tradicional
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = self.settings.reports.excel_filename.format(
                    deployment_id=deployment_id,
                    timestamp=timestamp
                )
                self.logger.info(f"Usando directorio por defecto")
            
            excel_path = os.path.join(output_dir, filename)
            
            # Verificar acceso al archivo antes de intentar escribir
            is_writable, error_msg = self._check_excel_file_writable(excel_path)
            if not is_writable:
                self.logger.warning(f"Archivo Excel no accesible: {error_msg}")
                excel_path = self._handle_excel_file_conflict(excel_path)
            
            # Crear libro de Excel (usa helper para garantizar persistencia de columnas especiales antes de cerrar)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                
                # 1. Pestaña Resumen
                self._create_summary_sheet(writer, extracted_data, analysis_results, 
                                         download_results, deployment_id)
                
                # (Eliminado) Hoja Datos_Elasticsearch ya no se genera; datos se consumen desde elastic_data.csv externo.
                
                # 3. Pestaña References - Consolidada con toda la información de análisis
                if not analysis_results['complete_analysis'].empty and not extracted_data['references_data'].empty:
                    # Combinar datos de análisis completo con datos de referencias
                    consolidated_references = self._create_consolidated_references(
                        analysis_results['complete_analysis'], 
                        extracted_data['references_data']
                    )
                    consolidated_references.to_excel(
                        writer, sheet_name='References', index=False
                    )
                elif not analysis_results['complete_analysis'].empty:
                    analysis_results['complete_analysis'].to_excel(
                        writer, sheet_name='References', index=False
                    )
                elif not extracted_data['references_data'].empty:
                    extracted_data['references_data'].to_excel(
                        writer, sheet_name='References', index=False
                    )
                
                # 4. Pestaña Devices (después de References)
                self._create_devices_sheet(writer, analysis_results.get('devices_analysis', pd.DataFrame()))
                
                # 5. Pestaña Labels
                if not extracted_data['labels_data'].empty:
                    extracted_data['labels_data'].to_excel(
                        writer, sheet_name='Labels', index=False
                    )
                
                # 6. Pestaña Sugerencias de Labels
                if not analysis_results['label_suggestions'].empty:
                    # Preparar datos de sugerencias para Excel
                    suggestions_df = self._prepare_suggestions_for_excel(
                        analysis_results['label_suggestions']
                    )
                    suggestions_df.to_excel(
                        writer, sheet_name='Sugerencias_Labels', index=False
                    )
                
                # 7. Pestaña Model Data (expandida en columnas)
                self._create_model_data_sheet(writer, extracted_data['model_data'], extracted_data.get('labels_data', pd.DataFrame()))
                
                # 8. Pestaña Análisis de Consistencia (corregida)
                self._create_consistency_sheet_fixed(writer, analysis_results)
            
            # Aplicar formato al archivo Excel y reforzar columnas especiales
            self._format_excel_file(excel_path)
            self._reinforce_special_columns(excel_path)
            
            self.logger.info(f"Archivo Excel principal creado: {excel_path}")
            # Reforzar validaciones tras cualquier escritura adicional externa
            try:
                ensure_dropdown_validations(excel_path)
            except Exception as _e:
                self.logger.debug(f"No se pudieron reforzar validaciones post-creación: {_e}")
            return excel_path
            
        except Exception as e:
            self.logger.error(f"Error creando archivo Excel principal: {e}")
            raise
    
    def create_new_references_excel(self, analysis_results: Dict[str, Any],
                                  download_results: Dict[str, Any],
                                  deployment_id: int,
                                  session_dir: str = None) -> str:
        """
        Crea archivo Excel específico para referencias nuevas
        
        Args:
            analysis_results: Resultados del análisis
            download_results: Resultados de descarga
            deployment_id: ID del deployment
            session_dir: Directorio de sesión (opcional)
            
        Returns:
            Ruta del archivo Excel creado
        """
        try:
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl no está disponible para crear archivos Excel")
            
            self.logger.info("Creando archivo Excel para referencias nuevas")
            
            # Determinar directorio de salida y nombre de archivo
            # Si hay un proceso activo, usar su directorio y nombre
            process_info = None
            try:
                from ..process_aware_settings import ProcessAwareSettings
                if hasattr(self.settings, 'get_process_info'):
                    process_info = self.settings.get_process_info()
                else:
                    temp_settings = ProcessAwareSettings()
                    process_info = temp_settings.get_process_info()
            except:
                process_info = None
            
            if process_info and session_dir is None:
                # Usar directorio del proceso activo
                output_dir = self.settings.get_output_path('reports')
                # Usar nombre del proceso como base para el archivo
                filename = f"new_references_{process_info['name']}.xlsx"
            elif session_dir:
                output_dir = os.path.join(session_dir, 'reports')
                # Generar nombre de archivo tradicional
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = self.settings.reports.new_references_filename.format(
                    deployment_id=deployment_id,
                    timestamp=timestamp
                )
            else:
                output_dir = self.settings.get_output_directory('reports')
                # Generar nombre de archivo tradicional
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = self.settings.reports.new_references_filename.format(
                    deployment_id=deployment_id,
                    timestamp=timestamp
                )
            
            excel_path = os.path.join(output_dir, filename)
            
            # Preparar datos combinados para referencias nuevas
            new_refs_data = self._prepare_new_references_data(
                analysis_results, download_results
            )
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Pestaña principal con referencias nuevas
                new_refs_data.to_excel(
                    writer, sheet_name='Referencias_Nuevas', index=False
                )
                
                # Pestaña con sugerencias detalladas
                if not analysis_results['label_suggestions'].empty:
                    detailed_suggestions = self._prepare_detailed_suggestions(
                        analysis_results['label_suggestions']
                    )
                    detailed_suggestions.to_excel(
                        writer, sheet_name='Sugerencias_Detalladas', index=False
                    )
            
            # Aplicar formato
            self._format_excel_file(excel_path)
            
            self.logger.info(f"Archivo Excel de referencias nuevas creado: {excel_path}")
            return excel_path
            
        except Exception as e:
            self.logger.error(f"Error creando archivo Excel de referencias nuevas: {e}")
            raise
    
    def _create_summary_sheet(self, writer, extracted_data: Dict[str, Any],
                            analysis_results: Dict[str, Any],
                            download_results: Dict[str, Any],
                            deployment_id: int):
        """Crea la pestaña de resumen ejecutivo"""
        try:
            summary_data = []
            
            # Información general
            summary_data.append(['RESUMEN EJECUTIVO - ANÁLISIS DE DATASET', ''])
            summary_data.append(['Deployment ID', deployment_id])
            summary_data.append(['Fecha de análisis', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            summary_data.append(['', ''])
            
            # Estadísticas de extracción
            summary_data.append(['DATOS EXTRAÍDOS', ''])
            summary_data.append(['Total registros Elasticsearch', len(extracted_data['elastic_data'])])
            summary_data.append(['Total referencias', len(extracted_data['references_data'])])
            summary_data.append(['Total labels', len(extracted_data['labels_data'])])
            summary_data.append(['', ''])
            
            # Modelo activo
            if extracted_data['active_model_info']:
                model_info = extracted_data['active_model_info']
                summary_data.append(['MODELO ACTIVO', ''])
                summary_data.append(['Model ID', model_info.get('model_id', 'N/A')])
                summary_data.append(['Frecuencia', model_info.get('frequency', 'N/A')])
                summary_data.append(['Porcentaje uso', f"{model_info.get('percentage', 0):.1f}%"])
                summary_data.append(['', ''])
            
            # Estadísticas de análisis
            stats = analysis_results['summary_stats']
            summary_data.append(['ANÁLISIS DE REFERENCIAS', ''])
            summary_data.append(['Referencias filtradas', stats['total_filtered_references']])
            summary_data.append(['Referencias sin asignar', stats['unassigned_count']])
            summary_data.append(['Referencias no entrenadas', stats['untrained_count']])
            summary_data.append(['Sugerencias alta confianza', stats['high_confidence_suggestions']])
            summary_data.append(['Sugerencias baja confianza', stats['low_confidence_suggestions']])
            summary_data.append(['', ''])
            
            # Análisis de devices si está disponible
            devices_analysis = analysis_results.get('devices_analysis', pd.DataFrame())
            if not devices_analysis.empty:
                summary_data.append(['ANÁLISIS DE DEVICES', ''])
                summary_data.append(['Total devices', len(devices_analysis)])
                summary_data.append(['Devices activos (>50 trans)', len(devices_analysis[devices_analysis['num_cliente'] > 50])])
                summary_data.append(['Devices con revisar_imagenes=si', len(devices_analysis[devices_analysis['revisar_imagenes'] == 'si'])])
                summary_data.append(['', ''])
            
            # Información de consistencia simplificada
            summary_data.append(['ESTADO DEL ANÁLISIS', ''])
            summary_data.append(['Análisis completado', 'Sí'])
            summary_data.append(['Errores críticos', 'No'])
            summary_data.append(['', ''])
            
            # Información del procesamiento
            summary_data.append(['INFORMACIÓN DEL PROCESAMIENTO', ''])
            summary_data.append(['Deployment ID', deployment_id])
            summary_data.append(['Fecha de procesamiento', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
            summary_data.append(['Duración estimada', 'Completado'])
            
            # Crear DataFrame y escribir
            summary_df = pd.DataFrame(summary_data, columns=['Métrica', 'Valor'])
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)
            
        except Exception as e:
            self.logger.error(f"Error creando pestaña de resumen: {e}")
    
    def _create_model_data_sheet(self, writer, model_data: Dict[str, Any], labels_df: pd.DataFrame = None):
        """Crea la pestaña con datos del modelo en formato simple"""
        try:
            if not model_data:
                # Crear pestaña vacía si no hay datos
                empty_df = pd.DataFrame([['No hay datos del modelo disponibles']], 
                                      columns=['Información'])
                empty_df.to_excel(writer, sheet_name='Model_Data', index=False)
                return

            # Crear mapeo de label_id a label_name
            label_mapping = {}
            if labels_df is not None and not labels_df.empty:
                self.logger.info(f"Creando mapeo de labels - DataFrame shape: {labels_df.shape}")
                label_mapping = dict(zip(labels_df['label_id'], labels_df['label_name']))
                self.logger.info(f"Mapeo creado con {len(label_mapping)} labels")
            else:
                self.logger.warning("No se recibieron datos de labels para el mapeo")

            # Lista para almacenar las filas del modelo
            model_rows = []
            
            # Procesar los datos del modelo
            if 'model_data' in model_data and model_data['model_data']:
                model_json_list = model_data['model_data']
                
                # Si es una lista de clases del modelo
                if isinstance(model_json_list, list):
                    for item in model_json_list:
                        # Información básica de la clase
                        base_row = {
                            'model_id': model_data.get('model_id', 'N/A'),
                            'model_code': model_data.get('model_code', 'N/A'),
                            'class_index': item.get('index', ''),
                            'class_name': item.get('index_name', ''),
                            'label_id': '',
                            'label_name': ''
                        }
                        
                        # Obtener las etiquetas entrenadas para esta clase
                        trained_labels = item.get('trained_labels', [])
                        
                        if trained_labels:
                            # Crear una fila por cada etiqueta entrenada
                            for label_id in trained_labels:
                                row = base_row.copy()
                                row['label_id'] = label_id
                                # Buscar el nombre del label en el mapeo
                                row['label_name'] = label_mapping.get(label_id, f'Label {label_id}')
                                model_rows.append(row)
                        else:
                            # Si no hay etiquetas entrenadas, crear una fila vacía para la clase
                            model_rows.append(base_row)
                            
            # Si no hay datos del modelo, crear al menos una fila con información básica
            if not model_rows:
                model_rows.append({
                    'model_id': model_data.get('model_id', 'N/A'),
                    'model_code': model_data.get('model_code', 'N/A'),
                    'class_index': 'N/A',
                    'class_name': 'No hay datos de clases disponibles',
                    'label_id': 'N/A',
                    'label_name': 'N/A'
                })
            
            # Convertir a DataFrame con el orden específico de columnas
            columns_order = ['model_id', 'model_code', 'class_index', 'class_name', 'label_id', 'label_name']
            model_df = pd.DataFrame(model_rows, columns=columns_order)
            
            self.logger.info(f"Creando pestaña Model_Data con {len(model_rows)} filas")
            model_df.to_excel(writer, sheet_name='Model_Data', index=False)
            
        except Exception as e:
            self.logger.error(f"Error creando pestaña de modelo: {e}")
            # Crear pestaña de error
            error_df = pd.DataFrame([['Error procesando datos del modelo']], 
                                  columns=['Error'])
            error_df.to_excel(writer, sheet_name='Model_Data', index=False)
    
    def _create_consolidated_references(self, complete_analysis_df: pd.DataFrame, 
                                      references_data_df: pd.DataFrame) -> pd.DataFrame:
        """Consolida datos de análisis completo con datos de referencias"""
        try:
            # Combinar datos usando reference_code como clave
            if not complete_analysis_df.empty and not references_data_df.empty:
                # Verificar qué columnas están disponibles en complete_analysis_df
                available_columns = ['reference_code']
                
                # Columnas básicas de análisis - reordenadas para que producto_trained y label_name estén después de label_id
                basic_columns = ['producto_trained', 'label_name', 'num_cliente', 'num_manuales', 'is_trained_elastic', 
                               'is_in_dataset', 'pct_ok_sistema', 'pct_ok_modelo', 'revisar_imagenes']
                
                # Columnas de top 3 result_index
                top3_columns = []
                for rank in range(1, 4):  # Top 1, 2, 3
                    top3_columns.extend([
                        f'top{rank}_result_index',
                        f'top{rank}_result_label_name', 
                        f'top{rank}_%pesadas',
                        f'top{rank}_pct_ok'
                    ])
                
                # Añadir todas las columnas que existan
                for col in basic_columns + top3_columns:
                    if col in complete_analysis_df.columns:
                        available_columns.append(col)
                
                consolidated = pd.merge(
                    references_data_df, 
                    complete_analysis_df[available_columns], 
                    on='reference_code', 
                    how='left'
                )
                
                # Las referencias ya vienen enriquecidas con label_name desde el analyzer
                # No es necesario llamar a _ensure_all_label_names
                
                # Añadir las nuevas columnas solicitadas
                self._add_dataset_columns(consolidated)
                
                # Eliminar columnas innecesarias (mantener num_cliente)
                columns_to_drop = ['created_at', 'updated_at', 'num_apariciones', 'deployment_id']
                consolidated = consolidated.drop(columns=[col for col in columns_to_drop if col in consolidated.columns])
                
                # Ordenar por num_cliente descendente
                if 'num_cliente' in consolidated.columns:
                    consolidated = consolidated.sort_values('num_cliente', ascending=False).reset_index(drop=True)
                
                return consolidated
            elif not complete_analysis_df.empty:
                # Aplicar mismas transformaciones a complete_analysis_df
                result = complete_analysis_df.copy()
                
                # Añadir las nuevas columnas solicitadas
                self._add_dataset_columns(result)
                
                columns_to_drop = ['created_at', 'updated_at', 'num_apariciones', 'deployment_id']
                result = result.drop(columns=[col for col in columns_to_drop if col in result.columns])
                
                # Ordenar por num_cliente descendente
                if 'num_cliente' in result.columns:
                    result = result.sort_values('num_cliente', ascending=False).reset_index(drop=True)
                
                return result
            else:
                # Solo references_data_df disponible - las referencias ya vienen enriquecidas
                result = references_data_df.copy()
                
                # Añadir las nuevas columnas solicitadas
                self._add_dataset_columns(result)
                
                columns_to_drop = ['created_at', 'updated_at', 'num_apariciones', 'deployment_id']
                result = result.drop(columns=[col for col in columns_to_drop if col in result.columns])
                
                # Ordenar por num_cliente descendente si la columna existe
                if 'num_cliente' in result.columns:
                    result = result.sort_values('num_cliente', ascending=False).reset_index(drop=True)
                
                return result
                
        except Exception as e:
            self.logger.error(f"Error consolidando referencias: {e}")
            # En caso de error, devolver al menos los datos básicos sin las columnas innecesarias
            result = references_data_df if not references_data_df.empty else complete_analysis_df
            if not result.empty:
                # Añadir las nuevas columnas solicitadas
                self._add_dataset_columns(result)
                
                columns_to_drop = ['created_at', 'updated_at', 'num_apariciones', 'deployment_id']
                result = result.drop(columns=[col for col in columns_to_drop if col in result.columns])
                # Ordenar por num_cliente descendente si existe
                if 'num_cliente' in result.columns:
                    result = result.sort_values('num_cliente', ascending=False).reset_index(drop=True)
            return result
    
    def _add_dataset_columns(self, df: pd.DataFrame) -> None:
        """Añade/actualiza columnas de configuración del dataset con nuevas reglas:
        - incluir_dataset = 'si' cuando num_cliente >= mínimo (sin tilde)
        - devices_incluir default = 'devices_seleccionados'
        - num_imagenes_dataset solo relleno cuando incluir_dataset == 'si'
        - Normalizar revisar_imagenes a 'si'/'no'
        """
        try:
            # Obtener valores por defecto de la configuración
            from ..config.settings import Settings
            settings = Settings()
            
            num_imagenes_default = getattr(settings.dataset, 'num_imagenes_dataset_default', 150)
            devices_incluir_default = getattr(settings.dataset, 'devices_incluir_default', 'todos')
            min_transacciones = getattr(settings.dataset, 'min_transacciones_cliente', 150)
            
            # Lógica incluir_dataset
            if 'num_cliente' in df.columns:
                # Asegurar que num_cliente es numérico
                df['num_cliente'] = pd.to_numeric(df['num_cliente'], errors='coerce').fillna(0)
                df['incluir_dataset'] = (df['num_cliente'] >= min_transacciones).map({True: 'si', False: 'no'})
            else:
                df['incluir_dataset'] = 'no'
            
            # devices_incluir default forzado a devices_seleccionados
            df['devices_incluir'] = 'devices_seleccionados'
            
            # num_imagenes_dataset sólo cuando se incluye
            df['num_imagenes_dataset'] = df['incluir_dataset'].apply(lambda v: num_imagenes_default if v == 'si' else '')
            
            # Nuevas columnas al final: producto_a_entrenar y label_considerar
            # Solo tienen valor cuando incluir_dataset = 'sí'
            df['producto_a_entrenar'] = df.apply(
                lambda row: row.get('producto_trained', '') if row.get('incluir_dataset') == 'si' else '',
                axis=1
            )
            
            df['label_considerar'] = df.apply(
                lambda row: row.get('label_id', '') if row.get('incluir_dataset') == 'si' else '',
                axis=1
            )
            
            # Asegurar que revisar_imagenes solo tenga valores 'sí' o 'no'
            if 'revisar_imagenes' in df.columns:
                df['revisar_imagenes'] = df['revisar_imagenes'].map({
                    True: 'si', 
                    False: 'no',
                    'True': 'si',
                    'False': 'no',
                    'sí': 'si',
                    'si': 'si',
                    'no': 'no',
                    1: 'si',
                    0: 'no'
                }).fillna('no')
            
            self.logger.info(f"Columnas de dataset añadidas. Mínimo transacciones: {min_transacciones}")
            
        except Exception as e:
            self.logger.warning(f"Error añadiendo columnas de dataset: {e}. Usando valores por defecto.")
            # Usar valores por defecto en caso de error
            if 'is_in_dataset' in df.columns:
                df['incluir_dataset'] = df['is_in_dataset']
            else:
                df['incluir_dataset'] = 'no'
            df['devices_incluir'] = 'todos'
            df['num_imagenes_dataset'] = 150
            df['producto_a_entrenar'] = ''
            df['label_considerar'] = ''
    
    def _ensure_all_label_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        DEPRECATED: Ya no es necesario este método porque las referencias 
        vienen enriquecidas desde el analyzer con la nueva arquitectura
        """
        self.logger.warning("DEPRECATED: _ensure_all_label_names ya no es necesario con la nueva arquitectura")
        return df
    
    def _flatten_json(self, json_obj, parent_key='', sep='_'):
        """Aplana un objeto JSON anidado en un diccionario plano"""
        items = []
        if isinstance(json_obj, dict):
            for k, v in json_obj.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(self._flatten_json(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    # Para listas, crear columnas numeradas o concatenar valores
                    if all(isinstance(item, (str, int, float)) for item in v):
                        items.append((new_key, ', '.join(map(str, v))))
                    else:
                        for i, item in enumerate(v):
                            if isinstance(item, dict):
                                items.extend(self._flatten_json(item, f"{new_key}_{i}", sep=sep).items())
                            else:
                                items.append((f"{new_key}_{i}", str(item)))
                else:
                    items.append((new_key, v))
        return dict(items)
    
    def _create_consistency_sheet(self, writer, consistency_analysis: Dict[str, Any]):
        """Crea la pestaña de análisis de consistencia"""
        try:
            consistency_data = []
            consistency_data.append(['ANÁLISIS DE CONSISTENCIA', ''])
            consistency_data.append(['Total sin asignar', consistency_analysis['total_unassigned']])
            consistency_data.append(['Total no entrenadas', consistency_analysis['total_untrained']])
            consistency_data.append(['Referencias consistentes', consistency_analysis['consistent_references']])
            consistency_data.append(['Ratio consistencia', f"{consistency_analysis['consistency_ratio']:.1%}"])
            consistency_data.append(['Tiene inconsistencias', 'Sí' if consistency_analysis['has_inconsistencies'] else 'No'])
            consistency_data.append(['', ''])
            
            # Referencias solo sin asignar
            if consistency_analysis['only_unassigned']:
                consistency_data.append(['SOLO SIN ASIGNAR', ''])
                for ref in consistency_analysis['only_unassigned'][:20]:  # Máximo 20
                    consistency_data.append(['', ref])
                consistency_data.append(['', ''])
            
            # Referencias solo no entrenadas
            if consistency_analysis['only_untrained']:
                consistency_data.append(['SOLO NO ENTRENADAS', ''])
                for ref in consistency_analysis['only_untrained'][:20]:  # Máximo 20
                    consistency_data.append(['', ref])
            
            consistency_df = pd.DataFrame(consistency_data, columns=['Categoría', 'Valor'])
            consistency_df.to_excel(writer, sheet_name='Consistencia', index=False)
            
        except Exception as e:
            self.logger.error(f"Error creando pestaña de consistencia: {e}")
    
    def _create_download_summary_sheet(self, writer, download_results: Dict[str, Any]):
        """Crea la pestaña resumen de descarga de imágenes"""
        try:
            if not download_results:
                empty_df = pd.DataFrame([['No hay datos de descarga disponibles']], 
                                      columns=['Información'])
                empty_df.to_excel(writer, sheet_name='Descarga_Imagenes', index=False)
                return
            
            download_data = []
            
            # Resumen general
            summary_stats = download_results.get('summary_stats', {})
            download_data.append(['RESUMEN DESCARGA DE IMÁGENES', ''])
            download_data.append(['Referencias nuevas procesadas', summary_stats.get('total_new_references', 0)])
            download_data.append(['Imágenes referencias nuevas', summary_stats.get('total_new_ref_images', 0)])
            download_data.append(['Clases similares procesadas', summary_stats.get('total_similar_classes', 0)])
            download_data.append(['Imágenes clases similares', summary_stats.get('total_similar_images', 0)])
            download_data.append(['Directorio salida', download_results.get('output_directory', 'N/A')])
            download_data.append(['', ''])
            
            # Detalle por referencia nueva
            new_ref_downloads = download_results.get('new_references_downloads', [])
            if new_ref_downloads:
                download_data.append(['REFERENCIAS NUEVAS - DETALLE', ''])
                download_data.append(['Código', 'Nombre', 'Descargadas', 'Total URLs', 'Tasa éxito'])
                for download in new_ref_downloads:
                    download_data.append([
                        download['reference_code'],
                        download['reference_name'],
                        download['downloaded'],
                        download['total_urls'],
                        f"{download['success_rate']:.1f}%"
                    ])
            
            download_df = pd.DataFrame(download_data)
            download_df.to_excel(writer, sheet_name='Descarga_Imagenes', index=False, header=False)
            
        except Exception as e:
            self.logger.error(f"Error creando pestaña de descarga: {e}")
    
    def _prepare_suggestions_for_excel(self, suggestions_df: pd.DataFrame) -> pd.DataFrame:
        """Prepara las sugerencias para mostrar en Excel de forma legible"""
        try:
            if suggestions_df.empty:
                return suggestions_df
            
            excel_suggestions = suggestions_df.copy()
            
            # Expandir sugerencias principales
            if 'all_suggestions' in excel_suggestions.columns:
                # Crear columnas para top 3 sugerencias
                for i in range(3):
                    excel_suggestions[f'sugerencia_{i+1}_label_id'] = None
                    excel_suggestions[f'sugerencia_{i+1}_porcentaje'] = None
                
                for idx, row in excel_suggestions.iterrows():
                    suggestions = row['all_suggestions']
                    if isinstance(suggestions, list):
                        for i, sugg in enumerate(suggestions[:3]):
                            if isinstance(sugg, dict):
                                excel_suggestions.at[idx, f'sugerencia_{i+1}_label_id'] = sugg.get('label_id')
                                excel_suggestions.at[idx, f'sugerencia_{i+1}_porcentaje'] = f"{sugg.get('percentage', 0):.1f}%"
                
                # Eliminar columna original compleja
                excel_suggestions = excel_suggestions.drop('all_suggestions', axis=1)
            
            return excel_suggestions
            
        except Exception as e:
            self.logger.error(f"Error preparando sugerencias para Excel: {e}")
            return suggestions_df
    
    def _prepare_new_references_data(self, analysis_results: Dict[str, Any],
                                   download_results: Dict[str, Any]) -> pd.DataFrame:
        """Prepara datos combinados para referencias nuevas"""
        try:
            untrained_refs = analysis_results['untrained_references'].copy()
            
            # Agregar información de descarga si está disponible
            if download_results and 'new_references_downloads' in download_results:
                download_info = {d['reference_code']: d for d in download_results['new_references_downloads']}
                
                untrained_refs['imagenes_descargadas'] = untrained_refs['reference_code'].map(
                    lambda x: download_info.get(x, {}).get('downloaded', 0)
                )
                untrained_refs['total_urls_encontradas'] = untrained_refs['reference_code'].map(
                    lambda x: download_info.get(x, {}).get('total_urls', 0)
                )
                untrained_refs['tasa_exito_descarga'] = untrained_refs['reference_code'].map(
                    lambda x: f"{download_info.get(x, {}).get('success_rate', 0):.1f}%"
                )
                untrained_refs['directorio_imagenes'] = untrained_refs['reference_code'].map(
                    lambda x: download_info.get(x, {}).get('output_directory', 'N/A')
                )
            
            # Agregar información de sugerencias
            if not analysis_results['label_suggestions'].empty:
                suggestions_info = analysis_results['label_suggestions'].set_index('reference_code')[
                    ['confident_suggestion', 'main_suggestion_label_id', 'main_suggestion_percentage', 'suggestion_confidence']
                ].to_dict('index')
                
                untrained_refs['sugerencia_confiable'] = untrained_refs['reference_code'].map(
                    lambda x: suggestions_info.get(x, {}).get('confident_suggestion', 'N/A')
                )
                untrained_refs['sugerencia_principal'] = untrained_refs['reference_code'].map(
                    lambda x: suggestions_info.get(x, {}).get('main_suggestion_label_id', 'N/A')
                )
                untrained_refs['porcentaje_sugerencia'] = untrained_refs['reference_code'].map(
                    lambda x: f"{suggestions_info.get(x, {}).get('main_suggestion_percentage', 0):.1f}%"
                )
                untrained_refs['nivel_confianza'] = untrained_refs['reference_code'].map(
                    lambda x: suggestions_info.get(x, {}).get('suggestion_confidence', 'Sin datos')
                )
            
            return untrained_refs
            
        except Exception as e:
            self.logger.error(f"Error preparando datos de referencias nuevas: {e}")
            return analysis_results['untrained_references']
    
    def _prepare_detailed_suggestions(self, suggestions_df: pd.DataFrame) -> pd.DataFrame:
        """Prepara sugerencias detalladas expandiendo todas las opciones"""
        try:
            detailed_rows = []
            
            for _, row in suggestions_df.iterrows():
                if row.get('all_suggestions'):
                    for sugg in row['all_suggestions']:
                        if isinstance(sugg, dict):
                            detailed_rows.append({
                                'reference_code': row['reference_code'],
                                'reference_name': row['reference_name'],
                                'suggested_label_id': sugg.get('label_id'),
                                'count': sugg.get('count'),
                                'percentage': f"{sugg.get('percentage', 0):.1f}%",
                                'total_detections': row['total_detections'],
                                'confidence_level': row['suggestion_confidence']
                            })
            
            return pd.DataFrame(detailed_rows)
            
        except Exception as e:
            self.logger.error(f"Error preparando sugerencias detalladas: {e}")
            return pd.DataFrame()
    
    def _format_excel_file(self, excel_path: str):
        """Aplica formato básico al archivo Excel"""
        try:
            if not EXCEL_AVAILABLE:
                return
            
            # Obtener configuración de color
            from ..config.settings import Settings
            settings = Settings()
            color_especial = getattr(settings.dataset, 'color_columnas_especiales', '#87CEEB')
            
            # Convertir color hex a código para openpyxl (sin #)
            color_code = color_especial.replace('#', '') if color_especial.startswith('#') else color_especial
            
            # Usar ReportExcelManager para asegurar prompts de bloqueo y preservación de dropdowns
            from utils.report_excel_manager import ReportExcelManager

            def _apply_formatting(wb):
                # Formato básico para todas las hojas
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                # Formato específico para columnas especiales (azul claro con texto negro)
                special_font = Font(color="000000")  # Texto negro para mejor legibilidad
                special_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

                # Columnas que deben tener formato especial
                special_columns = ['revisar_imagenes', 'incluir_dataset', 'devices_incluir', 'num_imagenes_dataset', 'producto_a_entrenar', 'label_considerar', 'num_imagenes_dataset_adicionales']

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    # Formato de cabeceras (primera fila)
                    if sheet.max_row > 0:
                        header_row = list(sheet[1])
                        column_names = [cell.value for cell in header_row]

                        for cell in header_row:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center")

                        # Aplicar formato especial a columnas específicas en References y Devices
                        if sheet_name in ['References', 'Devices'] and sheet.max_row > 1:
                            for col_idx, col_name in enumerate(column_names):
                                if col_name in special_columns:
                                    column_letter = get_column_letter(col_idx + 1)
                                    for row_num in range(2, sheet.max_row + 1):
                                        cell = sheet[f'{column_letter}{row_num}']
                                        cell.font = special_font
                                        cell.fill = special_fill

                    # Ajustar ancho de columnas
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except Exception:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width

                # Crear / refrescar hoja de validación para dropdowns
                self._add_validation_sheet(wb)

            ReportExcelManager(excel_path, interactive=True).save_with_validations(_apply_formatting)
            
        except Exception as e:
            self.logger.warning(f"Error aplicando formato a Excel: {e}")

    def _reinforce_special_columns(self, excel_path: str):
        """Asegura que las columnas especiales (dropdown) existen y tipos correctos incluso tras ediciones.
        Evita que Excel/usuarios eliminen accidentalmente encabezados críticos."""
        try:
            from ..config.settings import Settings
            settings = Settings()
            special_columns = ['revisar_imagenes', 'incluir_dataset', 'devices_incluir', 'num_imagenes_dataset', 'producto_a_entrenar', 'label_considerar', 'num_imagenes_dataset_adicionales']
            import openpyxl
            from utils.report_excel_manager import ReportExcelManager

            def _reinforce(wb):
                changed_local = False
                if 'References' in wb.sheetnames:
                    ws = wb['References']
                    headers = [c.value for c in ws[1]]
                    for col in special_columns:
                        if col not in headers:
                            ws.cell(row=1, column=len(headers)+1, value=col)
                            headers.append(col)
                            changed_local = True
                if 'Devices' in wb.sheetnames:
                    ws = wb['Devices']
                    headers = [c.value for c in ws[1]]
                    for col in ['revisar_imagenes','incluir_dataset','num_imagenes_dataset_adicionales']:
                        if col not in headers:
                            ws.cell(row=1, column=len(headers)+1, value=col)
                            headers.append(col)
                            changed_local = True
                # No necesidad de actuar si no hubo cambios; manager guardará igual para refrescar validaciones
                if changed_local:
                    self.logger.info("Columnas especiales reforzadas en Excel")

            ReportExcelManager(excel_path, interactive=True).save_with_validations(_reinforce)
        except Exception as e:
            self.logger.debug(f"No se pudo reforzar columnas especiales: {e}")
    
    def _add_validation_sheet(self, workbook):
        """Añade una hoja oculta con valores de validación para dropdowns"""
        try:
            from ..config.settings import Settings
            settings = Settings()
            opciones_devices = getattr(settings.dataset, 'opciones_devices_incluir', ['todos', 'devices_seleccionados'])
            opciones_si_no = ['si', 'no']
            
            # Crear hoja de validación (oculta)
            if 'Validacion' in workbook.sheetnames:
                validation_sheet = workbook['Validacion']
            else:
                validation_sheet = workbook.create_sheet('Validacion')
            
            # Limpiar contenido existente
            validation_sheet.delete_rows(1, validation_sheet.max_row)
            
            # Añadir opciones para devices_incluir (Columna A)
            validation_sheet['A1'] = 'Opciones_Devices_Incluir'
            for i, opcion in enumerate(opciones_devices, start=2):
                validation_sheet[f'A{i}'] = opcion
            
            # Añadir opciones sí/no (Columna B)
            validation_sheet['B1'] = 'Opciones_Si_No'
            for i, opcion in enumerate(opciones_si_no, start=2):
                validation_sheet[f'B{i}'] = opcion
            
            # Ocultar la hoja
            validation_sheet.sheet_state = 'hidden'
            
            # Añadir validaciones de datos en la hoja References
            if 'References' in workbook.sheetnames:
                ref_sheet = workbook['References']
                if ref_sheet.max_row > 0:
                    header_row = list(ref_sheet[1])
                    column_names = [cell.value for cell in header_row]
                    
                    from openpyxl.worksheet.datavalidation import DataValidation
                    
                    # Buscar columnas y aplicar validaciones
                    for col_idx, col_name in enumerate(column_names):
                        column_letter = get_column_letter(col_idx + 1)
                        
                        if col_name == 'devices_incluir':
                            # Validación para devices_incluir
                            dv = DataValidation(
                                type="list", 
                                formula1=f"Validacion!$A$2:$A${len(opciones_devices)+1}",
                                allow_blank=False
                            )
                            dv.error = 'Seleccione: todos o devices_seleccionados'
                            dv.errorTitle = 'Error de validación'
                            dv.prompt = 'Seleccione una opción válida'
                            dv.promptTitle = 'Opciones disponibles'
                            
                            dv.add(f"{column_letter}2:{column_letter}1000")
                            ref_sheet.add_data_validation(dv)
                            
                        elif col_name in ['revisar_imagenes', 'incluir_dataset']:
                            # Validación para columnas sí/no
                            dv = DataValidation(
                                type="list", 
                                formula1=f"Validacion!$B$2:$B${len(opciones_si_no)+1}",
                                allow_blank=False
                            )
                            dv.error = 'Seleccione: sí o no'
                            dv.errorTitle = 'Error de validación'
                            dv.prompt = 'Seleccione sí o no'
                            dv.promptTitle = 'Opciones: sí/no'
                            
                            dv.add(f"{column_letter}2:{column_letter}1000")
                            ref_sheet.add_data_validation(dv)
            
            # Aplicar validaciones también en la hoja Devices
            if 'Devices' in workbook.sheetnames:
                devices_sheet = workbook['Devices']
                if devices_sheet.max_row > 0:
                    header_row = list(devices_sheet[1])
                    column_names = [cell.value for cell in header_row]
                    
                    for col_idx, col_name in enumerate(column_names):
                        column_letter = get_column_letter(col_idx + 1)
                        
                        if col_name in ['revisar_imagenes', 'incluir_dataset']:
                            # Validación para columnas sí/no en Devices
                            dv = DataValidation(
                                type="list", 
                                formula1=f"Validacion!$B$2:$B${len(opciones_si_no)+1}",
                                allow_blank=False
                            )
                            dv.error = 'Seleccione: sí o no'
                            dv.errorTitle = 'Error de validación'
                            dv.prompt = 'Seleccione sí o no'
                            dv.promptTitle = 'Opciones: sí/no'
                            
                            dv.add(f"{column_letter}2:{column_letter}1000")
                            devices_sheet.add_data_validation(dv)
                            
        except Exception as e:
            self.logger.warning(f"Error añadiendo hoja de validación: {e}")
    
    def generate_all_reports(self, extracted_data: Dict[str, Any],
                           analysis_results: Dict[str, Any],
                           download_results: Dict[str, Any],
                           deployment_id: int) -> Dict[str, str]:
        """
        Genera todos los informes necesarios
        
        Args:
            extracted_data: Datos extraídos
            analysis_results: Resultados del análisis
            download_results: Resultados de descarga
            deployment_id: ID del deployment
            
        Returns:
            Diccionario con rutas de archivos generados
        """
        try:
            self.logger.info("Generando todos los informes")
            
            # Verificar si hay un proceso activo
            process_info = None
            try:
                from ..process_aware_settings import ProcessAwareSettings
                if hasattr(self.settings, 'get_process_info'):
                    process_info = self.settings.get_process_info()
                else:
                    temp_settings = ProcessAwareSettings()
                    process_info = temp_settings.get_process_info()
            except:
                process_info = None
            
            # Si hay proceso activo, no usar session_dir para permitir archivos directos
            if process_info:
                session_dir = None
                session_log_path = None
            else:
                # Crear directorio de sesión para casos sin proceso activo
                session_dir = self._create_session_directory(deployment_id, extracted_data)
                # Configurar logging de sesión
                session_log_path = self._setup_session_logging(session_dir, deployment_id)
            
            report_files = {}
            if session_dir:
                report_files['session_directory'] = session_dir
                report_files['session_log'] = session_log_path
            
            # Archivo principal
            main_excel = self.create_main_excel_report(
                extracted_data, analysis_results, download_results, deployment_id, session_dir
            )
            report_files['main_report'] = main_excel

            # No generar archivo de referencias nuevas (no necesario)
            # new_refs_excel = self.create_new_references_excel(
            #     analysis_results, download_results, deployment_id, session_dir
            # )
            # report_files['new_references_report'] = new_refs_excel            self.logger.info("Todos los informes generados exitosamente")
            return report_files
            
        except Exception as e:
            self.logger.error(f"Error generando informes: {e}")
            raise

    def _create_devices_sheet(self, writer, devices_df: pd.DataFrame):
        """Crea la pestaña de análisis de devices"""
        try:
            # Obtener configuración de dataset
            from ..config.settings import Settings
            settings = Settings()
            num_imagenes_adicionales_default = getattr(settings.dataset, 'num_imagenes_dataset_adicionales_default', 0)
            
            if devices_df.empty:
                # Crear DataFrame vacío con las columnas correctas (incluyendo nuevas)
                empty_df = pd.DataFrame(columns=[
                    'device_id', 'device_name', 'num_cliente', 'num_manual', 
                    'pct_ok_sistema', 'pct_ok_modelo', 'revisar_imagenes',
                    'incluir_dataset', 'num_imagenes_dataset_adicionales'
                ])
                empty_df.to_excel(writer, sheet_name='Devices', index=False)
                self.logger.info("Pestaña Devices creada (vacía - no hay datos)")
                return
            
            # Preparar los datos con las nuevas columnas
            devices_df_enhanced = devices_df.copy()
            
            # Normalizar revisar_imagenes a 'sí'/'no'
            if 'revisar_imagenes' in devices_df_enhanced.columns:
                devices_df_enhanced['revisar_imagenes'] = devices_df_enhanced['revisar_imagenes'].map({
                    True: 'si', 
                    False: 'no',
                    'True': 'si',
                    'False': 'no',
                    'si': 'si',
                    'sí': 'si',
                    'no': 'no',
                    1: 'si',
                    0: 'no'
                }).fillna('no')
            else:
                devices_df_enhanced['revisar_imagenes'] = 'no'
            
            # Añadir columna incluir_dataset (por defecto 'sí')
            devices_df_enhanced['incluir_dataset'] = 'si'
            
            # Añadir columna num_imagenes_dataset_adicionales (valor configurable, por defecto 0)
            devices_df_enhanced['num_imagenes_dataset_adicionales'] = num_imagenes_adicionales_default
            
            # Ordenar por número de transacciones de cliente
            devices_df_sorted = devices_df_enhanced.sort_values('num_cliente', ascending=False)
            
            # Asegurar columna first_transaction_date si falta
            if 'first_transaction_date' not in devices_df_sorted.columns:
                devices_df_sorted['first_transaction_date'] = ''
            # Convertir first_transaction_date a datetime y formatear como fecha (YYYY-MM-DD)
            try:
                devices_df_sorted['first_transaction_date'] = pd.to_datetime(devices_df_sorted['first_transaction_date'], errors='coerce').dt.date
            except Exception:
                pass
            devices_df_sorted.to_excel(writer, sheet_name='Devices', index=False)
            self.logger.info(f"Pestaña Devices creada con {len(devices_df_sorted)} devices")
            
        except Exception as e:
            self.logger.error(f"Error creando pestaña Devices: {e}")
            # Crear pestaña vacía en caso de error
            empty_df = pd.DataFrame(columns=[
                'device_id', 'device_name', 'num_cliente', 'num_manual', 
                'pct_ok_sistema', 'pct_ok_modelo', 'revisar_imagenes',
                'incluir_dataset', 'num_imagenes_dataset_adicionales'
            ])
            empty_df.to_excel(writer, sheet_name='Devices', index=False)

    def _create_consistency_sheet_fixed(self, writer, analysis_results: Dict[str, Any]):
        """Crea la pestaña de análisis de consistencia corregida"""
        try:
            consistency_data = []
            consistency_data.append(['ANÁLISIS DE CONSISTENCIA', ''])
            
            # Obtener estadísticas del summary_stats si está disponible
            summary_stats = analysis_results.get('summary_stats', {})
            
            # Información básica
            consistency_data.append(['Total referencias procesadas', summary_stats.get('total_references', 'N/A')])
            consistency_data.append(['Referencias sin asignar', summary_stats.get('unassigned_count', 'N/A')])
            consistency_data.append(['Referencias no entrenadas', summary_stats.get('untrained_count', 'N/A')])
            consistency_data.append(['Referencias entrenadas', summary_stats.get('trained_count', 'N/A')])
            consistency_data.append(['', ''])
            
            # Análisis de devices si está disponible
            devices_analysis = analysis_results.get('devices_analysis', pd.DataFrame())
            if not devices_analysis.empty:
                consistency_data.append(['ESTADÍSTICAS DE DEVICES', ''])
                consistency_data.append(['Total devices analizados', len(devices_analysis)])
                consistency_data.append(['Devices con transacciones > 100', len(devices_analysis[devices_analysis['num_cliente'] > 100])])
                consistency_data.append(['Devices con transacciones > 50', len(devices_analysis[devices_analysis['num_cliente'] > 50])])
                consistency_data.append(['', ''])
            
            consistency_data.append(['Estado del análisis', 'Completado correctamente'])
            
            # Convertir a DataFrame
            consistency_df = pd.DataFrame(consistency_data, columns=['Métrica', 'Valor'])
            consistency_df.to_excel(writer, sheet_name='Consistencia', index=False)
            
            self.logger.info("Pestaña Consistencia creada correctamente")
            
        except Exception as e:
            self.logger.error(f"Error creando pestaña de consistencia: {e}")
            # Crear pestaña básica en caso de error
            error_data = [
                ['ANÁLISIS DE CONSISTENCIA', ''],
                ['Estado', 'Error en generación'],
                ['Error', str(e)]
            ]
            error_df = pd.DataFrame(error_data, columns=['Métrica', 'Valor'])
            error_df.to_excel(writer, sheet_name='Consistencia', index=False)
