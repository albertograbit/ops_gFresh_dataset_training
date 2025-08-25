"""
Creador de datasets de entrenamiento
Descarga imágenes organizadas por clase y label_id según configuración Excel
"""

import pandas as pd
import os
import logging
import random
import glob
from pathlib import Path
import time
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Tuple
import yaml
from botocore.exceptions import ClientError
from datetime import datetime
import re
from utils.file_manager import read_csv_sc, is_file_locked, wait_for_writable
from utils.report_excel_manager import ReportExcelManager, ExcelLockedError


class DatasetCreator:
    """
    Creador de datasets de entrenamiento organizados por clase y label_id
    """
    
    def __init__(self, config_path: str = "config/settings.yaml", process_aware: bool = False):
        """
        Inicializar el creador de datasets
        
        Args:
            config_path: Ruta al archivo de configuración
            process_aware: Si usar configuración consciente de procesos
        """
        self.logger = logging.getLogger(__name__)
        self.process_aware = process_aware
        
        # Cargar variables de entorno del .env
        from dotenv import load_dotenv
        load_dotenv()
        
        # Cargar configuración
        if process_aware:
            from dataset_manager.process_aware_settings import ProcessAwareSettings
            process_settings = ProcessAwareSettings(config_path)
            full_config = process_settings.get_settings()
            self.config = full_config.get('dataset', {})
            
            # Obtener directorio base del proceso activo
            process_info = process_settings.get_process_info()
            if process_info and 'process_dir' in process_info:
                self.process_dir = Path(process_info['process_dir'])
                self.process_name = process_info['name']
            else:
                raise ValueError("No hay proceso activo configurado")
        else:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            self.config = config.get('dataset', {})
            self.process_dir = Path('./output')
            self.process_name = 'default'
        
        # Configuración de almacenamiento (S3 / MinIO) centralizada
        try:
            from dataset_manager.storage.storage_client import get_storage_client
            self.s3_client, storage_env = get_storage_client(self.logger, validate=True)
            self.s3_bucket = storage_env.bucket
            self.s3_region = storage_env.region
            self.storage_type = storage_env.storage_type
        except Exception as e:
            self.logger.warning(f"Fallo inicializando cliente de almacenamiento compartido: {e}. Se intentará inicialización diferida mínima.")
            self.s3_client = None
            self.s3_bucket = os.getenv('S3_BUCKET', 'grabit-data')
            self.s3_region = os.getenv('REMOTE_STORAGE_REGION', 'eu-west-2')
            self.storage_type = 's3'
        
        # Configuración de descarga
        self.download_timeout = 30
        self.max_retries = 3
        self.image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
        
        # Lista para rastrear imágenes descargadas
        self.downloaded_images_log = []
        # Cache para transacciones ya usadas (evita re-leer Excels repetidamente)
        self._used_transactions_cache = None

        # Concurrencia para descargas: configurable desde settings.image_processing.max_workers
        try:
            from dataset_manager.config.settings import Settings
            _settings = Settings()
            self.download_max_workers = max(1, int(getattr(_settings.image_processing, 'max_workers', 5)))
        except Exception:
            self.download_max_workers = 5

        self.logger.info(f"DatasetCreator inicializado:")
        self.logger.info(f"  - Proceso: {self.process_name}")
        self.logger.info(f"  - Directorio base: {self.process_dir}")
        self.logger.info(f"  - Bucket: {self.s3_bucket} (tipo={getattr(self,'storage_type','s3')})")

    def _init_s3_client(self):
        """Mantener método por compatibilidad; ya se usa almacenamiento compartido."""
        if self.s3_client is None:
            try:
                from dataset_manager.storage.storage_client import get_storage_client
                self.s3_client, storage_env = get_storage_client(self.logger, validate=False)
                self.s3_bucket = storage_env.bucket
                self.s3_region = storage_env.region
                self.storage_type = storage_env.storage_type
            except Exception as e:
                self.logger.error(f"No se pudo inicializar cliente de almacenamiento: {e}")
                self.s3_client = None

    def crear_dataset(self, deployment_id: int, excel_path: Optional[str] = None, 
                     dry_run: bool = False, force: bool = False,
                     fast: bool = False, limit_refs: Optional[int] = None,
                     limit_devices: Optional[int] = None,
                     images_per_ref: Optional[int] = None,
                     images_per_device: Optional[int] = None,
                     max_workers: Optional[int] = None,
                     no_filter_used: bool = False) -> Dict:
        """
        Crear dataset de entrenamiento
        
        Args:
            deployment_id: ID del deployment
            excel_path: Ruta al archivo Excel (opcional)
            dry_run: Si True, solo simula la descarga
            force: Si True, fuerza creación aunque ya exista
            
        Returns:
            Diccionario con resultados del procesamiento
        """
        try:
            self.logger.info(f"🎯 Iniciando creación de dataset para deployment {deployment_id}")
            t0 = time.time()
            # Permitir override de workers por ejecución
            if max_workers is not None:
                try:
                    mw = int(max_workers)
                    if mw > 0:
                        self.download_max_workers = mw
                        self.logger.info(f"⚙️  max_workers sobrescrito a {self.download_max_workers}")
                except Exception:
                    pass
            
            # 1. Encontrar archivo Excel si no se especifica
            if excel_path is None:
                excel_path = self._find_excel_file()
                if excel_path is None:
                    raise ValueError("No se encontró archivo Excel para procesar")
            
            self.logger.info(f"📊 Cargando datos desde: {excel_path}")
            
            # 2. Crear carpeta de dataset con versionado automático
            dataset_folder = self._create_dataset_folder(deployment_id, force)
            self.logger.info(f"📁 Carpeta dataset: {dataset_folder}")
            
            # 3. Cargar y validar datos de referencias
            t_refs0 = time.time()
            df_references = self._load_and_validate_references(excel_path)
            # Modo rápido: limitar número de referencias
            if fast and limit_refs is not None:
                if limit_refs > 0 and len(df_references) > limit_refs:
                    df_references = df_references.sample(n=limit_refs, random_state=42)
            self.logger.info(f"✅ Referencias válidas para dataset: {len(df_references)}")
            self.logger.info(f"⏱️  Tiempo referencias: {time.time()-t_refs0:.2f}s")
            
            # 4. (Lazy) No crear estructura completa aún; se creará al descargar imágenes
            class_structure = {}
            self.logger.info("🏷️  Creación diferida de carpetas (solo cuando haya imágenes)")
            
            # 5. Cargar datos de Elasticsearch
            t_es0 = time.time()
            df_elastic = self._load_elasticsearch_data(excel_path)
            self.logger.info(f"📈 Datos de Elasticsearch cargados: {len(df_elastic)} registros")
            self.logger.info(f"⏱️  Tiempo cargar Elasticsearch: {time.time()-t_es0:.2f}s")
            
            # 6. Procesar todas las referencias de una vez con pandas
            t_proc_refs0 = time.time()
            total_images = self._process_all_references_bulk(
                df_references, df_elastic, excel_path, dataset_folder, dry_run,
                fast=fast, images_per_ref=images_per_ref, no_filter_used=no_filter_used
            )
            self.logger.info(f"⏱️  Tiempo procesar referencias: {time.time()-t_proc_refs0:.2f}s")
            
            # 7. Procesar devices adicionales si los hay
            t_dev0 = time.time()
            additional_images = self._process_additional_devices(
                excel_path, df_elastic, df_references, dataset_folder, dry_run,
                fast=fast, limit_devices=limit_devices, images_per_device=images_per_device,
                no_filter_used=no_filter_used
            )
            total_images += additional_images
            self.logger.info(f"⏱️  Tiempo devices: {time.time()-t_dev0:.2f}s")
            
            # 9. Guardar registro de imágenes descargadas
            if not dry_run:
                t_log0 = time.time()
                self._save_dataset_log(excel_path, dataset_folder)
                self.logger.info(f"⏱️  Tiempo guardar log: {time.time()-t_log0:.2f}s")
            
            results = {
                'dataset_folder': str(dataset_folder),
                'references_processed': len(df_references),
                'classes_created': len(class_structure),
                'total_images': total_images,
                'excel_path': excel_path
            }
            
            self.logger.info(f"⏱️  Tiempo total: {time.time()-t0:.2f}s")
            self.logger.info(f"🎉 Dataset creado exitosamente: {results}")
            return results
            
        except Exception as e:
            self.logger.error(f"Error creando dataset: {e}")
            raise

    def _find_excel_file(self) -> Optional[str]:
        """Encontrar el archivo Excel del proceso activo"""
        reports_dir = self.process_dir / "reports"
        if not reports_dir.exists():
            return None
        
        # 1) Preferir el Excel del proceso; si existe doble extensión, priorizarla
        try:
            proc_name = getattr(self, 'process_name', '') or ''
            canonical_double = reports_dir / f"{proc_name}.xlsx.xlsx"
            if canonical_double.exists():
                return str(canonical_double)
            canonical = reports_dir / f"{proc_name}.xlsx"
            if canonical.exists():
                return str(canonical)
        except Exception:
            pass

        # 2) Fallback: elegir el más reciente válido, ignorando backups (con _dataset_) y temporales
        excel_files = list(reports_dir.glob("*.xlsx"))
        valid_files = []
        for file_path in excel_files:
            filename = file_path.name.lower()
            if filename.startswith('~') or filename.startswith('.'):
                continue
            if '_dataset_' in filename:
                continue
            valid_files.append(file_path)

        if not valid_files:
            return None
        latest_file = max(valid_files, key=lambda f: f.stat().st_mtime)
        return str(latest_file)

    def _create_dataset_folder(self, deployment_id: int, force: bool = False) -> Path:
        """Crear carpeta del dataset: dataset_<deployment>_<procDate>_<procVersion>_vXX"""
        datasets_dir = self.process_dir / "datasets"
        datasets_dir.mkdir(exist_ok=True)
        # Tomar fecha y versión base del nombre de proceso si es posible
        base_date = datetime.now().strftime("%d%m%y")
        base_proc_version = None  # por ejemplo, v10 del proceso
        try:
            proc_name = getattr(self, 'process_name', '') or ''
            m = re.match(r"dataset_(\d+)_(\d{6})_(v\d+)$", proc_name)
            if m:
                base_date = m.group(2)
                base_proc_version = m.group(3)
        except Exception:
            pass

        # Construir patrón de búsqueda para datasets existentes
        if base_proc_version:
            pattern = f"dataset_{deployment_id}_{base_date}_{base_proc_version}_v*"
        else:
            pattern = f"dataset_{deployment_id}_{base_date}_v*"

        existing_datasets = list(datasets_dir.glob(pattern))

        # Crear SIEMPRE una nueva carpeta de dataset en cada ejecución (no reutilizar la última)
        version = 1
        # Si ya existen, elegir la siguiente versión libre
        while True:
            if base_proc_version:
                folder_name = f"dataset_{deployment_id}_{base_date}_{base_proc_version}_v{version:02d}"
            else:
                folder_name = f"dataset_{deployment_id}_{base_date}_v{version:02d}"
            folder_path = datasets_dir / folder_name
            if not folder_path.exists():
                break
            version += 1

        folder_path.mkdir(exist_ok=True)
        return folder_path

    def _load_and_validate_references(self, excel_path: str) -> pd.DataFrame:
        """Cargar y validar datos de referencias desde Excel"""
        try:
            df = pd.read_excel(excel_path, sheet_name='References')
            
            # Validar columnas requeridas
            required_cols = ['reference_name', 'producto_a_entrenar', 'label_considerar']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise ValueError(f"Columnas faltantes en 'references': {missing_cols}")
            # Filtrar solo las que deben incluirse en el dataset (si existe la columna)
            if 'incluir_dataset' in df.columns:
                mask_si = (
                    df['incluir_dataset']
                    .astype(str)
                    .str.lower()
                    .str.strip()
                    .isin(['si', 'sí'])
                )
                df_valid = df[mask_si].copy()
                self.logger.info(f"Referencias totales: {len(df)} | Marcadas incluir_dataset=si: {len(df_valid)}")
            else:
                df_valid = df.copy()
                self.logger.info(f"Referencias totales (sin columna incluir_dataset): {len(df_valid)}")

            # Detectar filas con campos obligatorios vacíos (antes de dropear)
            missing_mask = (
                df_valid['producto_a_entrenar'].isna() | (df_valid['producto_a_entrenar'].astype(str).str.strip() == '') |
                df_valid['label_considerar'].isna() | (df_valid['label_considerar'].astype(str).str.strip() == '')
            )
            problematic = df_valid[missing_mask]
            if not problematic.empty:
                cols_show = ['reference_name', 'producto_a_entrenar', 'label_considerar']
                preview = problematic[cols_show].head(50)
                self.logger.error("Referencias con campos obligatorios vacíos (producto_a_entrenar / label_considerar):")
                self.logger.error("\n" + preview.to_string(index=False))
                raise ValueError(f"Se encontraron {len(problematic)} referencias con datos incompletos. Corrige el Excel y reintenta.")

            # Limpiar y validar datos (ya sabemos que no faltan obligatorios)
            df_valid['class'] = df_valid['producto_a_entrenar'].astype(str).str.strip()
            df_valid['label_id'] = pd.to_numeric(df_valid['label_considerar'], errors='coerce')
            if df_valid['label_id'].isna().any():
                invalid_label = df_valid[df_valid['label_id'].isna()][['reference_name','label_considerar']].head(50)
                self.logger.error("Referencias con label_considerar no numérico:")
                self.logger.error("\n" + invalid_label.to_string(index=False))
                raise ValueError("Existen valores no numéricos en label_considerar. Corrige y reintenta.")

            if df_valid.empty:
                raise ValueError("No hay referencias válidas para procesar tras validación")
            
            return df_valid
            
        except Exception as e:
            self.logger.error(f"Error cargando referencias: {e}")
            raise

    def _create_class_structure(self, df_references: pd.DataFrame, dataset_folder: Path, dry_run: bool) -> Dict:
        """Crear estructura de carpetas por clase y label_id"""
        class_structure = {}
        
        for _, row in df_references.iterrows():
            class_name = row['class']
            label_id = int(row['label_id'])  # Convertir a int para evitar decimales
            
            if class_name not in class_structure:
                class_structure[class_name] = []
            
            if label_id not in class_structure[class_name]:
                class_structure[class_name].append(label_id)
                
                # Crear directorio solo cuando sea necesario (lazy creation)
                if not dry_run:
                    class_dir = dataset_folder / class_name / str(label_id)
                    class_dir.mkdir(parents=True, exist_ok=True)
                    self.logger.debug(f"Directorio creado: {class_dir}")
        
        return class_structure

    def _load_elasticsearch_data(self, excel_path: str) -> pd.DataFrame:
        """Cargar datos de Elasticsearch desde el CSV externo (elastic_data.csv)."""
        try:
            # CSV ahora se guarda en reports/<process_name>/reports/elastic_data.csv o en la carpeta reports raíz del proceso
            # Buscar primero en carpeta reports del proceso activo
            candidates = []
            reports_dir = self.process_dir / 'reports'
            if reports_dir.exists():
                candidates.append(reports_dir / 'elastic_data.csv')
            # Legacy fallback: carpeta data/elastic_cache.csv (antiguo) por compatibilidad temporal
            legacy = self.process_dir / 'data' / 'elastic_cache.csv'
            if legacy.exists():
                candidates.append(legacy)
            for path in candidates:
                if path.exists():
                    try:
                        df = read_csv_sc(str(path))
                        self.logger.info(f"⚡ Cargado datos Elasticsearch desde CSV: {path} ({len(df)} filas)")
                        # Validar columnas mínimas
                        critical_columns = ['transaction_id', 'selected_reference_code', 'image_link']
                        missing_critical = [col for col in critical_columns if col not in df.columns]
                        if missing_critical:
                            raise ValueError(f"Columnas críticas faltantes en CSV {path}: {missing_critical}")
                        df = df.dropna(subset=critical_columns)
                        return df
                    except Exception as ce:
                        self.logger.warning(f"Problema leyendo {path}: {ce}")
            raise FileNotFoundError("No se encontró elastic_data.csv. Ejecute primero la extracción (download_info).")
        except Exception as e:
            self.logger.error(f"Error cargando datos de Elasticsearch: {e}")
            raise

    def _load_devices_filter(self, excel_path: str, df_references: pd.DataFrame) -> Optional[List]:
        """Cargar filtro de devices si existe la hoja"""
        try:
            df_devices = pd.read_excel(excel_path, sheet_name='Devices')
            if 'device_id' in df_devices.columns:
                # Si existe columna incluir_dataset, filtrar por 'si'/'sí' (case/whitespace-insensitive)
                if 'incluir_dataset' in df_devices.columns:
                    inc = (
                        df_devices['incluir_dataset']
                        .astype(str)
                        .str.lower()
                        .str.strip()
                        .isin(['si', 'sí'])
                    )
                    df_devices = df_devices[inc].copy()
                    self.logger.info(f"🔧 Filtro de devices (incluir_dataset=si): {len(df_devices)} devices")
                else:
                    self.logger.info("🔧 Hoja Devices sin columna 'incluir_dataset' - no se aplica filtro por inclusión")

                # Normalizar device_id a strings sin decimales
                series = df_devices['device_id'].dropna()
                nums = pd.to_numeric(series, errors='coerce')
                normalized = []
                for orig, num in zip(series, nums):
                    if pd.notna(num):
                        try:
                            normalized.append(str(int(num)))
                        except Exception:
                            normalized.append(str(orig).strip())
                    else:
                        normalized.append(str(orig).strip())
                devices_list = sorted(set(normalized))
                self.logger.info(f"🔧 Dispositivos permitidos por Devices: {len(devices_list)}")
                return devices_list
        except Exception as e:
            self.logger.debug(f"No se pudo cargar filtro de devices (opcional): {e}")
        
        return None

    def _filter_used_transactions(self, df_elastic: pd.DataFrame) -> pd.DataFrame:
        """Filtrar transacciones que ya fueron descargadas en datasets previos usando pandas"""
        try:
            # Asegurar que trabajamos sobre una copia para evitar SettingWithCopyWarning
            df_elastic = df_elastic.copy()
            # Usar cache si ya está disponible
            if self._used_transactions_cache is None:
                used_transaction_ids = set()
                # Buscar una sola vez todos los archivos Excel del proceso y acumular transacciones usadas
                reports_dir = self.process_dir / "reports"
                if reports_dir.exists():
                    for excel_file in reports_dir.glob("*.xlsx"):
                        try:
                            df_prev = pd.read_excel(excel_file, sheet_name='datasets_creados')
                            if 'transaction_id' in df_prev.columns:
                                prev_transactions = df_prev['transaction_id'].dropna().astype(str)
                                used_transaction_ids.update(prev_transactions)
                        except Exception:
                            continue  # Ignorar archivos sin la hoja
                # Guardar en cache (aunque sea vacío) para evitar relecturas
                self._used_transactions_cache = used_transaction_ids
            else:
                used_transaction_ids = self._used_transactions_cache
            
            if used_transaction_ids:
                # Filtrar usando pandas - mucho más eficiente
                df_elastic['transaction_id'] = df_elastic['transaction_id'].astype(str)
                initial_count = len(df_elastic)
                df_filtered = df_elastic[~df_elastic['transaction_id'].isin(used_transaction_ids)]
                filtered_count = initial_count - len(df_filtered)
                
                self.logger.info(f"🔄 Transacciones filtradas (ya descargadas): {filtered_count}")
                self.logger.info(f"📊 Transacciones disponibles: {len(df_filtered)}")
                
                return df_filtered
            else:
                self.logger.info("📊 No hay transacciones previas para filtrar")
                return df_elastic
                
        except Exception as e:
            self.logger.warning(f"Error filtrando transacciones previas: {e}")
            return df_elastic

    def _process_all_references_bulk(self, df_references: pd.DataFrame, df_elastic: pd.DataFrame,
                                   excel_path: str, dataset_folder: Path, dry_run: bool,
                                   fast: bool = False, images_per_ref: Optional[int] = None,
                                   no_filter_used: bool = False) -> int:
        """Procesar todas las referencias de una vez usando pandas (mucho más eficiente)"""
        try:
            self.logger.info("🚀 Procesando todas las referencias en lote con pandas...")
            
            # 1. Cargar filtro de devices si existe
            devices_filter = self._load_devices_filter(excel_path, df_references)
            
            # 2. Filtrar transacciones ya utilizadas
            df_elastic_filtered = df_elastic if no_filter_used else self._filter_used_transactions(df_elastic)
            
            # 3. Hacer join entre referencias y datos de elasticsearch por reference_code
            df_merged = df_references.merge(
                df_elastic_filtered,
                left_on='reference_code',
                right_on='selected_reference_code',
                how='inner'
            )
            
            self.logger.info(f"📊 Referencias con datos: {len(df_merged)} transacciones encontradas")
            
            if df_merged.empty:
                self.logger.warning("❌ No hay datos disponibles después del merge")
                return 0
            
            # 4. Aplicar filtro de devices si existe
            if devices_filter:
                df_merged = df_merged[df_merged['device_id'].astype(str).isin(devices_filter)]
                self.logger.info(f"🔧 Después del filtro de devices: {len(df_merged)} transacciones")
            
            # 5. Seleccionar número de imágenes por referencia
            selected_transactions = []
            
            for reference_name in df_references['reference_name'].unique():
                ref_info = df_references[df_references['reference_name'] == reference_name].iloc[0]
                ref_transactions = df_merged[df_merged['reference_name'] == reference_name]
                
                if ref_transactions.empty:
                    continue
                # Log candidatos tras filtros
                try:
                    self.logger.info(f"📌 Candidatos para {reference_name} tras filtros: {len(ref_transactions)}")
                except Exception:
                    pass
                
                # Determinar número de imágenes a seleccionar
                if fast and images_per_ref is not None and images_per_ref > 0:
                    num_images = min(images_per_ref, len(ref_transactions))
                else:
                    num_images = self._get_images_to_download_from_series(ref_info, len(ref_transactions))
                
                # Seleccionar transacciones aleatoriamente
                if num_images > 0:
                    sample_size = min(num_images, len(ref_transactions))
                    selected = ref_transactions.sample(n=sample_size, random_state=42)
                    selected_transactions.append(selected)
                    self.logger.info(f"✅ {reference_name}: {len(selected)} imágenes seleccionadas")
            
            if not selected_transactions:
                self.logger.warning("❌ No se seleccionaron transacciones")
                return 0
            
            # 6. Combinar todas las transacciones seleccionadas
            df_to_download = pd.concat(selected_transactions, ignore_index=True)
            self.logger.info(f"📥 Total de imágenes a descargar: {len(df_to_download)}")
            
            # 7. Descargar imágenes en lote
            images_downloaded = self._download_images_bulk(df_to_download, dataset_folder, dry_run)
            
            self.logger.info(f"🎉 Procesamiento en lote completado: {images_downloaded} imágenes")
            return images_downloaded
            
        except Exception as e:
            self.logger.error(f"Error en procesamiento en lote: {e}")
            return 0

    def _get_images_to_download_from_series(self, ref_row: pd.Series, available_images: int) -> int:
        """Determinar cuántas imágenes descargar para esta referencia (versión optimizada)"""
        # Buscar columnas que indiquen número de imágenes
        for col in ['num_imagenes_dataset', 'imagenes', 'num_images']:
            if col in ref_row and pd.notna(ref_row[col]):
                requested = int(ref_row[col])
                return min(requested, available_images)
        
        # Por defecto, descargar hasta 50 imágenes
        return min(50, available_images)

    def _download_images_bulk(self, df_transactions: pd.DataFrame, dataset_folder: Path, dry_run: bool) -> int:
        """Descargar imágenes en lote con concurrencia por clase para mayor rendimiento"""
        images_downloaded = 0

        # Agrupar por clase y label_id para crear carpetas eficientemente
        for (class_name, label_id), group in df_transactions.groupby(['class', 'label_id']):
            class_folder = dataset_folder / class_name / str(int(label_id))

            if not dry_run:
                class_folder.mkdir(parents=True, exist_ok=True)

            # Descargar en paralelo dentro de cada clase
            with ThreadPoolExecutor(max_workers=self.download_max_workers) as executor:
                futures = [
                    executor.submit(self._download_single_image_optimized, row, class_folder, dry_run)
                    for _, row in group.iterrows()
                ]
                for future in as_completed(futures):
                    try:
                        if future.result():
                            images_downloaded += 1
                    except Exception:
                        # Contabilizar como fallo silencioso (ya se registra en el método interno)
                        pass

        return images_downloaded

    def _download_single_image_optimized(self, transaction: pd.Series, dest_folder: Path, dry_run: bool) -> bool:
        """Descargar una imagen optimizada (sin logging excesivo)"""
        try:
            if self.s3_client is None:
                return False
            
            image_link = transaction.get('image_link')
            if isinstance(image_link, str):
                image_link = image_link.strip().strip('"').strip("'")
                if image_link.lower() in ('none', 'null'):  # normalizar
                    image_link = ''
            if not image_link or pd.isna(image_link):
                return False
            
            # Registrar el intento (formato simplificado)
            image_record = {
                'reference_name': transaction.get('reference_name'),
                'clase_name': transaction.get('class'),
                'label_id': transaction.get('label_id'),
                'transaction_id': transaction.get('transaction_id'),
                'transaction_start_time': transaction.get('transaction_start_time'),
                'device_id': transaction.get('device_id'),
                's3_key': '',
                'local_path': '',
                'download_status': 'failed',
                'error_reason': '',
                'dataset_folder': str(dest_folder.parent.parent.name),
                'motivo_descarga': 'reference',
                'valor_motivo': transaction.get('reference_name')
            }
            
            # Limpiar la ruta de S3
            s3_key = str(image_link).strip().lstrip('/')
            # Si no apunta a un archivo con extensión válida, listar y elegir uno
            key_for_download = s3_key
            if not any(s3_key.lower().endswith(ext) for ext in self.image_extensions):
                objs = self._list_s3_objects(s3_key)
                if not objs:
                    image_record['error_reason'] = 'Imagen no encontrada en S3'
                    self.downloaded_images_log.append(image_record)
                    return False
                key_for_download = objs[0]
            image_record['s3_key'] = key_for_download

            # Crear nombre de archivo con formato solicitado
            filename = self._create_dataset_filename(transaction, key_for_download)
            dest_path = dest_folder / filename
            image_record['local_path'] = str(dest_path)
            
            if dry_run:
                image_record['download_status'] = 'dry_run'
                image_record['error_reason'] = ''
                self.downloaded_images_log.append(image_record)
                return True
            
            # Descargar imagen
            if self._download_from_s3(key_for_download, dest_path):
                image_record['download_status'] = 'success'
                image_record['error_reason'] = ''
                self.downloaded_images_log.append(image_record)
                return True
            else:
                image_record['error_reason'] = 'Error descargando desde S3'
                self.downloaded_images_log.append(image_record)
                return False
                
        except Exception as e:
            return False

    def _process_reference_dataset(self, ref_row: pd.Series, df_elastic: pd.DataFrame, 
                                 devices_filter: Optional[List], dataset_folder: Path, 
                                 dry_run: bool) -> int:
        """Procesar una referencia y descargar sus imágenes"""
        reference_name = ref_row['reference_name']
        class_name = ref_row['class']
        label_id = int(ref_row['label_id'])  # Convertir a int
        
        self.logger.info(f"🔍 Procesando referencia: {reference_name} -> {class_name}/{label_id}")
        
        # Obtener product_ids para esta referencia
        product_ids = self._get_product_ids_for_reference(reference_name, ref_row)
        if not product_ids:
            self.logger.warning(f"No se encontraron product_ids para {reference_name}")
            return 0
        
        # Filtrar datos de Elasticsearch
        filtered_data = df_elastic[df_elastic['product_id'].isin(product_ids)]
        
        # Aplicar filtro de devices si existe
        if devices_filter:
            filtered_data = filtered_data[
                filtered_data['device_id'].astype(str).isin(devices_filter)
            ]
        
        # Filtrar transacciones ya utilizadas
        filtered_data = self._filter_used_transactions(filtered_data)
        
        if filtered_data.empty:
            self.logger.warning(f"No hay datos disponibles para {reference_name}")
            return 0
        
        # Obtener número de imágenes a descargar
        num_images = self._get_images_to_download(ref_row, len(filtered_data))
        
        # Seleccionar datos aleatoriamente
        sample_data = filtered_data.sample(n=min(num_images, len(filtered_data)), random_state=42)
        
        self.logger.info(f"📥 Descargando {len(sample_data)} imágenes para {reference_name}")
        
        # Descargar imágenes
        images_downloaded = 0
        class_folder = dataset_folder / class_name / str(label_id)
        
        for _, transaction in sample_data.iterrows():
            if self._download_dataset_image(transaction, class_folder, ref_row, dry_run):
                images_downloaded += 1
        
        self.logger.info(f"✅ {reference_name}: {images_downloaded} imágenes descargadas")
        return images_downloaded

    def _get_product_ids_for_reference(self, reference_name: str, ref_row: pd.Series) -> List[str]:
        """Obtener product_ids para una referencia"""
        # Usar reference_code como product_id
        if 'reference_code' in ref_row and pd.notna(ref_row['reference_code']):
            return [str(int(ref_row['reference_code']))]
        
        # Fallback: usar product_id si está disponible
        if 'product_id' in ref_row and pd.notna(ref_row['product_id']):
            return [str(ref_row['product_id'])]
        
        # No se encontró mapeo
        return []

    def _get_images_to_download(self, ref_row: pd.Series, available_images: int) -> int:
        """Determinar cuántas imágenes descargar para esta referencia"""
        # Buscar columnas que indiquen número de imágenes
        for col in ['num_imagenes_dataset', 'imagenes', 'num_images']:
            if col in ref_row and pd.notna(ref_row[col]):
                requested = int(ref_row[col])
                return min(requested, available_images)
        
        # Por defecto, descargar hasta 50 imágenes
        return min(50, available_images)

    def _download_dataset_image(self, transaction: pd.Series, dest_folder: Path, 
                              ref_row: pd.Series, dry_run: bool) -> bool:
        """Descargar una imagen para el dataset"""
        try:
            if self.s3_client is None:
                self.logger.error("Cliente S3 no inicializado")
                return False
            
            image_link = transaction.get('image_link')
            if isinstance(image_link, str):
                image_link = image_link.strip().strip('"').strip("'")
                if image_link.lower() in ('none', 'null'):
                    image_link = ''
            if not image_link or pd.isna(image_link):
                return False
            
            # Registrar el intento
            image_record = {
                'reference_name': ref_row['reference_name'],
                'clase_name': ref_row['producto_a_entrenar'],
                'label_id': ref_row['label_considerar'],
                'transaction_id': transaction.get('transaction_id'),
                'transaction_start_time': transaction.get('transaction_start_time'),
                'device_id': transaction.get('device_id'),
                's3_key': '',
                'local_path': '',
                'download_status': 'failed',
                'error_reason': '',
                'dataset_folder': str(dest_folder.parent.parent.name),
                'motivo_descarga': 'reference',
                'valor_motivo': ref_row['reference_name']
            }
            
            # Limpiar la ruta de S3
            s3_key = str(image_link).strip().lstrip('/')
            image_record['s3_key'] = s3_key
            
            # Buscar archivos de imagen en S3
            image_objects = self._list_s3_objects(s3_key)
            if not image_objects:
                image_record['error_reason'] = 'Imagen no encontrada en S3'
                self.downloaded_images_log.append(image_record)
                return False
            
            s3_object_key = image_objects[0]
            
            # Crear nombre de archivo con formato especificado
            filename = self._create_dataset_filename(transaction, s3_object_key)
            dest_path = dest_folder / filename
            image_record['local_path'] = str(dest_path)
            
            if dry_run:
                self.logger.debug(f"[DRY-RUN] {filename} -> {dest_folder}")
                image_record['download_status'] = 'dry_run'
                image_record['error_reason'] = ''
                self.downloaded_images_log.append(image_record)
                return True
            
            # Crear directorio si no existe
            dest_folder.mkdir(parents=True, exist_ok=True)
            
            # Descargar imagen
            if self._download_from_s3(s3_object_key, dest_path):
                image_record['download_status'] = 'success'
                image_record['error_reason'] = ''
                self.downloaded_images_log.append(image_record)
                return True
            else:
                image_record['error_reason'] = 'Error descargando desde S3'
                self.downloaded_images_log.append(image_record)
                return False
                
        except Exception as e:
            if 'image_record' in locals():
                image_record['error_reason'] = str(e)
                if image_record not in self.downloaded_images_log:
                    self.downloaded_images_log.append(image_record)
            self.logger.debug(f"Error descargando imagen: {e}")
            return False

    def _list_s3_objects(self, prefix: str) -> List[str]:
        """Listar objetos en S3 (o MinIO) que correspondan a un archivo o prefijo de carpeta.

        Mejora respecto a versión previa:
         - Si prefix ya parece un archivo (tiene extensión válida) se intenta head_object directo.
         - Aumenta MaxKeys a 100 para tener más opciones.
         - Logging de diagnóstico cuando no se encuentran imágenes.
        """
        try:
            if any(prefix.lower().endswith(ext) for ext in self.image_extensions):
                try:
                    self.s3_client.head_object(Bucket=self.s3_bucket, Key=prefix)
                    return [prefix]
                except Exception:
                    # Continuar a listado por si es un folder mal formateado
                    pass
            response = self.s3_client.list_objects_v2(
                Bucket=self.s3_bucket,
                Prefix=prefix,
                MaxKeys=100
            )
            objects: List[str] = []
            if 'Contents' in response:
                for obj in response['Contents']:
                    key = obj['Key']
                    if any(key.lower().endswith(ext) for ext in self.image_extensions):
                        objects.append(key)
            if not objects:
                self.logger.debug(f"Sin imágenes para prefijo '{prefix}' en bucket {self.s3_bucket}")
            return objects
        except Exception as e:
            self.logger.debug(f"Error listando objetos S3 con prefijo {prefix}: {e}")
            return []

    def _create_dataset_filename(self, transaction: pd.Series, s3_key: str) -> str:
        """Crear nombre de archivo: <fecha_adquisicion>_<transaction_id>_<original_filename>"""
        # 1) Fecha de adquisición
        ts = transaction.get('transaction_start_time')
        ts_str = None
        try:
            if pd.notna(ts):
                if hasattr(ts, 'to_pydatetime'):
                    dt = ts.to_pydatetime()
                elif isinstance(ts, (pd.Timestamp,)):
                    dt = ts.to_pydatetime()
                else:
                    # Intentar parsear desde string
                    dt = pd.to_datetime(ts, errors='coerce')
                    if pd.isna(dt):
                        dt = None
                if dt is not None:
                    ts_str = dt.strftime('%Y%m%d%H%M%S')
        except Exception:
            ts_str = None
        if not ts_str:
            ts_str = '00000000000000'

        # 2) Transaction ID
        transaction_id = str(transaction.get('transaction_id', 'unknown'))

        # 3) Nombre original
        original_filename = Path(s3_key).name if s3_key else ''
        # Asegurar extensión válida
        ext = Path(original_filename).suffix.lower()
        if not ext or ext not in self.image_extensions:
            # Si el original no tiene extensión válida, forzar .jpg
            original_stem = Path(original_filename).stem or 'image'
            original_filename = f"{original_stem}.jpg"

        return f"{ts_str}_{transaction_id}_{original_filename}"

    def _download_from_s3(self, s3_key: str, dest_path: Path) -> bool:
        """Descargar archivo desde S3"""
        try:
            self.s3_client.download_file(self.s3_bucket, s3_key, dest_path)
            return True
        except Exception as e:
            self.logger.debug(f"Error descargando {s3_key}: {e}")
            return False

    def _process_additional_devices(self, excel_path: str, df_elastic: pd.DataFrame, 
                                  df_references: pd.DataFrame, dataset_folder: Path, 
                                  dry_run: bool = False,
                                  fast: bool = False,
                                  limit_devices: Optional[int] = None,
                                  images_per_device: Optional[int] = None,
                                  no_filter_used: bool = False) -> int:
        """Procesar devices adicionales que tienen num_imagenes_dataset_adicionales > 0"""
        try:
            t0 = time.time()
            # Cargar devices desde Excel
            df_devices = pd.read_excel(excel_path, sheet_name='Devices')
            # Filtrar por incluir_dataset == si/sí si existe
            if 'incluir_dataset' in df_devices.columns:
                mask_inc = (
                    df_devices['incluir_dataset']
                    .astype(str)
                    .str.lower()
                    .str.strip()
                    .isin(['si'])
                )
                before = len(df_devices)
                df_devices = df_devices[mask_inc].copy()
                self.logger.info(f"🔧 Devices marcados para incluir_dataset=si: {len(df_devices)} (de {before})")
            # Recargar referencias desde Excel (garantizar incluir_dataset == 'sí')
            try:
                df_refs_full = pd.read_excel(excel_path, sheet_name='References')
                mask_si = df_refs_full.get('incluir_dataset', pd.Series(dtype=str)).astype(str).str.lower().str.strip().isin(['si'])
                df_refs_incluir = df_refs_full[mask_si].copy()
            except Exception:
                df_refs_incluir = df_references.copy()
            
            # Filtrar devices que necesitan imágenes adicionales
            devices_adicionales = df_devices[
                (df_devices['num_imagenes_dataset_adicionales'] > 0) & 
                (df_devices['num_imagenes_dataset_adicionales'].notna())
            ]
            # Modo rápido: limitar número de devices
            if fast and limit_devices is not None:
                if limit_devices > 0 and len(devices_adicionales) > limit_devices:
                    devices_adicionales = devices_adicionales.sample(n=limit_devices, random_state=42)
            
            if devices_adicionales.empty:
                self.logger.info("No hay devices con imágenes adicionales configuradas")
                return 0
            
            self.logger.info(f"🔧 Procesando {len(devices_adicionales)} devices adicionales")
            
            # Obtener todas las transacciones ya utilizadas (de referencias + devices anteriores)
            used_transactions = set()
            for record in self.downloaded_images_log:
                used_transactions.add(record['transaction_id'])
            
            total_additional_images = 0
            t_filter_total = 0.0
            t_download_total = 0.0
            
            # Procesar cada device adicional
            for _, device_row in devices_adicionales.iterrows():
                # Mantener string para filtros y asegurar entero para el log
                device_id_raw = device_row['device_id']
                device_id_str = str(device_id_raw).strip()
                try:
                    device_id_int = int(float(device_id_raw)) if pd.notna(device_id_raw) else None
                except Exception:
                    device_id_int = None
                if fast and images_per_device is not None and images_per_device > 0:
                    num_images = int(images_per_device)
                else:
                    num_images = int(device_row['num_imagenes_dataset_adicionales'])
                
                self.logger.info(f"📱 Procesando device {device_id_str}: {num_images} imágenes")
                
                # Filtrar datos de Elasticsearch para este device y productos del dataset
                t_f0 = time.time()
                dataset_reference_codes = df_refs_incluir['reference_code'].astype(str).str.strip().unique()
                device_elastic_data = df_elastic[
                    (df_elastic['device_id'].astype(str).str.strip() == device_id_str) &
                    (df_elastic['selected_reference_code'].astype(str).str.strip().isin(dataset_reference_codes))
                ]
                # Excluir transacciones ya usadas en datasets previos
                if not no_filter_used:
                    device_elastic_data = self._filter_used_transactions(device_elastic_data)
                
                if device_elastic_data.empty:
                    try:
                        total_device_rows = len(df_elastic[df_elastic['device_id'].astype(str).str.strip() == device_id_str])
                    except Exception:
                        total_device_rows = -1
                    self.logger.warning(
                        f"No hay datos para device {device_id_str} con productos del dataset (filas totales del device: {total_device_rows})"
                    )
                    continue
                
                # Filtrar transacciones ya utilizadas
                available_data = device_elastic_data[
                    ~device_elastic_data['transaction_id'].isin(used_transactions)
                ]
                t_filter_total += (time.time() - t_f0)
                
                if available_data.empty:
                    self.logger.warning(f"No hay transacciones disponibles para device {device_id_str}")
                    continue
                
                # Seleccionar imágenes aleatoriamente
                sample_size = min(num_images, len(available_data))
                selected_data = available_data.sample(n=sample_size, random_state=42)
                
                self.logger.info(f"📥 Descargando {len(selected_data)} imágenes para device {device_id_str}")
                
                # Descargar imágenes para cada transacción seleccionada
                device_images = 0
                t_d0 = time.time()
                for _, row in selected_data.iterrows():
                    # Encontrar el producto correspondiente para determinar la clase
                    product_match = df_refs_incluir[df_refs_incluir['reference_code'].astype(str) == str(row['selected_reference_code'])]
                    if product_match.empty:
                        continue
                    
                    class_name = product_match.iloc[0]['producto_a_entrenar']
                    label_id = int(product_match.iloc[0]['label_considerar'])
                    ref_name = product_match.iloc[0]['reference_name']
                    
                    # Crear carpeta de clase si no existe
                    class_folder = dataset_folder / class_name / str(label_id)
                    if not dry_run:
                        class_folder.mkdir(parents=True, exist_ok=True)
                    
                    # Descargar imagen
                    ok, abs_path = self._download_image_from_s3(row, class_folder, dry_run)
                    if ok:
                        device_images += 1
                        used_transactions.add(row['transaction_id'])
                        
                        # Registrar descarga con información del device (formato consistente)
                        self.downloaded_images_log.append({
                            'reference_name': ref_name,
                            'clase_name': class_name,
                            'label_id': label_id,
                            'transaction_id': row['transaction_id'],
                            'transaction_start_time': row.get('transaction_start_time', ''),
                            'device_id': device_id_int if device_id_int is not None else device_id_str,
                            's3_key': row.get('image_link', ''),
                            'local_path': abs_path,
                            'download_status': 'success',
                            'error_reason': '',
                            'dataset_folder': dataset_folder.name,
                            'motivo_descarga': 'device',
                            'valor_motivo': device_id_int if device_id_int is not None else device_id_str
                        })
                t_download_total += (time.time() - t_d0)
                
                total_additional_images += device_images
                self.logger.info(f"✅ Device {device_id_str}: {device_images} imágenes descargadas")
            
            self.logger.info(f"⏱️  Tiempo filtrar devices: {t_filter_total:.2f}s | tiempo descargar devices: {t_download_total:.2f}s")
            self.logger.info(f"🎯 Total de imágenes adicionales descargadas: {total_additional_images} (tiempo total devices: {time.time()-t0:.2f}s)")
            return total_additional_images
            
        except Exception as e:
            self.logger.error(f"Error procesando devices adicionales: {e}")
            return 0

    def _download_image_from_s3(self, row: pd.Series, dest_folder: Path, dry_run: bool) -> tuple[bool, str]:
        """Descargar una imagen desde S3 para devices adicionales. Devuelve (ok, ruta_absoluta|"")."""
        try:
            if self.s3_client is None:
                self.logger.error("Cliente S3 no inicializado")
                return (False, "")
            
            image_link = row.get('image_link')
            if not image_link or pd.isna(image_link):
                return (False, "")
            
            # Limpiar la ruta de S3
            s3_key = str(image_link).strip().lstrip('/')
            
            # Si el enlace apunta a un prefijo/carpeta, listar y elegir el primero válido
            s3_object_key = s3_key
            if not any(s3_key.lower().endswith(ext) for ext in self.image_extensions):
                image_objects = self._list_s3_objects(s3_key)
                if not image_objects:
                    return (False, "")
                s3_object_key = image_objects[0]
            
            # Crear nombre de archivo con formato solicitado
            filename = self._create_dataset_filename(row, s3_object_key)
            dest_path = dest_folder / filename
            
            if dry_run:
                self.logger.debug(f"[DRY-RUN] {filename} -> {dest_folder}")
                return (True, str(dest_path))
            
            # Crear directorio si no existe
            dest_folder.mkdir(parents=True, exist_ok=True)
            
            # Descargar imagen
            ok = self._download_from_s3(s3_object_key, dest_path)
            return (ok, str(dest_path) if ok else "")
            
        except Exception as e:
            self.logger.debug(f"Error descargando imagen para device: {e}")
            return (False, "")

    def _save_dataset_log(self, excel_path: str, dataset_folder: Path):
        """Guardar registro de dataset creado en Excel con gestión de archivo abierto"""
        try:
            if not self.downloaded_images_log:
                self.logger.info("No hay registros de dataset para guardar")
                return
            
            df_log = pd.DataFrame(self.downloaded_images_log)
            # Añadir columnas de origen si faltan
            if 'origen' not in df_log.columns:
                df_log['origen'] = 'crear_dataset'
            if 'dataset_original' not in df_log.columns:
                df_log['dataset_original'] = ''
            # Añadir relative_folder (ruta relativa dentro del dataset)
            if 'local_path' in df_log.columns and 'dataset_folder' in df_log.columns:
                def _rel(row):
                    try:
                        lp = Path(str(row['local_path']))
                        ds = str(row['dataset_folder'])
                        parts = list(lp.parts)
                        if ds in parts:
                            i = parts.index(ds)
                            rel_parts = parts[i+1:]
                            return '/'.join(rel_parts[:-1]) if len(rel_parts) > 1 else '.'
                    except Exception:
                        return ''
                    return ''
                df_log['relative_folder'] = df_log.apply(_rel, axis=1)
            else:
                df_log['relative_folder'] = ''
            # Normalizar tipos: device_id y valor_motivo (device) como enteros cuando aplique
            if 'device_id' in df_log.columns:
                # Convertir device_id a entero cuando sea numérico, preservando vacíos como <NA>
                df_log['device_id'] = pd.to_numeric(df_log['device_id'], errors='coerce').astype('Int64')
            if 'valor_motivo' in df_log.columns:
                mask_device = df_log.get('motivo_descarga', pd.Series(index=df_log.index)).eq('device')
                df_log.loc[mask_device, 'valor_motivo'] = pd.to_numeric(
                    df_log.loc[mask_device, 'valor_motivo'], errors='coerce'
                ).astype('Int64')
            sheet_name = 'datasets_creados'
            
            print(f"\n💾 Guardando registro del dataset en Excel...")
            print(f"📊 Total de registros: {len(df_log)}")

            manager = ReportExcelManager(excel_path, interactive=True, max_retries=5, retry_wait=1.5)
            try:
                manager.wait_until_unlocked()
            except ExcelLockedError as e:
                self.logger.error(f"No se pudo guardar registro (Excel abierto): {e}")
                print("⚠ Cierra el Excel y vuelve a ejecutar para registrar datasets_creados.")
                return

            def _modify(wb):
                from openpyxl.utils.dataframe import dataframe_to_rows
                if sheet_name in wb.sheetnames:
                    existing_df = pd.read_excel(excel_path, sheet_name=sheet_name)
                    if 'origen' not in existing_df.columns:
                        existing_df['origen'] = 'crear_dataset'
                    if 'dataset_original' not in existing_df.columns:
                        existing_df['dataset_original'] = ''
                    combined_df = pd.concat([existing_df, df_log], ignore_index=True)
                    self.logger.info(f"📝 Añadiendo {len(df_log)} registros a {len(existing_df)} existentes")
                    del wb[sheet_name]
                else:
                    combined_df = df_log
                    self.logger.info(f"📝 Creando nueva hoja con {len(df_log)} registros")
                pos = None
                if 'Devices' in wb.sheetnames:
                    pos = wb.sheetnames.index('Devices') + 1
                sheet = wb.create_sheet(sheet_name, index=pos if pos is not None else None)
                for r in dataframe_to_rows(combined_df, index=False, header=True):
                    sheet.append(r)
                sheet._final_count = len(combined_df)

            manager.save_with_validations(_modify)
            try:
                final_count = len(pd.read_excel(excel_path, sheet_name=sheet_name))
            except Exception:
                final_count = 'desconocido'
            print(f"✅ Registro guardado en hoja '{sheet_name}'")
            print(f"📊 Total de registros en la hoja: {final_count}")
            return
                        
        except Exception as e:
            self.logger.error(f"Error guardando registro de dataset: {e}")
            print("❌ No se guardó el registro (no se genera backup alternativo).")
